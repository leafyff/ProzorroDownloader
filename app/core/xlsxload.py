"""Читання вивантаженої книги Excel у нормалізовані таблиці.

Аналітика свідомо працює з файлом, а не з базою: користувач може взяти
будь-яке з попередніх вивантажень (у тому числі зроблене іншою версією
програми або на іншій машині) і порівняти періоди між собою.

Ціна такого рішення — колонки доводиться зіставляти за їхніми назвами, тож
зіставлення навмисно поблажливе: книга, у якій бракує аркуша чи колонки,
має читатися без помилок, просто з вужчим набором полів. Чого бракує —
видно у ``Dataset.missing``, і аналіз про це чесно пише.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Callable, Sequence

from openpyxl import load_workbook

from ..paths import long_path
from .xlsxfast import FastWorkbook

Row = dict[str, Any]

#: Аркуш зі змістом книги даних не містить — його пропускаємо.
SKIP_SHEETS = ("зміст", "contents")


# --- нормалізація значень -------------------------------------------------

#: Пробіли у звітах бувають нерозривними та вузькими — і для порівняння назв,
#: і для розбору чисел усі вони мають вважатися звичайним пробілом.
_SPACES = re.compile("[\\s   ]+")
#: Підрядки, через які клітинку таки треба чистити регулярним виразом.
DIRTY_SPACES = ("  ", "\t", "\n", "\r", "\xa0", "\u2007", "\u202f")
_APOSTROPHES = str.maketrans("’`ʼ´", "''''")


def norm_key(text: Any) -> str:
    """Ключ для порівняння назв аркушів і колонок."""
    s = str(text or "").translate(_APOSTROPHES).lower()
    return _SPACES.sub(" ", s).strip(" .:;,")


def as_text(value: Any) -> str:
    """Клітинка як охайний рядок.

    Переважна більшість значень уже охайна, тому спершу перевіряємо це
    кількома швидкими пошуками підрядка: регулярний вираз на кожній із
    півтора мільйона клітинок коштує більше, ніж увесь інший розбір разом.
    """
    if type(value) is str:
        text = value.strip()
        if not text:
            return ""
        for messy in DIRTY_SPACES:
            if messy in text:
                return _SPACES.sub(" ", text)
        return text
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return _SPACES.sub(" ", str(value)).strip()


def as_num(value: Any) -> float | None:
    """Число з клітинки. Текст із пробілами й комами теж приймається."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, (datetime, date, time)):
        # Дата в грошовій колонці — це помилка даних, а не число: інакше з
        # «2026-08-17» вийшло б 20 260 817.
        return None
    text = _SPACES.sub("", str(value)).replace("−", "-")
    if not text:
        return None
    # Десятковий роздільник у вивантаженнях трапляється і крапкою, і комою.
    if "," in text and "." in text:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    text = re.sub(r"[^\d.\-+eE]", "", text)
    try:
        return float(text)
    except ValueError:
        return None


def as_int(value: Any) -> int:
    number = as_num(value)
    return int(number) if number is not None else 0


def as_date(value: Any) -> str:
    """Дата у вигляді ``YYYY-MM-DD``; невпізнане — порожній рядок."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return ""
    text = str(value).strip()
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    match = re.search(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", text)
    if match:
        day, month, year = match.groups()
        return f"{year}-{month}-{day}"
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return match.group(0) if match else ""


def as_edrpou(value: Any) -> str:
    """Код ЄДРПОУ/РНОКПП рядком.

    Excel радо перетворює ``04527520`` на число й губить провідний нуль, тож
    короткі числові коди добиваємо нулями до восьми знаків.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        digits = str(int(value))
        return digits.zfill(8) if len(digits) < 8 else digits
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return ""
    return digits.zfill(8) if len(digits) < 8 else digits


# --- опис таблиць ---------------------------------------------------------

#: ``логічна таблиця → (назви аркушів, {поле: назви колонок})``.
#: Перша назва колонки — очікувана, решта — сумісність зі старими книгами.
TABLES: dict[str, dict] = {
    "tenders": {
        "sheets": ("закупівлі", "тендери"),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "date": ("дата оприлюднення", "дата"),
            "title": ("предмет закупівлі", "предмет"),
            "description": ("опис",),
            "cpv_list": ("коди дк021", "коди дк 021", "дк021"),
            "status": ("статус",),
            "method": ("процедура",),
            "category": ("категорія предмета",),
            "value": ("очікувана вартість",),
            "currency": ("валюта",),
            "vat": ("пдв включено",),
            "buyer": ("замовник",),
            "buyer_edrpou": ("єдрпоу замовника",),
            "region": ("регіон",),
            "locality": ("населений пункт",),
            "n_lots": ("лотів",),
            "n_bids": ("пропозицій",),
            "n_docs": ("файлів",),
            "tender_start": ("початок подання",),
            "tender_end": ("кінець подання",),
            "modified": ("остання зміна",),
            "url": ("посилання",),
        },
    },
    "lots": {
        "sheets": ("лоти",),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "title": ("лот",),
            "description": ("опис лоту",),
            "status": ("статус лоту",),
            "value": ("очікувана вартість лоту",),
            "currency": ("валюта",),
        },
    },
    "items": {
        "sheets": ("номенклатура", "позиції"),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "date": ("дата",),
            "description": ("позиція", "номенклатура"),
            "cpv": ("код дк021", "код дк 021"),
            "cpv_name": ("назва коду",),
            "quantity": ("кількість",),
            "unit": ("одиниця",),
            "lot_value": ("очікувана вартість лоту",),
            "unit_price": ("ціна за одиницю",),
            "buyer": ("замовник",),
            "buyer_edrpou": ("єдрпоу замовника",),
            "region": ("регіон",),
        },
    },
    "bids": {
        "sheets": ("пропозиції",),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "date": ("дата закупівлі",),
            "name": ("учасник",),
            "edrpou": ("єдрпоу учасника",),
            "region": ("регіон учасника",),
            "amount": ("сума пропозиції",),
            "currency": ("валюта",),
            "status": ("статус пропозиції",),
            "submitted": ("дата подання",),
        },
    },
    "awards": {
        "sheets": ("переможці",),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "date": ("дата закупівлі",),
            "name": ("переможець",),
            "edrpou": ("єдрпоу переможця",),
            "amount": ("сума",),
            "currency": ("валюта",),
            "status": ("статус рішення",),
            "decided": ("дата рішення",),
        },
    },
    "contracts": {
        "sheets": ("договори",),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "contract_id": ("номер договору",),
            "name": ("постачальник",),
            "edrpou": ("єдрпоу постачальника",),
            "amount": ("сума договору",),
            "currency": ("валюта",),
            "status": ("статус договору",),
            "signed": ("дата підписання",),
            "buyer": ("замовник",),
            "buyer_edrpou": ("єдрпоу замовника",),
            "region": ("регіон",),
            "cpv_list": ("коди дк021", "коди дк 021"),
        },
    },
    "documents": {
        # Реєстр документів — найбільший аркуш книги: на десять тисяч
        # закупівель це понад сто тисяч рядків. Аналізу з нього потрібні лише
        # власник файлу, розділ, назва та формат — за назвою визначається, що
        # це за документ. Решту колонок (тип, дата, розмір, стан, шлях,
        # посилання) свідомо не читаємо: це майже мільйон клітинок, які нікуди
        # далі не йдуть. Якщо колись знадобляться — досить дописати рядок.
        "sheets": ("документи",),
        "columns": {
            "tender_id": ("номер закупівлі",),
            "scope": ("розділ",),
            "owner_edrpou": ("єдрпоу власника",),
            "title": ("назва файлу",),
            "format": ("формат",),
        },
    },
    "products": {
        "sheets": ("товари каталогу", "товари"),
        "columns": {
            "title": ("товар",),
            "brand": ("бренд",),
            "category": ("категорія",),
            "cpv": ("код дк021",),
            "cpv_name": ("назва коду",),
            "barcode": ("штрихкод",),
            "description": ("опис",),
            "price_low": ("ціна, нижній квартиль",),
            "price_high": ("ціна, верхній квартиль",),
            "currency": ("валюта",),
            "vat": ("пдв включено",),
            "price_date": ("дата ціни",),
            "marketplace": ("майданчик",),
            "vendor": ("постачальник",),
            "status": ("статус",),
            "n_images": ("фото, шт",),
            "n_specs": ("характеристик, шт",),
            "description_len": ("довжина опису",),
            "url": ("посилання на картку",),
        },
    },
    "specs": {
        "sheets": ("характеристики товарів", "характеристики"),
        "columns": {
            "title": ("товар",),
            "brand": ("бренд",),
            "category": ("категорія",),
            "barcode": ("штрихкод",),
            "name": ("характеристика",),
            "value": ("значення",),
            "number": ("число",),
            "unit": ("одиниця",),
        },
    },
}

#: Як приводити значення кожного поля.
CASTS: dict[str, Callable[[Any], Any]] = {
    "value": as_num, "amount": as_num, "lot_value": as_num, "unit_price": as_num,
    "quantity": as_num, "price_low": as_num, "price_high": as_num, "number": as_num,
    "size": as_int, "n_lots": as_int, "n_bids": as_int, "n_docs": as_int,
    "n_images": as_int, "n_specs": as_int, "description_len": as_int,
    "date": as_date, "signed": as_date, "submitted": as_date, "decided": as_date,
    "published": as_date, "modified": as_date, "price_date": as_date,
    "tender_start": as_date, "tender_end": as_date,
    "edrpou": as_edrpou, "buyer_edrpou": as_edrpou, "owner_edrpou": as_edrpou,
}

#: Людські назви логічних таблиць — для повідомлень про те, чого бракує.
TABLE_LABELS = {
    "tenders": "Закупівлі", "lots": "Лоти", "items": "Номенклатура",
    "bids": "Пропозиції", "awards": "Переможці", "contracts": "Договори",
    "documents": "Документи", "products": "Товари каталогу",
    "specs": "Характеристики товарів",
}


@dataclass
class Dataset:
    """Прочитана книга: кожна логічна таблиця — список словників."""

    path: Path
    tenders: list[Row] = field(default_factory=list)
    lots: list[Row] = field(default_factory=list)
    items: list[Row] = field(default_factory=list)
    bids: list[Row] = field(default_factory=list)
    awards: list[Row] = field(default_factory=list)
    contracts: list[Row] = field(default_factory=list)
    documents: list[Row] = field(default_factory=list)
    products: list[Row] = field(default_factory=list)
    specs: list[Row] = field(default_factory=list)

    #: Логічні таблиці, яких у книзі не знайшлося.
    missing: list[str] = field(default_factory=list)
    #: Назви аркушів як вони є у файлі.
    sheet_names: list[str] = field(default_factory=list)
    #: Колонки, яких забракло: ``таблиця → перелік полів``.
    missing_columns: dict[str, list[str]] = field(default_factory=dict)

    def table(self, name: str) -> list[Row]:
        return getattr(self, name, [])

    @property
    def total_rows(self) -> int:
        return sum(len(self.table(name)) for name in TABLES)

    def counts(self) -> list[tuple[str, int]]:
        return [(TABLE_LABELS[name], len(self.table(name))) for name in TABLES]


def _match_columns(headers: list[str], spec: dict[str, tuple[str, ...]]) -> dict[str, int]:
    """Зіставляє назви колонок аркуша з полями таблиці.

    Спершу точні збіги (щоб «Замовник» не забрав колонку «ЄДРПОУ замовника»),
    потім часткові — і лише серед колонок, які ще ніхто не зайняв.
    """
    normalized = [norm_key(h) for h in headers]
    used: set[int] = set()
    found: dict[str, int] = {}

    for field_name, names in spec.items():
        for wanted in names:
            for i, header in enumerate(normalized):
                if i not in used and header == wanted:
                    found[field_name] = i
                    used.add(i)
                    break
            if field_name in found:
                break

    for field_name, names in spec.items():
        if field_name in found:
            continue
        for wanted in names:
            for i, header in enumerate(normalized):
                if i not in used and header and (
                        header.startswith(wanted) or wanted in header):
                    found[field_name] = i
                    used.add(i)
                    break
            if field_name in found:
                break
    return found


def _read_rows(rows_of: Callable, title: str,
               spec: dict[str, tuple[str, ...]]) -> tuple[list[Row], list[str]]:
    """Перетворює аркуш на список словників за описом таблиці.

    Заголовок віддається читачеві назад: знаючи, які колонки насправді
    потрібні, той може не розбирати решту. На реєстрі документів це більшість
    аркуша.
    """
    state: dict[str, Any] = {}

    def on_header(raw: Sequence[Any]) -> set[int]:
        headers = [as_text(cell) for cell in raw]
        columns = _match_columns(headers, spec)
        state["missing"] = [name for name in spec if name not in columns]
        # Порядок полів і потрібні перетворення визначаємо один раз, а не на
        # кожному з двохсот тисяч рядків.
        state["plan"] = [(name, columns.get(name), CASTS.get(name, as_text))
                         for name in spec]
        return set(columns.values())

    rows: list[Row] = []
    append = rows.append
    for number, raw in enumerate(rows_of(title, on_header)):
        if number == 0:
            if "plan" not in state:         # читач не покликав нас сам
                on_header(raw or ())
            continue
        if not raw:
            continue
        width = len(raw)
        row: Row = {}
        empty = True
        for field_name, index, cast in state["plan"]:
            value = raw[index] if index is not None and index < width else None
            if value is not None and value != "":
                empty = False
            row[field_name] = cast(value)
        if not empty:
            append(row)
    if "plan" not in state:                 # аркуш порожній
        return [], list(spec)
    return rows, state["missing"]


def _fast_streams(path: Path):
    book = FastWorkbook(long_path(path))
    return book.names, book.rows, book.close


def _openpyxl_streams(path: Path):
    book = load_workbook(long_path(path), read_only=True, data_only=True)
    sheets = {ws.title: ws for ws in book.worksheets}

    def rows(name: str, on_header: Callable | None = None):
        """Ті самі рядки, що й у швидкого читача.

        ``openpyxl`` розбирає книгу цілком і звузити читання не вміє, тому
        відповідь ``on_header`` тут просто ігнорується — сигнатура однакова
        лише щоб виклик згори не розрізняв читачів.
        """
        first = True
        for raw in sheets[name].iter_rows(values_only=True):
            if first:
                first = False
                if on_header is not None:
                    on_header(raw)
            yield raw

    return list(sheets), rows, book.close


def load(path: Path | str,
         on_progress: Callable[[str, int, int], None] | None = None) -> Dataset:
    """Читає книгу у ``Dataset``. Аркуші, яких немає, просто пропускаються.

    Спершу працює власний потоковий читач: на великих книгах він у рази
    швидший за ``openpyxl``. Якщо книга виявиться незвичною й розбір
    зірветься — та сама книга читається ``openpyxl``, який розуміє геть усі
    закутки формату. Помилку показуємо лише тоді, коли не впорався і він.
    """
    path = Path(path)
    try:
        return _load_with(path, _fast_streams, on_progress)
    except Exception:
        return _load_with(path, _openpyxl_streams, on_progress)


def _load_with(path: Path, streams: Callable, on_progress) -> Dataset:
    data = Dataset(path=path)
    names, rows_of, close = streams(path)
    try:
        data.sheet_names = list(names)
        by_name: dict[str, str] = {}
        for title in names:
            by_name.setdefault(norm_key(title), title)
        total = len(TABLES)
        for done, (name, spec) in enumerate(TABLES.items(), start=1):
            if on_progress:
                on_progress(f"Читаємо аркуш «{TABLE_LABELS[name]}»", done, total)
            title = next((by_name[key] for key in spec["sheets"] if key in by_name), None)
            if title is None:
                # Пошук за частковим збігом: назву аркуша могли трохи змінити.
                title = next((value for key, value in by_name.items()
                              if key not in SKIP_SHEETS
                              and any(key.startswith(w) for w in spec["sheets"])), None)
            if title is None:
                data.missing.append(TABLE_LABELS[name])
                continue
            rows, missing_columns = _read_rows(rows_of, title, spec["columns"])
            setattr(data, name, rows)
            if missing_columns:
                data.missing_columns[TABLE_LABELS[name]] = missing_columns
    finally:
        close()
    return data


def find_workbooks(folder: Path | str, limit: int = 200) -> list[Path]:
    """Книги ``.xlsx`` у теці завантажень — свіжіші вгорі.

    Дивимось лише у сам каталог, а не вглиб: вивантаження лягають саме сюди,
    а в підтеках лежать документи закупівель — серед них бувають десятки тисяч
    файлів, і обхід усього дерева підвішував би сторінку на кожному відкритті.
    Файл із будь-якого іншого місця можна взяти кнопкою «Огляд…».
    """
    folder = Path(folder)
    if not folder.is_dir():
        return []
    found = [item for item in folder.glob("*.xlsx")
             if not item.name.startswith("~$")]      # тимчасовий файл Excel
    found.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return found[:limit]

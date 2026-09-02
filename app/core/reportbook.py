"""Книга звіту аналітики: сторінка застосунку — аркуш Excel.

Раніше вивантаження було пласким: кожна таблиця ставала окремим аркушем, а
плитки, висновки й графіки лишалися тільки на екрані. Книгу з такого звіту не
можна було ані показати, ані роздрукувати — числа є, а сенсу немає.

Тут книга повторює те, що бачить око у застосунку. Один аркуш — одна вкладка
або один портрет гравця, а всередині той самий порядок: показники, висновки,
графіки, таблиці. Кожен графік стоїть **праворуч від власних даних**: ліворуч
табличка, з якої він побудований, праворуч — сама діаграма. Тому будь-яку
діаграму можна перерахувати очима, а не вірити на слово.

Діаграми тут — рідні діаграми Excel, а не картинки: вони живуть на своїх
даних, тож у книзі їх можна крутити, перефарбовувати й тягнути в презентацію.
Кольори беруться з тієї самої палітри, що й графіки на екрані
(``app.core.report.SERIES_COLORS``), щоб один ряд не був синім у програмі
й зеленим у звіті.
"""
from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart,
)
from openpyxl.chart import Series as XLSeries
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties, Paragraph, ParagraphProperties, RegularTextRun,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ..paths import long_path
from .report import (
    OUTLIER_COLOR, OWN_COLOR, SERIES_COLORS, Block, ChartData, Profile, Report,
    Sheet, is_percent_column,
)

# --- вигляд ---------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PAGE_FONT = Font(size=16, bold=True, color="1F3864")
BLOCK_FONT = Font(size=13, bold=True, color="1F3864")
GROUP_FONT = Font(size=11, bold=True, color="44546A")
NAME_FONT = Font(size=11, bold=True)
HINT_FONT = Font(size=9, italic=True, color="707070")
LINK_FONT = Font(color="0563C1", underline="single")
TILE_FILL = PatternFill("solid", fgColor="EEF2FA")
TABLE_STYLE = TableStyleInfo(name="TableStyleLight8", showRowStripes=True,
                             showColumnStripes=False)
HEADER_ALIGN = Alignment(vertical="center", wrap_text=True)

#: Ширина діаграми в сантиметрах. Ширша за типові 15 см: підписи в нас
#: довгі — назви компаній і торгових марок, — і на вужчій вони обрізаються.
CHART_WIDTH = 20.0
#: Типова висота діаграми; смугова росте від кількості рядків.
CHART_HEIGHT = 9.5
#: Висота рядка Excel типово 15 пунктів ≈ 0,53 см. Потрібна, щоб порахувати,
#: скільки рядків займе діаграма, і не покласти наступну поверх неї.
ROW_CM = 0.53
#: Скільки колонок лишаємо між табличкою даних і діаграмою.
CHART_GAP = 1
#: Ширина колонки на оглядовому аркуші. Даних там немає, тож колонки потрібні
#: лише як сітка, до якої чіпляються діаграми: 9 знаків ≈ 1,8 см, отже
#: двадцятисантиметрова діаграма займає рівно 11 колонок.
OVERVIEW_COL = 9
#: Перша колонка правої діаграми на оглядовому аркуші (11 зайнятих + проміжок).
OVERVIEW_RIGHT = 13
#: Скільки колонок відводимо під сітку оглядового аркуша.
OVERVIEW_COLS = 24
#: Найдовша ширина колонки: далі колонка їде за край екрана.
MAX_WIDTH = 60
#: Скільки перших рядків таблиці міряємо, добираючи ширину колонки. Ширина
#: однаково впирається у ``MAX_WIDTH``, а міряти всі 50 тисяч — це ще один
#: ``str()`` на кожну клітинку книги.
WIDTH_ROWS = 400
#: Скільки точок хмари пишемо в книгу. На екрані їх до 4000, але в книзі це
#: 4000 рядків, що відсунули б наступну таблицю за обрій; викиди лишаються всі.
SCATTER_CAP = 1000

#: Символи, які Excel не пускає в назву аркуша.
_BAD_SHEET = re.compile(r"[\[\]:*?/\\]")

#: Сторінка закупівлі на порталі. Портал відкриває її просто за людським
#: номером — тим самим, що стоїть у таблиці, — тож ані UUID, ані зайвого
#: запиту для посилання не потрібно.
TENDER_URL = "https://prozorro.gov.ua/tender/{}"
#: Людський номер закупівлі: ``UA-2026-08-21-000123-a``. Шаблон навмисно
#: прив'язаний до країв рядка: у «Ключі» зауважень поруч лежать і описи
#: позицій, і кілька номерів через кому — таке посилання назвати не може.
#: Перевірено на всіх 3 966 977 номерах індексу ЦБД: інших форм немає.
TENDER_ID = re.compile(r"^UA-\d{4}-\d{2}-\d{2}-\d{6}-[a-z]$")

# --- вигляд діаграм -------------------------------------------------------
# Excel малює діаграму без жодних вказівок так, як 2007-го: чорна рамка,
# чорна сітка, чорний текст. Числа від цього не змінюються, а от читати
# книгу стає важко, тож кольори тексту й ліній задані тут явно.

#: Колір назви діаграми — той самий, що й у заголовків аркуша.
CHART_TITLE_COLOR = "1F3864"
#: Колір підписів осей, легенди й підписів значень.
CHART_TEXT_COLOR = "404040"
#: Лінії сітки: помітні, але тихіші за самі дані.
CHART_GRID_COLOR = "D9D9D9"
#: Лінії осей — на півтону темніші за сітку.
CHART_AXIS_COLOR = "BFBFBF"
#: Розмір шрифту в сотих пункта, як велить OOXML: 1200 — це 12 пт.
TITLE_SIZE = 1200
LABEL_SIZE = 900
#: Найбільша кількість смуг, за якої підписи значень ще не зливаються.
#: Виміряно на 20-сантиметровій діаграмі: 16 рядків читаються, далі підпис
#: сідає на сусідній.
LABEL_ROWS = 16


def _rgb(color: str) -> str:
    """``#3d7eff`` в ``3D7EFF``: OOXML не знає ані решітки, ані малих літер."""
    return str(color).lstrip("#").upper()


def _sheet_title(name: str, used: set[str]) -> str:
    """Назва аркуша: без заборонених символів, до 31 знака, неповторна.

    Назви компаній довші за межу Excel, тож хвіст відрізається — разом із
    комою чи крапкою, на якій випало обрізати.
    """
    base = _trim(_BAD_SHEET.sub(" ", str(name or "Аркуш"))) or "Аркуш"
    title, n = base, 2
    while title.lower() in used:
        suffix = f" {n}"
        title = _trim(base[:31 - len(suffix)]) + suffix
        n += 1
    used.add(title.lower())
    return title


def _trim(text: str) -> str:
    """Назва аркуша без хвостів. Апостроф скраю Excel не приймає взагалі."""
    return re.sub(r"\s+", " ", text).strip()[:31].strip(" ,.;:-—'`")


def _column_specs(headers: Sequence[str],
                  rows: Sequence[Sequence[Any]]) -> tuple[list[str | None], list[bool]]:
    """Формат кожної колонки й ознака «тут номери закупівель».

    Формат — за назвою, як і в таблиці на екрані. Частка («Частка»,
    «Дисконт», «Розрив»…) — відсоток; решта чисел — розряди, а копійки лише
    там, де вони справді є: у сумі договору на 90 000 два нулі після коми —
    просто шум, а в ціні за одиницю округлення до гривні спотворює
    порівняння.

    Колонка з номерами впізнається за самими значеннями, а не за назвою:
    номер лежить і в «Закупівлі», і в «Ключі» зауважень, а поруч у тому ж
    «Ключі» — описи позицій. За назвою вийшло б або пропустити половину,
    або пообіцяти посилання там, де його нема з чого зробити.

    Один прохід по таблиці, а не по проходу на колонку: у «Зауваженнях до
    даних» п'ять тисяч рядків і одинадцять колонок, і поколонковий обхід
    будував одинадцять копій стовпців у пам'яті заради двох ознак.
    """
    width = len(headers)
    numeric = [False] * width
    whole = [True] * width
    links = [False] * width
    for row in rows:
        for i, value in zip(range(width), row):
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric[i] = True
                if whole[i] and not float(value).is_integer():
                    whole[i] = False
            # Дешева прикмета перед шаблоном: у книзі сотні тисяч клітинок,
            # і майже жодна з них не починається з «UA-».
            elif not links[i] and isinstance(value, str) and value[:3] == "UA-":
                links[i] = TENDER_ID.match(value) is not None
    out: list[str | None] = []
    for i, header in enumerate(headers):
        if not numeric[i]:
            out.append(None)
        elif is_percent_column(header):
            out.append("0.0%")
        else:
            out.append("#,##0" if whole[i] else "#,##0.00")
    return out, links


def _chart_format(chart: ChartData) -> str:
    """Формат клітинок і осі діаграми.

    Відсотки в даних графіків уже помножені на сто (``discount * 100``), тож
    формат має лише дописати знак, а не ділити ще раз — звідси ``0.0"%"``
    замість ``0.0%``.

    Гроші на графіках — це майже завжди суми в мільйонах, де копійки лише
    засмічують вісь. Але той самий формат дістається й ціні за одиницю: там
    «1 235» замість «1 234,56» вже спотворює порівняння, тож дрібні величини
    лишаються з копійками.
    """
    if chart.unit == "%":
        return '0.0"%"'
    if chart.money_axis:
        top = max((abs(v) for s in chart.series for v in s.values
                   if isinstance(v, (int, float))), default=0)
        if 0 < top < 1000:
            return "#,##0.00"
    return "#,##0"


def _series_header(name: str, unit: str) -> str:
    label = name or "Значення"
    if unit and unit.lower() not in label.lower():
        return f"{label}, {unit}"
    return label


def chart_table(chart: ChartData) -> Sheet:
    """Дані діаграми як звичайна таблиця — те, з чого вона намальована.

    Саме ця табличка лягає ліворуч від діаграми в книзі. Точкова хмара — це
    пари ``(x, y)``, і викиди в ній виносяться окремою колонкою: рідна
    діаграма Excel фарбує ряд цілком, тож два кольори — це два ряди.
    """
    if chart.kind == "scatter":
        series = chart.series[0] if chart.series else None
        points = list(series.points) if series else []
        accent = set(series.accent) if series else set()
        chosen = _thinned(len(points), accent)
        rows = [[points[i][0],
                 None if i in accent else points[i][1],
                 points[i][1] if i in accent else None]
                for i in chosen]
        name = chart.y_title or (series.name if series and series.name else "Значення")
        return [chart.x_title or "X", name, "Викиди"], rows

    with_labels = [s for s in chart.series if s.labels]
    if not with_labels:
        return [], []
    labels = with_labels[0].labels
    first = "Період" if chart.kind in ("line", "area") else "Категорія"
    headers = [first] + [_series_header(s.name, chart.unit) for s in chart.series]
    rows = []
    for i, label in enumerate(labels):
        row: list[Any] = [label]
        for series in chart.series:
            row.append(series.values[i] if i < len(series.values) else None)
        rows.append(row)
    return headers, rows


def _thinned(total: int, accent: set[int]) -> list[int]:
    """Індекси точок, які пишемо в книгу: усі викиди плюс проріджена решта."""
    if total <= SCATTER_CAP:
        return list(range(total))
    rest = [i for i in range(total) if i not in accent]
    room = max(SCATTER_CAP - len(accent), 1)
    stride = max(1, len(rest) // room)
    return sorted(accent | set(rest[::stride][:room]))


class _Placed:
    """Діаграма, вже покладена на аркуш разом зі своїми даними.

    Зберігає не малюнок, а **адресу**: аркуш і прямокутник клітинок. Об'єкт
    діаграми openpyxl належить одному аркушу й на другий його не перекласти,
    зате за цією адресою можна побудувати другу таку саму, яка читає ті самі
    клітинки.
    """

    __slots__ = ("chart", "ws", "head", "n_rows", "n_series", "fmt", "section")

    def __init__(self, chart: ChartData, ws: Worksheet, head: int, n_rows: int,
                 n_series: int, fmt: str) -> None:
        self.chart = chart
        self.ws = ws
        self.head = head
        self.n_rows = n_rows
        self.n_series = n_series
        self.fmt = fmt
        self.section = ""

    def rebuild(self):
        """Така сама діаграма на тих самих клітинках — для іншого аркуша."""
        return _excel_chart(self.chart, self.ws, self.head, self.n_rows,
                            self.n_series, self.fmt)


class _Page:
    """Один аркуш книги з курсором рядка й пам'яттю про ширину колонок."""

    def __init__(self, ws: Worksheet, title: str, *subtitles: str) -> None:
        self.ws = ws
        self.row = 1
        self.widths: dict[int, int] = {}
        self.n_tables = 0
        self.n_charts = 0
        #: Куди лягла кожна діаграма аркуша — щоб оглядовий аркуш міг
        #: побудувати таку саму на **тих самих** клітинках, не переписуючи
        #: даних до себе.
        self.placed: list[_Placed] = []
        # Сітка вимкнена навмисно: аркуш читається як звіт, а межі таблиць
        # малює стиль самої таблиці.
        ws.sheet_view.showGridLines = False
        self.text(title, PAGE_FONT)
        for subtitle in subtitles:
            if subtitle:
                self.text(subtitle, HINT_FONT)
        self.skip()

    # --- службове ---------------------------------------------------------

    def _put(self, row: int, col: int, value: Any, *, font: Font | None = None,
             fill: PatternFill | None = None, fmt: str | None = None,
             width: bool = False):
        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        if cell.data_type in ("f", "e"):
            # Назва предмета закупівлі цілком може початися з «=», а поле в
            # книзі — містити «#N/A»: openpyxl зробив би з першого формулу, а
            # з другого — код помилки. Текст із даних лишається текстом.
            cell.data_type = "s"
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        if width:
            size = len(str(value if value is not None else ""))
            self.widths[col] = max(self.widths.get(col, 0), min(MAX_WIDTH, size))
        return cell

    def text(self, value: str, font: Font | None = None) -> None:
        """Рядок тексту на всю ширину.

        Ширину колонки такий рядок не задає: заголовок блока на 90 знаків
        розтягнув би першу колонку так, що таблиці під ним стали б нечитні.
        Порожні сусідні клітинки й так дають тексту розлитися вправо.
        """
        self._put(self.row, 1, value, font=font)
        self.row += 1

    def skip(self, rows: int = 1) -> None:
        self.row += rows

    def finish(self) -> None:
        for col, width in self.widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = \
                min(MAX_WIDTH, max(10, width + 2))

    # --- складники --------------------------------------------------------

    def tiles(self, tiles: Sequence[tuple[str, str]]) -> None:
        """Плитки показників — списком «назва / значення».

        На екрані вони стоять сіткою по чотири, але в книзі ширина колонки
        спільна на весь аркуш: сітка з плиток розсунула б колонки таблиць.
        """
        if not tiles:
            return
        self.text("Показники", GROUP_FONT)
        for label, value in tiles:
            self._put(self.row, 1, label, fill=TILE_FILL, width=True)
            self._put(self.row, 2, value, font=NAME_FONT, fill=TILE_FILL, width=True)
            self.row += 1
        self.skip()

    def notes(self, notes: Sequence[str], title: str = "Висновки",
              bullet: str = "•  ") -> None:
        if not notes:
            return
        self.text(title, GROUP_FONT)
        for note in notes:
            self.text(bullet + str(note))
        self.skip()

    def table(self, name: str, sheet: Sheet | None, book: _BookState) -> None:
        """Повна таблиця звіту — з фільтром і смугастим стилем.

        Таблиця оформлена як «розумна таблиця» Excel, а не просто рядками:
        автофільтр аркуша буває лише один, а таких таблиць на аркуші кілька,
        і кожній потрібні свої стрілки для фільтрування.
        """
        if not sheet:
            return
        headers, rows = sheet
        if not headers or not rows:
            return
        self.text(name, NAME_FONT)
        head = self.row
        names = _unique_headers(headers)
        for col, header in enumerate(names, start=1):
            cell = self._put(head, col, header, font=HEADER_FONT, fill=HEADER_FILL,
                             width=True)
            cell.alignment = HEADER_ALIGN
        self._body(head, names, rows)
        last = head + len(rows)
        ref = f"A{head}:{get_column_letter(len(names))}{last}"
        # Колонки таблиці описуємо самі. Інакше openpyxl перед записом бере
        # ``ws[ref]`` — зріз усього діапазону — лише щоб прочитати рядок
        # заголовків: зайвий обхід усіх 50 тисяч клітинок книги. Разом із
        # колонками доводиться задавати й фільтр: openpyxl створює його в тій
        # самій функції, і без нього таблиця лишається без стрілок відбору.
        table = Table(
            displayName=book.table_name(), ref=ref, autoFilter=AutoFilter(ref=ref),
            tableColumns=[TableColumn(id=i, name=header)
                          for i, header in enumerate(names, start=1)])
        table.tableStyleInfo = TABLE_STYLE
        self.ws.add_table(table)
        self.n_tables += 1
        self.row = last + 2

    def _body(self, head: int, names: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
        """Тіло таблиці. Тут найгарячіший цикл книги — звідси й дослівність.

        На великій вибірці це під пів мільйона клітинок, тож замість
        ``_put`` з іменованими параметрами йде прямий цикл, а ширину колонки
        міряють лише перші рядки: далі вона все одно впирається в межу.

        Номер закупівлі стає посиланням на її сторінку на порталі. Посилання
        справжнє, а не формулою ``HYPERLINK``: тоді в клітинці лишається сам
        номер, і відбір, сортування та зведені таблиці бачать текст, а не
        результат обчислення.
        """
        ws = self.ws
        width = len(names)
        formats, links = _column_specs(names, rows)
        columns = list(zip(range(1, width + 1), formats, links))
        for offset, row in enumerate(rows, start=1):
            line = head + offset
            measure = offset <= WIDTH_ROWS
            for (col, fmt, link), value in zip(columns, row):
                cell = ws.cell(row=line, column=col)
                cell.value = value
                if cell.data_type in ("f", "e"):
                    cell.data_type = "s"
                if fmt is not None:
                    cell.number_format = fmt
                elif link and isinstance(value, str) and TENDER_ID.match(value):
                    cell.hyperlink = TENDER_URL.format(value)
                    cell.font = LINK_FONT
                if measure and value is not None:
                    size = min(MAX_WIDTH, len(str(value)))
                    if size > self.widths.get(col, 0):
                        self.widths[col] = size

    def chart(self, chart: ChartData) -> None:
        """Дані діаграми ліворуч, сама діаграма — праворуч від них."""
        headers, rows = chart_table(chart)
        if not headers or not rows:
            return
        top = self.row
        self.text(chart.title, NAME_FONT)
        if chart.hint:
            self.text(chart.hint, HINT_FONT)
        head = self.row
        for col, header in enumerate(headers, start=1):
            self._put(head, col, header, font=HEADER_FONT, fill=HEADER_FILL, width=True)
        fmt = _chart_format(chart)
        for offset, row in enumerate(rows, start=1):
            self._put(head + offset, 1, row[0], width=True)
            for col, value in enumerate(row[1:], start=2):
                self._put(head + offset, col, value, fmt=fmt)
        drawn = _excel_chart(chart, self.ws, head, len(rows), len(headers) - 1, fmt)
        self.ws.add_chart(drawn,
                          f"{get_column_letter(len(headers) + CHART_GAP + 1)}{top}")
        self.n_charts += 1
        self.placed.append(_Placed(chart, self.ws, head, len(rows),
                                   len(headers) - 1, fmt))
        height = math.ceil(drawn.height / ROW_CM) + 1
        self.row = max(head + len(rows) + 1, top + height) + 1


def _unique_headers(headers: Sequence[str]) -> list[str]:
    """Excel не терпить двох однакових заголовків в одній таблиці."""
    out: list[str] = []
    seen: dict[str, int] = {}
    for header in headers:
        name = str(header or "Колонка")
        low = name.lower()
        seen[low] = seen.get(low, 0) + 1
        out.append(name if seen[low] == 1 else f"{name} {seen[low]}")
    return out


def _excel_chart(chart: ChartData, ws: Worksheet, head: int, n_rows: int,
                 n_series: int, fmt: str):
    """Рідна діаграма Excel на щойно записаних даних."""
    last = head + n_rows
    cats = Reference(ws, min_col=1, min_row=head + 1, max_row=last)
    accents = [set(s.accent) for s in chart.series]

    if chart.kind == "scatter":
        drawn = ScatterChart()
        drawn.x_axis.title = _chart_title(chart.x_title or "X", LABEL_SIZE,
                                          CHART_TEXT_COLOR)
        drawn.y_axis.title = _chart_title(chart.y_title or "Y", LABEL_SIZE,
                                          CHART_TEXT_COLOR)
        for i in range(n_series):
            values = Reference(ws, min_col=2 + i, min_row=head, max_row=last)
            item = XLSeries(values, cats, title_from_data=True)
            color = _rgb(OUTLIER_COLOR if i else SERIES_COLORS[0])
            item.marker = Marker(symbol="circle", size=5)
            item.marker.graphicalProperties = GraphicalProperties(solidFill=color)
            # Без цього Excel з'єднав би точки лінією й хмара стала б клубком.
            item.graphicalProperties.line.noFill = True
            drawn.series.append(item)
    elif chart.kind == "pie":
        drawn = PieChart()
        drawn.add_data(Reference(ws, min_col=2, min_row=head, max_row=last),
                       titles_from_data=True)
        drawn.set_categories(cats)
        # Кожне «показувати» доводиться вимикати поіменно. Сказавши лише
        # ``showPercent``, решту Excel бере на власний розсуд і виводить усе,
        # що знає: «Сума, грн; Машини для обробки даних (апаратна частина);
        # 76 901 586; 65%» — чотири рядки на кожен сектор, які накривали
        # і сусідні підписи, і назву діаграми.
        drawn.dataLabels = DataLabelList(
            showPercent=True, showVal=False, showCatName=False, showSerName=False,
            showLegendKey=False, showBubbleSize=False, showLeaderLines=True,
            dLblPos="bestFit", txPr=_label_text())
        if drawn.series:
            drawn.series[0].data_points = [
                _point(i, SERIES_COLORS[i % len(SERIES_COLORS)]) for i in range(n_rows)]
    else:
        if chart.kind == "line":
            drawn = LineChart()
        elif chart.kind == "area":
            drawn = AreaChart()
        else:
            drawn = BarChart()
            drawn.type = "bar" if chart.kind == "hbar" else "col"
            drawn.grouping = "clustered"
            # Проміжок міряється у відсотках ширини смуги, тож на двох
            # категоріях звичні 60% дають смугу завтовшки з третину діаграми:
            # вона читається як заливка, а не як стовпчик. Висоту при цьому
            # чіпати не можна — під нею стоїть розкладка аркуша.
            drawn.gapWidth = 60 if n_rows > 4 else 150
        drawn.add_data(Reference(ws, min_col=2, max_col=1 + n_series,
                                 min_row=head, max_row=last), titles_from_data=True)
        drawn.set_categories(cats)
        for i, item in enumerate(drawn.series):
            color = _rgb(SERIES_COLORS[i % len(SERIES_COLORS)])
            if chart.kind == "line":
                item.graphicalProperties.line.solidFill = color
                item.graphicalProperties.line.width = 22000
                item.marker = Marker(symbol="circle", size=5)
                item.marker.graphicalProperties = GraphicalProperties(solidFill=color)
                item.smooth = False
            else:
                item.graphicalProperties.solidFill = color
                item.graphicalProperties.line.noFill = True
            # Наші ТОВ виділені тим самим кольором, що й на екрані: у
            # стовпчиках виділення важливіше за колір ряду.
            marks = accents[i] if i < len(accents) else set()
            if marks and chart.kind in ("bar", "hbar", "hist"):
                item.data_points = [_point(idx, OWN_COLOR) for idx in sorted(marks)]

    drawn.title = _chart_title(chart.title)
    drawn.width = CHART_WIDTH
    drawn.height = _chart_height(chart, n_rows)
    # Рамки навколо діаграми немає: аркуш і так без сітки, а чорний
    # прямокутник лише додавав ліній до й без того щільної сторінки.
    drawn.graphical_properties = GraphicalProperties(
        solidFill="FFFFFF", ln=LineProperties(noFill=True))
    if chart.kind != "pie":
        # openpyxl лишає осі «видаленими», якщо їх не чіпати, — на аркуші
        # виходила діаграма без жодного підпису.
        drawn.x_axis.delete = False
        drawn.y_axis.delete = False
        _dress_axis(drawn.x_axis)
        _dress_axis(drawn.y_axis, grid=True)
    if chart.kind not in ("pie", "scatter"):
        # У openpyxl вісь X — завжди категорії, а вісь Y — завжди значення;
        # смугова діаграма лише малює їх боком. Перевірено на живому Excel:
        # формат, покладений на вісь X смугової, дістається підписам категорій.
        drawn.y_axis.numFmt = fmt
        if chart.kind == "hbar":
            # Excel кладе першу категорію донизу, а в нас перший рядок —
            # найбільший, і він має бути згори. Розворот категорій піднімає
            # разом із ними й вісь значень, тож її саму просять перетнути
            # категорії на останній — тобто лишитися внизу.
            #
            # «Перетин на максимумі» має стояти саме на осі значень.
            # Покладений на вісь категорій, він **знебарвлює всю діаграму**:
            # виміряно 02.09.2026 на живому Excel — смуги виходять білі з
            # чорним обведенням, хоча кольори в книзі є (об'єктна модель
            # Excel віддає ту саму заливку 16743997), а ще така діаграма не
            # вивантажується в картинку. Смугових діаграм у звіті більшість
            # (91 зі 181), тож у чорно-білу перетворювалася майже вся книга.
            drawn.x_axis.scaling.orientation = "maxMin"
            drawn.y_axis.crosses = "max"
    if chart.kind in ("bar", "hbar", "hist") and n_series == 1 and n_rows <= LABEL_ROWS:
        # Підпис значення просто на смузі: інакше висоту стовпчика доводиться
        # міряти оком по сітці, а числа й так лежать у таблиці поруч.
        drawn.dataLabels = DataLabelList(
            showVal=True, showCatName=False, showSerName=False, showPercent=False,
            showLegendKey=False, showBubbleSize=False, numFmt=fmt,
            dLblPos="outEnd", txPr=_label_text())
    named = [s for s in chart.series if s.name]
    if len(named) < 2 and chart.kind != "pie":
        drawn.legend = None
    elif drawn.legend is not None:
        drawn.legend.position = "b"
        drawn.legend.overlay = False
        drawn.legend.txPr = _label_text()
    return drawn


def _chart_title(text: str, size: int = TITLE_SIZE,
                 color: str = CHART_TITLE_COLOR) -> Title:
    """Назва діаграми — над малюнком, а не поверх нього.

    ``overlay`` Excel вважає увімкненим, доки в книзі не сказано протилежне,
    тож назва лягала просто на верхню лінію сітки, а на смуговій — ще й на
    підписи осі значень. Саме це й читалося як «текст поверх тексту».

    Так само підписані й осі точкової хмари — там без назви осі видно лише
    два стовпці голих чисел.
    """
    props = CharacterProperties(sz=size, b=True, solidFill=color)
    para = Paragraph(pPr=ParagraphProperties(defRPr=props),
                     r=[RegularTextRun(t=str(text or ""))])
    return Title(tx=Text(rich=RichText(p=[para])), overlay=False)


def _label_text() -> RichText:
    """Спільний вигляд дрібного тексту діаграми: осі, легенда, підписи."""
    props = CharacterProperties(sz=LABEL_SIZE, solidFill=CHART_TEXT_COLOR)
    body = RichTextProperties(vert="horz")
    # ``r=[]`` навмисно: тут описано, **як** виглядає текст, а не який він.
    # Без цього openpyxl кладе всередину порожній рядок тексту, і в книзі
    # з'являється ``<a:r><a:t/></a:r>`` — ні на що не схожий, зате видний.
    return RichText(bodyPr=body,
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=props),
                                 endParaRPr=props, r=[])])


def _dress_axis(axis, *, grid: bool = False) -> None:
    """Вісь у кольорах звіту: тиха лінія, сіра сітка, читабельний підпис."""
    axis.txPr = _label_text()
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=CHART_AXIS_COLOR))
    axis.majorTickMark = "out"
    axis.minorTickMark = "none"
    if grid:
        axis.majorGridlines = ChartLines(
            spPr=GraphicalProperties(ln=LineProperties(solidFill=CHART_GRID_COLOR)))
    else:
        axis.majorGridlines = None


def _point(index: int, color: str) -> DataPoint:
    """Одна пофарбована точка ряду. ``spPr`` — це і є «графічні властивості»."""
    return DataPoint(idx=index, spPr=GraphicalProperties(solidFill=_rgb(color)))


def _chart_height(chart: ChartData, n_rows: int) -> float:
    """Смугова діаграма росте від кількості рядків, решта — стала."""
    if chart.kind == "hbar":
        return min(20.0, max(7.0, 2.2 + 0.62 * n_rows))
    return CHART_HEIGHT


class _BookState:
    """Спільний стан книги: неповторні назви аркушів і таблиць."""

    def __init__(self) -> None:
        self.titles: set[str] = set()
        self._tables = 0

    def table_name(self) -> str:
        """Ім'я «розумної таблиці» — без пробілів і з підкресленням.

        Ім'я таблиці — це іменований діапазон, а він не має права виглядати
        як адреса клітинки. Перевірено: з іменем ``tbl1`` Excel відмовляється
        відкривати книгу взагалі («знайдено нечитабельний вміст»), бо ``TBL``
        — це справжня колонка (13584-та), і ``tbl1`` для нього адреса.
        Підкреслення перед номером робить таку сплутанину неможливою.
        """
        self._tables += 1
        return f"Таблиця_{self._tables}"


def _write_block(page: _Page, block: Block, book: _BookState) -> None:
    """Блок звіту в тому ж порядку, що й на екрані."""
    page.text(block.title, BLOCK_FONT)
    if block.hint:
        page.text(block.hint, HINT_FONT)
    page.skip()
    page.tiles(block.tiles)
    page.notes(block.notes)
    if block.charts:
        page.text("Графіки", GROUP_FONT)
        page.skip()
        for chart in block.charts:
            page.chart(chart)
    if any(sheet and sheet[1] for _name, sheet in block.tables):
        page.text("Таблиці", GROUP_FONT)
        page.skip()
        for name, sheet in block.tables:
            page.table(name, sheet, book)


def _profile_subtitle(profile: Profile) -> str:
    place = f"місце №{profile.rank}" if profile.rank else "поза рейтингом"
    kind = "наше ТОВ" if profile.is_ours else "конкурент"
    return f"ЄДРПОУ {profile.edrpou}  ·  {kind}  ·  {place}"


def _source_line(report: Report) -> str:
    """Звідки й за що звіт — рядком, коли аркуша «Зміст» немає."""
    source = Path(report.source).name if report.source else "—"
    period = f"{report.period[0] or '—'} — {report.period[1] or '—'}"
    return f"{source}  ·  період {period}  ·  складено {report.generated or '—'}"


def _write_profile(wb: Workbook, book: _BookState, profile: Profile,
                   title: str, *subtitles: str) -> _Page:
    """Аркуш одного гравця: його портрет плюс сильні та слабкі сторони."""
    page = _Page(wb.create_sheet(title), profile.name or profile.edrpou,
                 _profile_subtitle(profile), *subtitles)
    if profile.block:
        _write_block(page, profile.block, book)
    page.notes(profile.strengths, "Сильні сторони", "+  ")
    page.notes(profile.weaknesses, "Слабкі сторони", "−  ")
    page.finish()
    return page


def _write_overview(ws: Worksheet, report: Report,
                    placed: Sequence[_Placed]) -> int:
    """Оглядовий аркуш: усі діаграми звіту в одному місці, по дві в ряд.

    Розділів у книзі дев'ять, і щоб побачити картину цілком, доводилося
    клацати по вкладках, тримаючи попередню в голові. Тут вони поруч — у тому
    самому порядку, у якому їх рахує аналіз.

    Даних аркуш **не дублює**: кожна діаграма читає ті самі клітинки на своєму
    рідному аркуші. Тому книга не важчає, числа лишаються в одному місці, а
    підпис під назвою каже, де саме їх шукати. Заразом це знімає найдовшу
    діаграму з дороги: хмара «ціна × кількість» має тисячу рядків даних, і
    переписана сюди вона відсунула б усе наступне за обрій.

    Портрети гравців сюди не входять: у них 137 діаграм на 39 аркушах, і
    оглядовий аркуш перетворився б на другу книгу.
    """
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Усі графіки звіту"
    ws["A1"].font = PAGE_FONT
    ws["A2"] = _source_line(report)
    ws["A2"].font = HINT_FONT
    ws["A3"] = ("Кожна діаграма побудована на даних свого розділу — вони лишилися "
                "на його аркуші, тут лише малюнки. Портрети окремих компаній "
                "мають власні аркуші й сюди не входять.")
    ws["A3"].font = HINT_FONT
    for col in range(1, OVERVIEW_COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = OVERVIEW_COL

    row = 5
    drawn = 0
    for section in dict.fromkeys(item.section for item in placed):
        group = [item for item in placed if item.section == section]
        if not group:
            continue
        cell = ws.cell(row=row, column=1, value=section)
        cell.font = BLOCK_FONT
        sheet = ws.cell(row=row + 1, column=1,
                        value=f"дані — на аркуші «{group[0].ws.title}»")
        sheet.font = HINT_FONT
        row += 2
        # По дві діаграми в ряд; висоту ряду задає вища з пари, інакше нижча
        # накрила б заголовок наступного розділу.
        for i in range(0, len(group), 2):
            pair = group[i:i + 2]
            height = 0.0
            for offset, item in enumerate(pair):
                figure = item.rebuild()
                column = 1 if offset == 0 else OVERVIEW_RIGHT
                ws.add_chart(figure, f"{get_column_letter(column)}{row}")
                height = max(height, figure.height)
                drawn += 1
            row += math.ceil(height / ROW_CM) + 2
        row += 1
    return drawn


def write_report(path: Path, report: Report) -> Path:
    """Записує весь звіт: аркуш під кожну вкладку й кожного гравця.

    Портрети гравців розкладені по аркушу на компанію, а не по аркушу на
    таблицю: двадцять конкурентів дали б сотню аркушів, у яких нічого не
    знайти, а так один аркуш — один портрет, як одна сторінка в застосунку.
    """
    book = _BookState()
    wb = Workbook()
    wb.remove(wb.active)
    contents = wb.create_sheet(_sheet_title("Зміст", book.titles))
    # Аркуш створюємо одразу, щоб він став другим у книзі, а наповнюємо
    # останнім: його діаграми посилаються на клітинки, яких ще немає.
    overview_title = _sheet_title("Графіки", book.titles)
    overview = wb.create_sheet(overview_title, 1)
    entries: list[tuple[str, str, str, int, int]] = []
    placed: list[_Placed] = []

    for section, blocks in report.sections.items():
        title = _sheet_title(section, book.titles)
        page = _Page(wb.create_sheet(title), section)
        for block in blocks:
            _write_block(page, block, book)
        page.finish()
        entries.append(("Розділ", section, title, page.n_tables, page.n_charts))
        for item in page.placed:
            item.section = section
            placed.append(item)

    # Номер у назві аркуша — це місце в переліку, а не в рейтингу ринку:
    # вкладки Excel не сортуються, тож без нього двадцять портретів лежали б
    # у випадковому на вигляд порядку.
    for people, prefix, kind in ((report.ours, "Ми", "Наше ТОВ"),
                                 (report.competitors, "К", "Конкурент")):
        for n, profile in enumerate(people, start=1):
            if not profile.block:
                continue
            title = _sheet_title(f"{prefix}{n} {profile.name or profile.edrpou}",
                                 book.titles)
            page = _write_profile(wb, book, profile, title)
            entries.append((kind, profile.label, title, page.n_tables, page.n_charts))

    # Той самий графік стоїть і в «Підсумку», і у своєму розділі — на екрані
    # це доречно, а поруч на одному аркуші виглядало б помилкою.
    unique: list[_Placed] = []
    seen: set[str] = set()
    for item in placed:
        if item.chart.title not in seen:
            seen.add(item.chart.title)
            unique.append(item)
    n_charts = _write_overview(overview, report, unique)
    entries.insert(0, ("Огляд", "Усі графіки звіту", overview_title, 0, n_charts))

    _write_contents(contents, report, entries)
    return _save(wb, path)


def write_profile_report(path: Path, report: Report, profile: Profile) -> Path:
    """Записує звіт по одній компанії — нашій або конкуренту.

    Книга з одного аркуша: рівно те, що показує сторінка цієї компанії в
    застосунку. Аркуша «Зміст» тут немає — переліковувати нічого, — тож
    звідки взяті дані й за який період, сказано просто під назвою компанії.
    """
    book = _BookState()
    wb = Workbook()
    wb.remove(wb.active)
    title = _sheet_title(profile.name or profile.edrpou, book.titles)
    _write_profile(wb, book, profile, title, _source_line(report))
    return _save(wb, path)


def _save(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(long_path(path))
    return path


def _write_contents(ws: Worksheet, report: Report,
                    entries: Sequence[tuple[str, str, str, int, int]]) -> None:
    """Перший аркуш: звідки дані, за який період і що де лежить."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Звіт аналітики Prozorro"
    ws["A1"].font = PAGE_FONT
    facts = [("Джерело", Path(report.source).name if report.source else "—"),
             ("Період", f"{report.period[0] or '—'} — {report.period[1] or '—'}"),
             ("Складено", report.generated or "—"),
             ("Аркушів", len(entries)),
             ("Таблиць", sum(e[3] for e in entries)),
             ("Графіків", sum(e[4] for e in entries))]
    row = 3
    for label, value in facts:
        ws.cell(row=row, column=1, value=label).font = NAME_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1
    if report.notes:
        row += 1
        ws.cell(row=row, column=1, value="Застереження").font = GROUP_FONT
        row += 1
        for note in report.notes:
            ws.cell(row=row, column=1, value="•  " + str(note))
            row += 1
    row += 1
    head = row
    for col, header in enumerate(["Що це", "Назва", "Аркуш", "Таблиць", "Графіків"],
                                 start=1):
        cell = ws.cell(row=head, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for offset, (kind, name, title, n_tables, n_charts) in enumerate(entries, start=1):
        ws.cell(row=head + offset, column=1, value=kind)
        ws.cell(row=head + offset, column=2, value=name)
        # Посилання формулою, а не гіперпокликанням: так Excel сам знаходить
        # аркуш за назвою, навіть якщо в ній пробіли й крапки. Назва компанії
        # майже завжди в лапках («ТОВ "КОМЕЛ"»), а лапка всередині рядка
        # формули обриває цей рядок — тож подвоюємо її, як велить Excel.
        target = quote_sheetname(title).replace('"', '""')
        label = title.replace('"', '""')
        link = ws.cell(row=head + offset, column=3,
                       value=f'=HYPERLINK("#{target}!A1","{label}")')
        link.font = Font(color="0563C1", underline="single")
        ws.cell(row=head + offset, column=4, value=n_tables).number_format = "#,##0"
        ws.cell(row=head + offset, column=5, value=n_charts).number_format = "#,##0"
    for col, width in ((1, 14), (2, 52), (3, 34), (4, 10), (5, 11)):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=head + 1, column=1)

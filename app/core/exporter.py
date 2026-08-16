"""Вивантаження результатів у XLSX і CSV."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..paths import long_path

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_xlsx(path: Path, sheets: dict[str, tuple[Sequence[str], Sequence[Sequence]]]) -> Path:
    """Записує кілька аркушів: ``{назва: (заголовки, рядки)}``."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(name[:31] or "Аркуш")
        ws.append(list(headers))
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append(list(row))
        # Грошові й лічильні колонки — з розрядами, щоб суми читалися очима.
        for line in ws.iter_rows(min_row=2):
            for cell in line:
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(cell.value, int) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0"
        widths = [len(str(h)) for h in headers]
        for row in rows[:400]:
            for i, value in enumerate(row):
                if i < len(widths):
                    widths[i] = max(widths[i], min(60, len(str(value or ""))))
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, width + 2))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(long_path(path))
    return path


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Sequence]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(long_path(path), "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(headers)
        writer.writerows(rows)
    return path

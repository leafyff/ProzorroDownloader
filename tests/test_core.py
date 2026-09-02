"""Швидкі перевірки ядра без звернень до мережі.

Запуск:  .venv\\Scripts\\python -m tests.test_core
"""
from __future__ import annotations

import io
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

if __package__ in (None, ""):                       # прямий запуск файлу
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import SearchPreset, Settings
from app.core.classifiers import expand_prefixes, label_for, significant_prefix
from app.core.db import Database
from app.core.downloader import (
    FileDownloader, document_extension, is_signature, normalize_extensions, tender_folder,
)
from app.core.extract import iter_documents, latest_versions, parse_tender
from app.core.market import (
    MARKET_BASE, PRODUCT_URL, SEARCH_URL, _unwrap, class_codes, parse_product,
)
from app.core.pipeline import Pipeline
from app.core.resolver import IndexBuilder, tender_id_of
from app.paths import safe_name

FAILED: list[str] = []


def check(name: str, condition: bool, detail: object = "") -> None:
    print(f"  {'✓' if condition else '✗'} {name}" + (f" — {detail}" if detail else ""))
    if not condition:
        FAILED.append(name)


def test_classifiers() -> None:
    print("=== класифікатор ДК021 ===")
    check("порожній список", expand_prefixes([]) == [])
    # 17 — єдиний двоцифровий розділ, якого в ДК021 немає.
    check("невідомий префікс", expand_prefixes(["17"]) == [])
    check("рідкісний код", expand_prefixes(["99999"]) == ["99999999-9"])
    check("сміття на вході", expand_prefixes(["", "  ", None]) == [])
    check("розділ 30 розгортається", len(expand_prefixes(["30"])) == 401)
    check("значущий префікс коду", significant_prefix("30213300-8") == "302133")
    check("значущий префікс розділу", significant_prefix("30000000-9") == "30")
    check("підпис відомої гілки", label_for("777").startswith("77700000-7 —"))
    check("підпис невідомої гілки", label_for("17") == "17")


def test_safe_names() -> None:
    print("\n=== безпечні імена файлів ===")
    check("заборонені символи", safe_name('a<b>c:d"e/f\\g|h?i*j') == "a_b_c_d_e_f_g_h_i_j")
    check("зарезервоване ім'я", safe_name("CON.txt") == "_CON.txt")
    check("порожнє ім'я", safe_name("") == "_")
    # Windows мовчки відкидає крапки в кінці імені — прибираємо їх самі, щоб
    # записаний у базі шлях збігався з реальним файлом на диску.
    check("крапки по краях", safe_name("  ...назва...  ") == "назва")
    long_name = safe_name("x" * 300 + ".pdf", max_len=40)
    check("обрізання зі збереженням розширення",
          len(long_name) <= 40 and long_name.endswith(".pdf"))


def test_extract() -> None:
    print("\n=== розбір картки закупівлі ===")
    check("порожня картка не падає", parse_tender({})["row"]["uuid"] == "")
    check("картка без документів", parse_tender({"id": "x"})["docs"] == [])

    tender = {
        "id": "t1", "tenderID": "UA-2026-01-01-000001-a",
        "documents": [
            {"url": "https://x/1", "title": "a.pdf", "datePublished": "2026-01-01", "id": "d1"},
            {"url": "https://x/2", "title": "a.pdf", "datePublished": "2026-01-02", "id": "d1"},
        ],
        "bids": [{
            "id": "b1", "tenderers": [{"name": "ТОВ", "identifier": {"id": "123"}}],
            "documents": [{"url": "https://x/3", "title": "b.pdf"}],
            "financialDocuments": [{"url": "https://x/4", "title": "c.pdf"}],
        }],
        "awards": [{
            "id": "a1", "suppliers": [{"name": "ТОВ", "identifier": {"id": "123"}}],
            "documents": [{"url": "https://x/5", "title": "d.pdf"}],
        }],
    }
    docs = list(iter_documents(tender))
    check("знайдено всі файли", len(docs) == 5, len(docs))
    check("області визначено",
          sorted({d["scope"] for d in docs}) == ["award", "bid", "tender"])
    check("власника пропозиції визначено",
          {d["owner_edrpou"] for d in docs if d["scope"] == "bid"} == {"123"})
    check("контейнери правильні",
          sorted({d["container"] for d in docs}) == [
              "awards[0].documents", "bids[0].documents",
              "bids[0].financialDocuments", "documents"])
    check("ключі документів унікальні", len({d[0] for d in parse_tender(tender)["docs"]}) == 5)
    # Дві версії одного doc_id у тому самому контейнері згортаються в останню.
    versions = latest_versions([dict(d) for d in docs])
    check("лишається остання версія", len(versions) == 4, len(versions))

    # У картці закупівлі постачальник указаний лише в рішенні про переможця,
    # а договір посилається на нього через awardID.
    with_contract = {
        "id": "t2", "tenderID": "UA-2026-01-01-000002-a",
        "awards": [{"id": "aw1", "status": "active",
                    "suppliers": [{"name": "ТОВ Постачальник",
                                   "identifier": {"id": "12345678"}}]}],
        "contracts": [{"id": "c1", "contractID": "UA-2026-01-01-000002-a-a1",
                       "awardID": "aw1", "status": "active",
                       "value": {"amount": 1000, "currency": "UAH"}}],
    }
    parsed = parse_tender(with_contract)
    contract = parsed["contracts"][0]
    check("ЄДРПОУ постачальника підтягнуто з нагороди", contract[8] == "12345678", contract[8])
    check("назву постачальника теж підтягнуто", contract[7] == "ТОВ Постачальник")
    # Якщо постачальник таки вказаний у самому договорі — беремо саме його.
    with_contract["contracts"][0]["suppliers"] = [
        {"name": "ТОВ Інший", "identifier": {"id": "87654321"}}]
    check("власний постачальник договору має пріоритет",
          parse_tender(with_contract)["contracts"][0][8] == "87654321")


def test_index_chunks() -> None:
    print("\n=== межі відрізків індексу ===")
    chunks = IndexBuilder.month_chunks(date(2025, 1, 15), date(2025, 3, 10))
    check("три місячні відрізки", len(chunks) == 3)
    check("початок не раніше заданого", chunks[0][0] == date(2025, 1, 15))
    check("кінець не пізніше заданого", chunks[-1][1] == date(2025, 3, 10))
    check("один день", IndexBuilder.month_chunks(date(2025, 5, 5), date(2025, 5, 5))
          == [(date(2025, 5, 5), date(2025, 5, 5))])
    check("номер закупівлі з номера договору",
          tender_id_of("UA-2026-08-13-007779-a-a1") == "UA-2026-08-13-007779-a")
    check("номер без суфікса не змінюється",
          tender_id_of("UA-2026-08-13-007779-a") == "UA-2026-08-13-007779-a")


def test_batch_files() -> None:
    """Файли запуску мають лишатися суто ASCII.

    Кирилиця в .bat ламає cmd.exe: після ``chcp`` він губить позицію у файлі
    й починає виконувати уривки рядків як команди. Помилка виглядає як
    «'середовище' is not recognized», а програма просто не стартує.
    """
    print("\n=== файли запуску ===")
    root = Path(__file__).resolve().parent.parent
    for name in ("run.bat", "run_console.bat"):
        path = root / name
        if not path.exists():
            check(f"{name} існує", False)
            continue
        data = path.read_bytes()
        bad = [b for b in data if b > 127]
        check(f"{name}: лише ASCII", not bad, f"не-ASCII байтів: {len(bad)}")
        check(f"{name}: без BOM", not data.startswith(b"\xef\xbb\xbf"))
        text = data.decode("ascii", errors="replace")
        check(f"{name}: запускає застосунок", "-m app.main" in text)


def test_market() -> None:
    print("\n=== картки товарів е-каталогу ===")
    # Каталог класифікує товари на рівні класу, закупівлі — на рівні коду.
    check("коди підіймаються до класу",
          class_codes(["30213300-8", "30213100-6"]) == ["30210000-4"],
          class_codes(["30213300-8", "30213100-6"]))
    check("порожній перелік", class_codes([]) == [])
    check("сміття не ламає", class_codes(["", "12", None]) == [])

    card = {
        "id": "p1", "status": "active",
        "title": "Ноутбук Acer TravelMate TMP215-55",
        "description": "Ноутбук 15.6\" FHD IPS",
        "categoryTitle": "Ноутбуки",
        "owner": "uss.gov.ua",
        "images": ["https://x/1.png", "https://x/2.png"],
        "identifier": {"id": "4711474580108", "scheme": "EAN-13"},
        "classification": {"scheme": "ДК021",
                           "description": "ДК 021:2015: 30210000-4 — Машини для обробки даних"},
        "latestPrice": {"lowerQuartile": 38235, "upperQuartile": 42900,
                        "currency": "UAH", "valueAddedTaxIncluded": False,
                        "date": "2026-02-10T10:00:04+02:00"},
        "requirementResponses": [
            {"requirement": "Бренд", "values": ["ACER"], "unitName": ""},
            {"requirement": "Діагональ екрану", "values": [15.6], "unitName": "дюйм"},
            {"requirement": "Мова розкладки", "values": ["англійська", "українська"],
             "unitName": ""},
            {"requirement": "Сенсорний екран", "values": [False], "unitName": ""},
        ],
    }
    row, specs = parse_product(card)
    check("бренд витягнуто з характеристик", row["brand"] == "ACER", row["brand"])
    check("штрихкод", row["barcode"] == "4711474580108")
    check("код ДК021 з опису класифікації", row["cpv"] == "30210000-4", row["cpv"])
    check("ціновий діапазон", (row["price_low"], row["price_high"]) == (38235.0, 42900.0))
    check("фото полічені", row["n_images"] == 2)
    check("характеристики полічені", row["n_specs"] == 4)
    check("довжина опису", row["description_len"] == len(card["description"]))
    check("характеристик у довгому форматі", len(specs) == 4)
    numeric = next(s for s in specs if s[1] == "Діагональ екрану")
    check("числове значення окремо", numeric[3] == 15.6 and numeric[4] == "дюйм")
    multi = next(s for s in specs if s[1] == "Мова розкладки")
    check("кілька значень через кому", multi[2] == "англійська, українська", multi[2])
    check("логічне значення як текст",
          next(s for s in specs if s[1] == "Сенсорний екран")[2] == "False")
    # Стисла картка з пошуку доповнює детальну, якщо в тій чогось бракує.
    row2, _ = parse_product({"id": "p2"}, {"title": "З пошуку", "status": "inactive"})
    check("дані з пошуку підставляються",
          row2["title"] == "З пошуку" and row2["status"] == "inactive")


def test_market_api() -> None:
    """Картки читаються з market-api.prozorro.gov.ua, а не з проксі порталу."""
    print("\n=== офіційний API каталогу ===")
    check("картки з market-api", PRODUCT_URL.startswith(MARKET_BASE), PRODUCT_URL)
    check("портал лишається тільки для пошуку",
          "/api/products/" not in SEARCH_URL and SEARCH_URL.endswith("/search/products"),
          SEARCH_URL)

    # market-api загортає об'єкт у {"data": ...}, проксі віддавав його плоским.
    check("конверт знімається", _unwrap({"data": {"id": "p3"}}) == {"id": "p3"})
    check("плоска відповідь не псується", _unwrap({"id": "p4"})["id"] == "p4")
    check("сміття не ламає", _unwrap(None) == {} and _unwrap([1]) == {})

    # Відповідь у форматі каталогу: value замість values, unit замість unitName,
    # фото — об'єктами з відносним шляхом, опису немає взагалі.
    card = {
        "id": "0c4fd49c61c34727b2c25916ca615361",
        "status": "active",
        "title": "Інтерактивна панель SWEDX 85\"",
        "description": None,
        "relatedCategory": "32320000-005501-32348248",
        "owner": "uss.gov.ua",
        "dateCreated": "2026-08-25T18:07:53.763762+03:00",
        "dateModified": "2026-08-25T18:07:53.763762+03:00",
        "classification": {"scheme": "ДК021", "id": "32320000-2",
                           "description": "Телевізійне й аудіовізуальне обладнання"},
        "images": [{"url": "/static/images/04/4c/ad/ab58d04.jpeg",
                    "title": "panel.jpg", "hash": "md5:5102bb"}],
        "requirementResponses": [
            {"requirement": "Бренд", "values": ["SWEDX"],
             "classification": {"scheme": "ESPD211", "id": "CRITERION.OTHER"}},
            {"requirement": "Діагональ екрану", "value": 85.0,
             "unit": {"code": "INH", "name": "дюйм"}},
            {"requirement": "Вага", "value": 2.4, "unit": {"code": "KGM", "name": "кг"}},
            {"requirement": "Сенсорний екран", "value": True},
            {"requirement": "Порожня", "value": None},
        ],
    }
    row, specs = parse_product(card, category_title="Інтерактивні панелі НУШ")

    check("бренд із values", row["brand"] == "SWEDX", row["brand"])
    check("код ДК021 з поля id", row["cpv"] == "32320000-2", row["cpv"])
    check("відносне фото стало абсолютним",
          row["images"] == MARKET_BASE + "/static/images/04/4c/ad/ab58d04.jpeg",
          row["images"])
    check("фото полічені", row["n_images"] == 1)
    check("порожній опис не ламає базу",
          row["description"] == "" and row["description_len"] == 0)
    check("назва категорії підставлена",
          row["category"] == "Інтерактивні панелі НУШ", row["category"])
    check("посилання складається саме", row["url"] == MARKET_BASE + "/api/products/" + card["id"],
          row["url"])
    check("майданчик", row["marketplace"] == "uss.gov.ua")
    check("дати обрізані до дня",
          row["date_created"] == "2026-08-25" and row["date_modified"] == "2026-08-25")
    check("ціни немає — і не вигадуємо",
          row["price_low"] is None and row["price_high"] is None)

    single = next(s for s in specs if s[1] == "Діагональ екрану")
    check("ціле число без зайвого «.0»", single[2] == "85" and single[3] == 85.0, single)
    check("одиниця з unit.name", single[4] == "дюйм", single[4])
    fraction = next(s for s in specs if s[1] == "Вага")
    check("дробове значення лишається дробовим",
          fraction[2] == "2.4" and fraction[3] == 2.4, fraction)
    boolean = next(s for s in specs if s[1] == "Сенсорний екран")
    check("логічне value як текст", boolean[2] == "True", boolean[2])
    empty = next(s for s in specs if s[1] == "Порожня")
    check("value=null дає порожній рядок", empty[2] == "" and empty[3] is None, empty)
    check("усі характеристики збережені", len(specs) == 5 and row["n_specs"] == 5)

    # Значення, які лягають у базу, мають бути скалярами SQLite.
    bad = [k for k, v in row.items() if not isinstance(v, (str, int, float, type(None)))]
    check("рядок придатний для SQLite", not bad, bad)


def test_product_groups() -> None:
    """Той самий товар під різними назвами має бути одним рядком.

    Пастки тут не вигадані: усі приклади взяті з живого вивантаження, а порядок
    і поле ``beats`` у ``data/product_types.json`` підібрані саме під них.
    """
    print("\n=== об'єднання назв товарів ===")
    from app.core.products import by_catalog, catalog, head_noun, normalize, product_group

    check("довідник завантажився", len(catalog()) > 40, len(catalog()))

    # 1. Те, з чого все почалося: одна річ під п'ятьма назвами.
    mouse = ["Миші", "Мишка", "Миша", "Миша безпровідна Gembird", "Комп’ютерна миша",
             "Мишка комп'ютерна", "маніпулятор миша"]
    groups = {product_group(t) for t in mouse}
    check("миша в усіх написаннях — один товар", groups == {"Миша"}, groups)

    # 2. Апостроф буває чотирьох накреслень, і це різні символи.
    pc = {product_group(t) for t in ("Комп'ютер", "Комп’ютер", "Компʼютер",
                                     "Персональні комп'ютери", "ПК")}
    check("апостроф не розриває товар на кілька", pc == {"Персональний комп'ютер"}, pc)

    # 3. Число й відмінок.
    check("однина і множина разом",
          product_group("Монітори") == product_group("Монітор") == "Монітор")
    check("абревіатура і повна назва разом",
          product_group("БФП лазерний") == product_group("Багатофункціональні пристрої")
          == "Багатофункціональний пристрій (БФП)")

    # 4. Уточнення після товару не робить його іншим товаром, але й не краде
    #    аксесуар: українською головне слово стоїть першим.
    pairs = [
        ("Килимок для мишки Gembird", "Килимок для миші"),
        ("Кабель монітора 1.8м VGA", "Кабель, перехідник"),
        ("Кулер для процесора Deepcool", "Кулер, вентилятор, термопаста"),
        ("Клавіатура для ноутбука ASUS", "Клавіатура"),
        ("Сумка для ноутбука CASE LOGIC", "Сумка або чохол"),
        ("Блок живлення ноутбука", "Блок живлення"),
    ]
    wrong = [(text, product_group(text)) for text, want in pairs
             if product_group(text) != want]
    check("уточнення не перетягує товар до себе", not wrong, wrong)

    # 5. Виняток `beats`: опис починається із загального слова, а товар вужчий.
    check("SSD не тоне в загальному «накопичувачі»",
          product_group("Накопичувач SSD USB Type-C 2TB Kingston") == "Накопичувач SSD")
    check("сканер штрих-коду не зливається зі сканером",
          product_group("Сканер штрих-коду Symbol LS2208") == "Сканер штрих-коду")
    check("«принтер БФП» — це БФП",
          product_group("принтер БФП HP Smart Tank") == "Багатофункціональний пристрій (БФП)")

    # 6. Код ДК021 і назва класу — це заголовок, а не товар.
    check("товар беруть із дужок після коду класу",
          product_group("30230000-0 Комп’ютерне обладнання (монітор)") == "Монітор",
          product_group("30230000-0 Комп’ютерне обладнання (монітор)"))
    check("сам лише код класу товару не називає",
          normalize("30230000-0 Комп’ютерне обладнання") == "",
          normalize("30230000-0 Комп’ютерне обладнання"))

    # 7. Друкарські помилки й латиниця замість кирилиці.
    check("латинська H у «Haкопичувач»",
          product_group("Haкопичувач SSD M.2 2280") == "Накопичувач SSD")
    check("описка в назві", product_group("Монтітор") == "Монітор")

    # 8. Різні речі не мають злипатися.
    check("системний блок — це не блок живлення",
          product_group("Системний блок НЕО") != product_group("Блок живлення"),
          (product_group("Системний блок НЕО"), product_group("Блок живлення")))
    check("ноутбук — це не сумка для ноутбука",
          product_group("Ноутбук ASUS") != product_group("Сумка для ноутбука"))

    # 9. Запасне правило для того, чого немає в довіднику.
    check("незнайоме слово зводиться до основи",
          head_noun("Зрощувачі оптичні") == head_noun("Зрощувач"),
          (head_noun("Зрощувачі оптичні"), head_noun("Зрощувач")))
    check("прикметник не стає товаром", head_noun("Твердотільний накопичувач") == "накопичувач",
          head_noun("Твердотільний накопичувач"))
    check("порожній опис — порожня група", product_group("   ") == "")
    check("довідник мовчить про сторонній товар", by_catalog("Бетонозмішувач") == "")


def test_two_product_tables(tmp: Path) -> None:
    """У звіті дві таблиці товарів: об'єднана й «як у документах»."""
    print("\n=== дві таблиці товарів ===")
    from app.core import insight
    from app.core.xlsxload import Dataset

    def item(tender: str, text: str, qty: float = 1.0):
        return {"tender_id": tender, "lot_id": "", "description": text, "cpv": "30213300-8",
                "cpv_name": "", "quantity": qty, "unit": "шт", "unit_price": 100.0,
                "buyer": "Замовник", "buyer_edrpou": "04527520", "date": "2026-03-01"}

    tenders = [{"tender_id": f"UA-{i}", "date": "2026-03-01", "title": "Мишки",
                "description": "", "cpv_list": "30213300-8", "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 1000.0,
                "currency": "UAH", "vat": "так", "buyer": "Замовник",
                "buyer_edrpou": "04527520", "region": "Київська область", "locality": "Київ",
                "n_lots": 0, "n_bids": 1, "n_docs": 0, "tender_start": "", "tender_end": "",
                "modified": "", "url": ""} for i in range(4)]
    items = [item("UA-0", "Миші"), item("UA-1", "Мишка"), item("UA-2", "Миша дротова"),
             item("UA-3", "Монітори")]
    data = Dataset(path=tmp / "товари.xlsx", tenders=tenders, items=items)
    report = insight.analyse(data)

    tables = dict(report.sections["Товари і ТМ"][0].tables)
    check("є обидві таблиці",
          "Товари ринку (об'єднані)" in tables
          and "Товари ринку (як у документах)" in tables, list(tables))
    grouped, raw = tables["Товари ринку (об'єднані)"], tables["Товари ринку (як у документах)"]
    check("об'єднана коротша", len(grouped[1]) == 2 and len(raw[1]) == 4,
          (len(grouped[1]), len(raw[1])))
    check("миші зведені в один рядок",
          any(r[0] == "Миша" and r[1] == 3 for r in grouped[1]), grouped[1])
    check("видно, що саме злилося",
          "Варіанти назви" in grouped[0]
          and any("Мишка" in str(r[-1]) for r in grouped[1]), grouped[1])
    # Головний інваріант: об'єднання не має нічого губити дорогою.
    check("сума позицій однакова в обох таблицях",
          sum(r[1] for r in grouped[1]) == sum(r[1] for r in raw[1]) == 4,
          (sum(r[1] for r in grouped[1]), sum(r[1] for r in raw[1])))
    check("кількість товару теж збігається",
          sum(r[3] for r in grouped[1]) == sum(r[3] for r in raw[1]))
    check("у примітці сказано про об'єднання",
          any("формулювань" in n for n in report.sections["Товари і ТМ"][0].notes),
          report.sections["Товари і ТМ"][0].notes)


def test_gone_cards(tmp: Path) -> None:
    """Картки, яких каталог більше не віддає, не мають заливати журнал.

    Каталог відповідає ``404 Category not found`` на товар, чия категорія
    вилучена з довідника, — це стан його даних, а не збій. Таких карток бувають
    сотні, тож поштучно вони не потрібні: потрібен підсумок. А от справжні збої
    навпаки — мають лишитися видимими.
    """
    import requests

    from app.core.pipeline import MARKET_WARN_LIMIT, _card_gone

    print("\n=== картки, яких каталог більше не віддає ===")

    def http_error(code: int) -> requests.HTTPError:
        response = requests.Response()
        response.status_code = code
        return requests.HTTPError(f"HTTP {code}", response=response)

    check("404 — картки немає", _card_gone(http_error(404)))
    check("503 — це збій, а не відсутність", not _card_gone(http_error(503)))
    check("помилка без відповіді — теж збій", not _card_gone(ValueError("розрив зв'язку")))

    logs: list[tuple[str, str]] = []
    preset = SearchPreset(cpv_codes=["30213300-8"], cpv_prefixes=[],
                          date_from="2026-01-01", date_to="2026-12-31")
    pipe = Pipeline(Settings(), preset, Database(tmp / "gone.db"),
                    on_log=lambda level, msg: logs.append((level, msg)))

    class FakeMarket:
        """Каталог, у якому половина карток застаріла, а одна просто впала."""

        def search(self, *, cpv=None, text="", max_pages=500):
            rows = [{"id": f"{i:04d}", "status": "active",
                     "title": f"Товар {i}", "dateModified": "2026-08-01T00:00:00+03:00"}
                    for i in range(12)]
            yield 1, len(rows), rows

        def product(self, pid: str):
            if pid == "0000":
                raise http_error(500)
            if int(pid) % 2:
                raise http_error(404)
            return {"id": pid, "status": "active", "title": f"Товар {pid}"}

        @staticmethod
        def category_title(_cid: str) -> str:
            return ""

    pipe.market = FakeMarket()
    saved = pipe.collect_products()

    check("справні картки збережені", saved == 5, saved)
    per_card = [m for level, m in logs if m.startswith("Картка ")]
    check("застарілі картки поштучно не перелічуються", len(per_card) <= MARKET_WARN_LIMIT,
          per_card)
    check("справжній збій лишився видимим",
          any("HTTP 500" in m for m in per_card), per_card)
    summary = [m for _l, m in logs if "не віддав" in m]
    check("є підсумок про застарілі картки", summary and "6 карток" in summary[0], summary)
    check("підсумок пояснює причину",
          summary and "категорію" in summary[0] and "не наша помилка" in summary[0],
          summary)
    check("інші збої теж полічені",
          any("з інших причин" in m for _l, m in logs), [m for _l, m in logs][-3:])
    pipe.client.close()


def test_search_quota() -> None:
    """Портал дає рівно 60 запитів на вікно в 60 с — маємо в них вкладатися."""
    import threading
    import time as _t
    from app.core.http import HOST_QUOTAS, WindowLimiter
    from app.core.http import Cancelled as HttpCancelled

    print("\n=== квота пошуку порталу ===")
    limit, window = HOST_QUOTAS["prozorro.gov.ua"]
    check("бюджет не перевищує виміряні 60 на 60 с",
          limit <= 60 and window >= 60.0, (limit, window))
    check("і не занижений даремно", limit >= 50, limit)

    # Мала копія тієї самої квоти, щоб перевірка була швидкою.
    lim = WindowLimiter(5, 0.4)
    t0 = _t.monotonic()
    for _ in range(5):
        lim.acquire()
    check("бюджет вікна витрачається без штучних пауз", _t.monotonic() - t0 < 0.2,
          round(_t.monotonic() - t0, 3))
    lim.acquire()
    check("наступний запит чекає на звільнення вікна", _t.monotonic() - t0 >= 0.4,
          round(_t.monotonic() - t0, 3))

    # 429 попри наш облік — перечікуємо вікно, а не женемо далі.
    lim2 = WindowLimiter(5, 0.4)
    lim2.acquire()
    lim2.penalize()
    t1 = _t.monotonic()
    lim2.acquire()
    check("після відмови сервера перечікуємо вікно", _t.monotonic() - t1 >= 0.35,
          round(_t.monotonic() - t1, 3))

    # Очікування не має ковтати зупинку користувача.
    lim3 = WindowLimiter(1, 30.0)
    lim3.acquire()
    stop = threading.Event()
    stop.set()
    stopped = False
    t2 = _t.monotonic()
    try:
        lim3.acquire(stop)
    except HttpCancelled:
        stopped = True
    check("зупинка перериває очікування одразу",
          stopped and _t.monotonic() - t2 < 1.0, round(_t.monotonic() - t2, 3))


def _plan_search(depths, page_size=20, per_day=40):
    """Фейковий пошук для перевірок плану гортання.

    ``depths`` — скільки записів має запит: ключ ``(коди, процедури)``,
    значення — загальна кількість. Записи розкладаються від найновішого по
    ``per_day`` на добу, тож дата на останній доступній сторінці показує,
    наскільки глибоко цей запит дістає, — саме це й перевіряє проба. Щільність
    узята близькою до справжньої: 10 000 записів по 40 на добу вкривають
    близько восьми місяців, тобто року вони не покривають.
    """
    from app.core.api import SEARCH_MAX_PAGE, SEARCH_MAX_RESULTS
    from datetime import date as _d, timedelta as _td

    class FakeSearch:
        def __init__(self):
            self.queries = []      # тіла всіх запитів
            self.walked = []       # плани, які справді гортались

        def _total(self, body):
            key = (tuple(sorted(body.get("cpv") or [])),
                   tuple(sorted(body.get("proc_type") or [])))
            return depths.get(key, depths.get("*", 0))

        def _row(self, n):
            day = (_d(2026, 12, 31) - _td(days=n // per_day)).isoformat()
            return {"tenderID": f"UA-{day}-{n:06d}",
                    "enquiryPeriod": {"startDate": day + "T00:00:00+03:00"}}

        def query(self, entity, body):
            # Сервер відхиляє порожні поля з HTTP 422 («status must have at
            # least 1 items», «text must be a string»), тож фейк поводиться
            # так само: інакше проба могла б надсилати те, чого API не приймає.
            for key, value in body.items():
                if value == "" or value == []:
                    raise AssertionError(
                        f"порожнє поле {key!r} у тілі запиту — сервер відповів би 422")
            # І так само відхиляє задовгі значення процедури: «The proc_type.0
            # may not be greater than 30 characters».
            for value in body.get("proc_type") or []:
                if len(value) > 30:
                    raise AssertionError(
                        f"proc_type {value!r} — {len(value)} символів, сервер відповів би 422")
            self.queries.append(dict(body))
            total = self._total(body)
            page = int(body.get("page") or 1)
            start_n = (page - 1) * page_size
            rows = [self._row(start_n + i) for i in range(page_size)
                    if start_n + i < min(total, SEARCH_MAX_RESULTS)]
            return {"page": page, "total": total, "data": rows}

        def pages(self, entity="tenders", **kw):
            self.walked.append((tuple(sorted(kw.get("cpv") or [])),
                                tuple(sorted(kw.get("proc_type") or []))))
            total = self._total(kw)
            shown = min(total, SEARCH_MAX_RESULTS)
            for page in range(1, SEARCH_MAX_PAGE + 1):
                start_n = (page - 1) * page_size
                if start_n >= shown:
                    yield page, shown, []
                    return
                rows = [self._row(start_n + i)
                        for i in range(min(page_size, shown - start_n))]
                yield page, shown, rows
                if page * page_size >= shown:
                    return

    return FakeSearch()


def test_search_plan(tmp: Path) -> None:
    """Пошук має з'ясовувати глибину пробою, а не 500 сторінками гортання."""
    from app.core.api import SEARCH_MAX_RESULTS

    print("\n=== план гортання пошуку ===")
    codes = ["30213300-8", "30213100-6", "30231300-0", "30232110-8"]

    def run(depths, date_from="2026-01-01", methods=None):
        preset = SearchPreset(cpv_codes=list(codes), cpv_prefixes=[],
                              methods=list(methods or []),
                              date_from=date_from, date_to="2026-12-31")
        db = Database(tmp / f"plan{abs(hash((str(depths), date_from)))}.db")
        pipe = Pipeline(Settings(), preset, db)
        fake = _plan_search(depths)
        pipe.search = fake
        cards = pipe.discover()
        pipe.client.close()
        db.close()
        return fake, cards

    # 1. Вибірка вміщається у стелю — один план, без уточнень.
    fake, cards = run({"*": 200})
    check("невелику вибірку гортаємо одним запитом", len(fake.walked) == 1, fake.walked)
    check("усі коди пішли разом", len(fake.walked[0][0]) == 4, fake.walked[0])
    check("знайдене зібрано", len(cards) == 200, len(cards))
    check("у тілі запиту немає порожніх полів",
          all(v not in ("", []) for q in fake.queries for v in q.values()),
          [q for q in fake.queries if any(v in ("", []) for v in q.values())][:1])

    # 2. Стеля вичерпана, а період не покрито — коди діляться навпіл.
    #    Головне: рішення про поділ ухвалюється пробою, а не після 500 сторінок.
    fake, _ = run({"*": SEARCH_MAX_RESULTS})
    probed = [tuple(sorted(q.get("cpv") or [])) for q in fake.queries]
    check("проміжний набір із двох кодів справді пробувався",
          any(len(p) == 2 for p in probed), sorted(set(probed))[:4])
    check("уточнені плани не гортаються — лише пробуються",
          all(len(p) in (1, 2, 4) for p in probed), sorted(set(probed))[:4])
    # Проба — це один-два запити на план; гортання стелі коштувало б 500.
    plans = len(set(zip(probed, [tuple(sorted(q.get("proc_type") or []))
                                 for q in fake.queries])))
    check("на план витрачено одиниці запитів, а не сотні",
          len(fake.queries) <= 3 * plans, (len(fake.queries), plans))
    walked_codes = [w[0] for w in fake.walked]
    check("жоден код не загублено",
          {c for w in walked_codes for c in w} == set(codes),
          sorted({c for w in walked_codes for c in w}))

    # 3. Код лишився один, а глибини бракує — беремо розріз за процедурою.
    heavy = {"*": 100, (tuple(sorted(codes)), ()): SEARCH_MAX_RESULTS}
    for half in (tuple(sorted(codes[:2])), tuple(sorted(codes[2:]))):
        heavy[(half, ())] = SEARCH_MAX_RESULTS
    for one in codes:
        heavy[((one,), ())] = SEARCH_MAX_RESULTS
    fake, _ = run(heavy)
    with_proc = [w for w in fake.walked if w[1]]
    check("дійшло до розрізу за процедурою", with_proc, fake.walked[:4])
    check("процедури беруться поштучно",
          all(len(w[1]) == 1 for w in with_proc), with_proc[:3])
    check("задовгі назви процедур не надсилаються",
          all(len(w[1][0]) <= 30 for w in with_proc),
          [w[1][0] for w in with_proc if len(w[1][0]) > 30])

    # 4. Глибше ділити нікуди — попереджаємо, але не зациклюємось.
    fake, _ = run({"*": SEARCH_MAX_RESULTS}, methods=["aboveThresholdUA"])
    check("один код і одна процедура не спричиняють нескінченний поділ",
          len(fake.walked) <= 8, len(fake.walked))


def _buyer_search(total_records=15000, n_buyers=8, page_size=20, per_day=40):
    """Фейк пошуку зі справжньою стелею й фільтром за замовником.

    Уся видача — ``total_records`` записів від найновіших, по ``per_day`` на
    добу; замовник у кожного свій за колом. Сервер, як і живий, показує не
    більше 10 000 записів на запит і не дає сторінки за 500-ту — тож дістати
    давніші можна лише вужчим запитом.
    """
    from app.core.api import SEARCH_MAX_PAGE, SEARCH_MAX_RESULTS
    from datetime import date as _d, timedelta as _td

    class FakeSearch:
        def __init__(self):
            self.queries = []

        @staticmethod
        def _matching(wanted) -> list[int]:
            return [n for n in range(total_records)
                    if not wanted or f"{n % n_buyers:08d}" in wanted]

        @staticmethod
        def _row(n: int) -> dict:
            day = (_d(2026, 12, 31) - _td(days=n // per_day)).isoformat()
            return {"tenderID": f"UA-{day}-{n:06d}",
                    "procuringEntity": {"identifier": {"id": f"{n % n_buyers:08d}"}}}

        def query(self, entity, body):
            self.queries.append(dict(body))
            page = int(body.get("page") or 1)
            if page > SEARCH_MAX_PAGE:
                raise AssertionError("сторінка за 500-ту — сервер відповів би 422")
            rows = self._matching(set(body.get("buyer") or ()))
            shown = min(len(rows), SEARCH_MAX_RESULTS)
            window = rows[(page - 1) * page_size:page * page_size]
            return {"total": shown, "data": [self._row(n) for n in window if n < len(rows)]}

        def pages(self, entity="tenders", **kw):
            self.queries.append(dict(kw))
            rows = self._matching(set(kw.get("buyer") or ()))
            shown = min(len(rows), SEARCH_MAX_RESULTS)
            for page in range(1, SEARCH_MAX_PAGE + 1):
                start = (page - 1) * page_size
                if start >= shown:
                    yield page, shown, []
                    return
                yield page, shown, [self._row(n) for n in rows[start:start + page_size]]
                if page * page_size >= shown:
                    return

    return FakeSearch()


def test_buyer_split(tmp: Path) -> None:
    """Переповнений запит перегортається групами замовників.

    Стеля стоїть не на кількості записів, а на адресації: сторінки за 500-ту
    немає. Тому глибші записи бере лише вужчий запит — а замовник для цього
    годиться, бо в кожної закупівлі він рівно один і список ЄДРПОУ сервер
    об'єднує через АБО. Перелік замовників беремо з карток, які й так
    перегорнуто, тож він не коштує жодного запиту.
    """
    print("\n=== розріз переповненого запиту за замовником ===")
    logs: list[tuple[str, str]] = []
    # Один код і одна процедура — розрізати вже нема чим, окрім замовника.
    preset = SearchPreset(cpv_codes=["30213300-8"], cpv_prefixes=[],
                          methods=["reporting"],
                          date_from="2026-01-01", date_to="2026-12-31")
    pipe = Pipeline(Settings(), preset, Database(tmp / "buyers.db"),
                    on_log=lambda level, msg: logs.append((level, msg)))
    pipe.search = _buyer_search()
    cards = pipe.discover()

    check("узято більше, ніж дозволяє стеля", len(cards) > 10_000, len(cards))
    # Найдавніший запис лежить на позиції 14 999 — далеко за стелею.
    deepest = min(pipe._published(row) for row in cards.values())
    check("дісталися початку періоду", deepest <= "2026-01-02", deepest)
    check("розріз ішов саме за замовником",
          any(q.get("buyer") for q in pipe.search.queries),
          [q.get("buyer") and len(q["buyer"]) for q in pipe.search.queries][:6])
    check("сторінок за 500-ту не просили",
          all(int(q.get("page") or 1) <= 500 for q in pipe.search.queries))
    check("сказано, скільки додав розріз",
          any("Розріз за замовником додав" in m for _l, m in logs),
          [m for _l, m in logs if "замовник" in m][:1])
    check("і що повноти це не гарантує",
          any("повноти це не гарантує" in m for _l, m in logs),
          [m for _l, m in logs if "гарантує" in m][:1])
    check("зайвого попередження про стелю немає",
          not any("упирається у стелю" in m for _l, m in logs),
          [m for _l, m in logs if "стелю" in m][:1])
    pipe.client.close()

    # Коли замовників не видно (стара видача без procuringEntity), розрізати
    # нема чим — і програма має чесно попередити, а не мовчати.
    blind = Pipeline(Settings(), preset, Database(tmp / "blind.db"),
                     on_log=lambda level, msg: logs.append((level, msg)))
    blind.search = _plan_search({"*": 10_000})
    logs.clear()
    blind.discover()
    check("без замовників у видачі — чесне попередження",
          any("не видно достатньо різних ЄДРПОУ" in m for _l, m in logs)
          and any("упирається у стелю" in m for _l, m in logs),
          [m for _l, m in logs][-2:])
    blind.client.close()


def test_unnamed_procedure(tmp: Path) -> None:
    """«Відбір за рамковою угодою» пошук назвати не може — беремо зі стрічки.

    Значення ``proc_type`` довші за 30 символів сервер відхиляє
    (`The proc_type.0 may not be greater than 30 characters`), а
    ``closeFrameworkAgreementSelectionUA`` має 34. Такі закупівлі випадають
    саме тоді, коли запит доводиться різати за процедурою, — і добираються з
    індексу, куди тип процедури тепер потрапляє зі стрічки змін ЦБД.
    """
    from app.core.api import SEARCH_MAX_RESULTS

    print("\n=== процедура, якої пошук не називає ===")
    long_method = "closeFrameworkAgreementSelectionUA"
    check("назва справді довша за межу", len(long_method) > 30, len(long_method))

    codes = ["30213300-8", "30213100-6"]
    logs: list[tuple[str, str]] = []
    preset = SearchPreset(cpv_codes=list(codes), cpv_prefixes=[],
                          date_from="2026-01-01", date_to="2026-12-31")
    db = Database(tmp / "unnamed.db")
    pipe = Pipeline(Settings(), preset, db, on_log=lambda level, msg: logs.append((level, msg)))
    # Кожен запит упирається у стелю, тож поділ дійде до розрізу за процедурою.
    pipe.search = _plan_search({"*": SEARCH_MAX_RESULTS})
    pipe.discover()

    check("довгу процедуру записано як пропущену",
          pipe.unnamed_methods == {long_method}, pipe.unnamed_methods)
    ceiling = [m for level, m in logs if "стелю пошуку" in m]
    check("попередження про стелю каже, докуди дістає запит",
          ceiling and "дістає лише до" in ceiling[0], ceiling[:1])
    check("і не радить зсувати вікно в минуле",
          ceiling and "зсунути вікно в минуле не вийде" in ceiling[0], ceiling[:1])

    # Її ж можна позначити у фільтрі «Тип процедури» — і тоді вона потрапляла
    # прямо в тіло запиту, а сервер відповідав 422 і валив увесь збір.
    picked_preset = SearchPreset(cpv_codes=list(codes), cpv_prefixes=[],
                                 methods=["aboveThresholdUA", long_method],
                                 date_from="2026-01-01", date_to="2026-12-31")
    chosen = Pipeline(Settings(), picked_preset, Database(tmp / "chosen.db"))
    chosen.search = _plan_search({"*": 100})
    chosen.discover()
    check("обрану у фільтрі довгу процедуру в запит не кладемо",
          all(all(len(v) <= 30 for v in (q.get("proc_type") or []))
              for q in chosen.search.queries),
          [q.get("proc_type") for q in chosen.search.queries][:2])
    check("але й не забуваємо про неї", chosen.unnamed_methods == {long_method},
          chosen.unnamed_methods)
    chosen.client.close()

    # Якщо в фільтрі лишити тільки її, шукати нема чого: пошук цю процедуру не
    # знає, тож усі запити були б марними, а закупівлі однаково з індексу.
    only = SearchPreset(cpv_codes=list(codes), cpv_prefixes=[], methods=[long_method],
                        date_from="2026-01-01", date_to="2026-12-31")
    solo = Pipeline(Settings(), only, Database(tmp / "solo.db"),
                    on_log=lambda level, msg: logs.append((level, msg)))
    solo.search = _plan_search({"*": 100})
    logs.clear()
    solo.discover()
    check("із самою лише такою процедурою пошук не витрачає жодного запиту",
          solo.search.queries == [], solo.search.queries[:2])
    check("і про це сказано", any("тільки зі стрічки" in m for _l, m in logs),
          [m for _l, m in logs][:2])
    check("процедуру передано на добір", solo.unnamed_methods == {long_method},
          solo.unnamed_methods)
    solo.client.close()

    # Індекс без типу процедури (зібраний давніше) — програма має сказати прямо.
    db.index_put([("UA-2026-05-05-000001-a", "uuid-old", "", "2026-05-05", "", "", "", "", "")])
    logs.clear()
    check("зі старого індексу добирати нічого", pipe.pickup_unnamed({}) == {}, "")
    check("і про це сказано прямо",
          any("перебудуйте індекс" in m for _l, m in logs), [m for _l, m in logs][:1])

    # Індекс зі стрічки: тип процедури є.
    db.index_put([
        ("UA-2026-03-11-000010-a", "uuid-a", "", "2026-03-11", "", long_method, "", "", ""),
        ("UA-2026-07-22-000011-a", "uuid-b", "", "2026-07-22", "", long_method, "", "", ""),
        ("UA-2026-08-01-000012-a", "uuid-c", "", "2026-08-01", "", "reporting", "", "", ""),
        ("UA-2025-12-31-000013-a", "uuid-d", "", "2025-12-31", "", long_method, "", "", ""),
        ("UA-2027-01-01-000014-a", "uuid-e", "", "2027-01-01", "", long_method, "", "", ""),
    ])
    logs.clear()
    picked = pipe.pickup_unnamed({"UA-2026-03-11-000010-a": "uuid-a"})
    check("узято лише потрібну процедуру за період",
          picked == {"UA-2026-07-22-000011-a": "uuid-b"}, picked)
    check("сказано, скільки знайдено й скільки з них нових",
          any("добираю її зі стрічки" in m for _l, m in logs), [m for _l, m in logs][:1])
    check("неповний індекс не замовчується",
          any("тип процедури відомий лише для" in m for _l, m in logs),
          [m for _l, m in logs][:1])

    rows, with_method = db.index_method_coverage("2026-01-01", "2026-12-31")
    check("покриття індексу рахується за датою з номера", (rows, with_method) == (4, 3),
          (rows, with_method))

    # З неповним індексом добирати нема звідки: у ньому лише знайдене пошуком.
    settings = Settings()
    settings.keep_full_index = False
    lean = Pipeline(settings, preset, db, on_log=lambda level, msg: logs.append((level, msg)))
    lean.unnamed_methods = {long_method}
    logs.clear()
    check("без повного індексу добір не вигадує даних", lean.pickup_unnamed({}) == {}, "")
    check("і причину названо саме цю",
          any("повний індекс" in m for _l, m in logs), [m for _l, m in logs][:1])
    lean.client.close()
    pipe.client.close()
    db.close()


def test_search_progress(tmp: Path) -> None:
    """Смуга має показувати відсоток, а плитки — рости під час пошуку.

    Саме цього бракувало: поступ звітувався лише після завершення пакета, а
    пакет один на весь період, тож вікно до кінця пошуку показувало початковий
    стан. Пізніше смугу зробили «живою» без загальної величини — і зник
    відсоток. Тепер поступ рахується сторінками: скільки пройдено зі скількох
    очікуваних (верхню оцінку дає проба).
    """
    print("\n=== поступ і лічильники під час пошуку ===")

    seen: list[tuple] = []
    ref = {}

    def on_progress(stage, done, total):
        pipe = ref.get("p")
        seen.append((stage, done, total, pipe.result.found if pipe else -1))

    preset = SearchPreset(cpv_codes=["30213300-8", "30213100-6"], cpv_prefixes=[],
                          date_from="2026-01-01", date_to="2026-12-31")
    db = Database(tmp / "progress.db")
    pipe = Pipeline(Settings(), preset, db, on_progress=on_progress)
    ref["p"] = pipe
    pipe.search = _plan_search({"*": 240})
    cards = pipe.discover()
    pipe.client.close()
    db.close()

    check("звіт надходить під час гортання, а не лише в кінці", len(seen) >= 12, len(seen))
    stages = [st for i, (st, *_) in enumerate(seen) if i == 0 or seen[i - 1][0] != st]
    # Спершу з'ясовуємо глибину запитів, і лише потім гортаємо: інакше до
    # кінця планування невідома загальна кількість сторінок, а отже й відсоток.
    check("спершу планування, потім пошук", stages == ["Планую пошук", "Пошук закупівель"],
          stages)
    check("смуга визначена — відсоток можна порахувати",
          all(t > 0 for _s, _d, t, _f in seen), {t for _s, _d, t, _f in seen})
    check("пройдене не перевищує загального",
          all(d <= t for _s, d, t, _f in seen),
          [(d, t) for _s, d, t, _f in seen if d > t][:3])
    check("поступ рухається вперед", seen[-1][1] > seen[0][1], (seen[0][1], seen[-1][1]))
    check("наприкінці смуга повна", seen[-1][1] == seen[-1][2], seen[-1][1:3])

    found_seq = [f for _s, _d, _t, f in seen]
    check("лічильник знайденого росте разом із пошуком",
          found_seq[-1] > found_seq[0] and any(f > 0 for f in found_seq[:3]),
          found_seq[:4])
    check("і сходиться з підсумком", found_seq[-1] == len(cards) == 240,
          (found_seq[-1], len(cards)))



def test_host_policies() -> None:
    """Кожен сервер має свою межу — і хости не мають плутатися між собою."""
    from app.core.http import HOST_QUOTAS, HOST_RATES, HttpClient

    print("\n=== межі за серверами ===")
    check("ім'я хоста з URL",
          HttpClient.host_of("https://market-api.prozorro.gov.ua/api/products/x")
          == "market-api.prozorro.gov.ua")
    check("порт і користувач відкидаються",
          HttpClient.host_of("https://user@Example.COM:8443/a") == "example.com")
    check("сміття не ламає", HttpClient.host_of("не-url") == "")

    c = HttpClient(timeout=5, max_retries=0, rps=12.0)
    try:
        portal = "https://prozorro.gov.ua/api/search/tenders"
        market = "https://market-api.prozorro.gov.ua/api/products/x"
        cdb = "https://public.api.openprocurement.org/api/2.5/tenders"
        docs = "https://public-docs.prozorro.gov.ua/get/abc"

        check("портал під квотою", c._limiter_for(portal) is not None)
        # Ось у чому була пастка: market-api містить prozorro.gov.ua підрядком,
        # тож зіставлення за підрядком садило каталог на квоту порталу — хоча
        # квота в нього своя (перевірено: портал у блоці, каталог відповідає).
        check("каталог НЕ під квотою порталу", c._limiter_for(market) is None)
        check("ЦБД не під квотою", c._limiter_for(cdb) is None)
        check("сервер документів не під квотою", c._limiter_for(docs) is None)

        check("каталог має власний темп", c._rate_for(market) is not None)
        check("ЦБД має власний темп", c._rate_for(cdb) is not None)
        check("портал темпом не обмежують — у нього квота", c._rate_for(portal) is None)
        check("сервер документів іде під загальним лімітом", c._rate_for(docs) is None)

        check("квота порталу — виміряні 60 на 60 с",
              HOST_QUOTAS["prozorro.gov.ua"][0] <= 60
              and HOST_QUOTAS["prozorro.gov.ua"][1] >= 60.0,
              HOST_QUOTAS["prozorro.gov.ua"])
        check("темпи не завищені понад виміряне",
              all(rate <= 30.0 for rate in HOST_RATES.values()), HOST_RATES)

        # Виміряна межа — оцінка, а не обіцянка сервера. Тому темп для таких
        # хостів має підлаштовуватись сам: інакше стала константа ловила б
        # 429 знову й знову (саме це й спостерігалося на двох тисячах запитів
        # до Центральної бази поспіль).
        rl = c._rate_for(cdb)
        before = rl.min_interval
        rl.penalize()
        check("після відмови темп знижується", rl.min_interval > before,
              (before, rl.min_interval))
        slowed = rl.min_interval
        for _ in range(25):
            rl.reward()
        check("за серії вдалих запитів темп повертається",
              rl.min_interval < slowed, (slowed, rl.min_interval))
        check("але не швидше за виміряну межу", rl.min_interval >= rl.base_interval,
              (rl.min_interval, rl.base_interval))
    finally:
        c.close()



def test_output_dir(tmp: Path) -> None:
    """Книга має лягати в теку з налаштувань — там її шукає сторінка аналітики."""
    from app.paths import default_output_dir, export_path

    print("\n=== тека вивантаження ===")
    check("без вказівки — типова тека",
          export_path("книга").parent == default_output_dir(),
          export_path("книга").parent)

    chosen = tmp / "моя тека"
    got = export_path("книга", folder=chosen)
    check("задана тека шанується", got.parent == chosen, got.parent)
    check("теку створено", chosen.is_dir())
    check("у назві є відмітка часу й розширення",
          got.name.startswith("книга-") and got.suffix == ".xlsx", got.name)

    # Саме через цю розбіжність книга «зникала»: конвеєр писав її в теку
    # проєкту, а аналітика перебирала теку з налаштувань.
    s = Settings()
    s.output_dir = str(chosen)
    check("конвеєр і аналітика дивляться в одну теку",
          export_path("prozorro-дані", folder=s.output_dir).parent == Path(s.output_dir))


def test_output_dir_inside_project(tmp: Path) -> None:
    """Типово все має лягати в ``downloads`` проєкту — і на чужій машині теж."""
    import json

    from app.paths import (DOWNLOADS_DIRNAME, PROJECT_ROOT, default_output_dir,
                           resolve_output_dir, store_output_dir)

    print("\n=== усе своє — в теку проєкту ===")
    check("типова тека — downloads у проєкті",
          default_output_dir() == PROJECT_ROOT / DOWNLOADS_DIRNAME,
          default_output_dir())
    check("теку створено одразу", default_output_dir().is_dir())

    for empty in ("", "   ", None):
        check(f"порожнє значення ({empty!r}) → downloads",
              resolve_output_dir(empty) == PROJECT_ROOT / DOWNLOADS_DIRNAME,
              resolve_output_dir(empty))
    check("відносний шлях відлічується від кореня проєкту, а не від cwd",
          resolve_output_dir("downloads/архів") == PROJECT_ROOT / "downloads" / "архів",
          resolve_output_dir("downloads/архів"))

    # settings.json лежить у репозиторії: абсолютний шлях у ньому вів би збір
    # на чужій машині в тéку з іншого профілю. Тому пишемо його відносним.
    check("типова тека зберігається відносною",
          store_output_dir(default_output_dir()) == DOWNLOADS_DIRNAME,
          store_output_dir(default_output_dir()))
    outside = tmp / "поза проєктом"
    check("свідомо обрана тека поза проєктом лишається абсолютною",
          store_output_dir(outside) == str(outside), store_output_dir(outside))

    # Обіг через файл: збереження → читання не має виводити теку з проєкту.
    s = Settings()
    check("у пам'яті шлях абсолютний — його підставляють просто в Path",
          Path(s.output_dir).is_absolute(), s.output_dir)
    saved = tmp / "налаштування.json"
    s.save(saved)
    check("у файлі — відносний шлях",
          json.loads(saved.read_text(encoding="utf-8"))["output_dir"] == DOWNLOADS_DIRNAME,
          json.loads(saved.read_text(encoding="utf-8"))["output_dir"])
    check("після читання — знову тека проєкту",
          Path(Settings.load(saved).output_dir) == PROJECT_ROOT / DOWNLOADS_DIRNAME,
          Settings.load(saved).output_dir)

    # Налаштування зі старим типовим значенням — абсолютним шляхом з чужого
    # профілю — після перезапису стають переносними.
    alien = tmp / "чужі.json"
    alien.write_text(json.dumps({"output_dir": "Z:/Хтось/ProzorroDownloader"},
                                ensure_ascii=False), encoding="utf-8")
    check("чужа абсолютна тека лишається вибором користувача",
          Settings.load(alien).output_dir.startswith("Z:"),
          Settings.load(alien).output_dir)


def test_one_date() -> None:
    """Пошук мусить відбирати за тією самою датою, за якою сортує видачу."""
    from app.core.pipeline import LOOKBACK_SLACK_DAYS, Pipeline

    print("\n=== одна дата на весь пошук ===")
    # Дата оприлюднення закодована в номері; саме їй дорівнює dateCreated у ЦБД,
    # за яким потім фільтрує _passes_filters (перевірено на живих даних).
    card = {"tenderID": "UA-2026-08-21-010792-a",
            "tenderPeriod": {"startDate": "2026-08-27T00:00:00+03:00"},
            "enquiryPeriod": {"startDate": "2026-08-21T10:00:00+03:00"}}
    check("дата береться з номера", Pipeline._published(card) == "2026-08-21",
          Pipeline._published(card))
    check("а не з початку подання пропозицій",
          Pipeline._published(card) != card["tenderPeriod"]["startDate"][:10])

    # Без номера лишається запасний варіант.
    check("запасний варіант — період уточнень",
          Pipeline._published({"enquiryPeriod": {"startDate": "2026-01-05T00:00:00+02:00"}})
          == "2026-01-05")
    check("потім період подання",
          Pipeline._published({"tenderPeriod": {"startDate": "2026-02-07T00:00:00+02:00"}})
          == "2026-02-07")
    check("порожня картка не ламає", Pipeline._published({}) == "")

    # Поки дати збігаються, гортати глибше за період майже не треба.
    check("запас на гортання невеликий", LOOKBACK_SLACK_DAYS <= 7, LOOKBACK_SLACK_DAYS)


def test_resolve_choice(tmp: Path) -> None:
    """Спосіб розпізнавання обирається за розміром вибірки."""
    print("\n=== вибір способу розпізнавання ===")
    db = Database(tmp / "choice.db")

    def budget(days):
        preset = SearchPreset(date_from="2026-01-01",
                              date_to=(date(2026, 1, 1) + timedelta(days=days - 1)).isoformat())
        return Pipeline(Settings(), preset, db)._summary_budget()

    check("на тиждень вигідно опитати одиниці", budget(7) == 10, budget(7))
    check("на рік — уже сотні", budget(365) == 547, budget(365))
    check("запас росте з періодом", budget(30) < budget(90) < budget(365))
    db.close()


def test_summary_resolver(tmp: Path) -> None:
    """Прямий розв'язувач порталу: номер → внутрішній UUID."""
    from app.core.api import SUMMARY_URL
    from app.core.resolver import Resolver

    print("\n=== прямий розв'язувач порталу ===")
    check("адреса на порталі й приймає номер",
          SUMMARY_URL.startswith("https://prozorro.gov.ua/") and "{tender_id}" in SUMMARY_URL,
          SUMMARY_URL)

    asked = []

    class FakeClient:
        def get_json(self, url, params=None):
            asked.append(url)
            tid = url.rsplit("/", 2)[-2]
            if tid.endswith("-zzz"):
                raise RuntimeError("немає такої закупівлі")
            return {"tenderID": tid, "id": "uuid-" + tid[-4:]}

    db = Database(tmp / "summary.db")
    res = Resolver.__new__(Resolver)
    res.client = FakeClient()
    res.db = db
    res.concurrency = 2
    res.search_concurrency = 2
    res._log = lambda level, msg: None
    res._lock = __import__("threading").Lock()

    ids = ["UA-2026-01-01-000001-a", "UA-2026-01-01-000002-a", "UA-2026-01-01-000-zzz"]
    seen = []
    found = res.from_summary(ids, lambda st, d, t: seen.append((st, d, t)))

    check("розпізнано те, що існує", len(found) == 2, found)
    check("UUID узято з поля id", found["UA-2026-01-01-000001-a"] == "uuid-01-a",
          found.get("UA-2026-01-01-000001-a"))
    check("відсутня закупівля не валить розбір",
          "UA-2026-01-01-000-zzz" not in found)
    check("один запит на закупівлю", len(asked) == 3, len(asked))
    check("поступ визначений — видно відсоток",
          seen and all(t == 3 for _s, _d, t in seen), seen)
    check("знайдене лягає в локальний індекс",
          db.index_lookup(ids).get("UA-2026-01-01-000002-a") == "uuid-02-a")
    db.close()


def test_index_shards() -> None:
    """Період ріжеться так, щоб потоки індексації справді працювали."""
    from app.core.resolver import IndexBuilder

    print("\n=== відрізки індексації ===")
    since, until = date(2026, 8, 18), date(2026, 8, 25)
    chunks = IndexBuilder.date_chunks(since, until, 8)
    check("на вісім діб — вісім відрізків", len(chunks) == 8, len(chunks))
    check("відрізки покривають період без дір",
          chunks[0][0] == since and chunks[-1][1] == until
          and all(chunks[i][1] + timedelta(days=1) == chunks[i + 1][0]
                  for i in range(len(chunks) - 1)), chunks)

    # Саме тут була вада: місячне різання давало один відрізок на короткий
    # період, і вісім потоків гортали стрічку по черзі.
    check("помісячно вийшов би один відрізок",
          len(IndexBuilder.month_chunks(since, until)) == 1)

    check("більше потоків, ніж діб — не дробимо зайве",
          len(IndexBuilder.date_chunks(since, since, 8)) == 1)
    long_chunks = IndexBuilder.date_chunks(date(2025, 1, 1), date(2025, 12, 31), 4)
    check("довгий період — рівно за кількістю потоків", len(long_chunks) == 4,
          len(long_chunks))
    check("і теж без дір",
          long_chunks[0][0] == date(2025, 1, 1) and long_chunks[-1][1] == date(2025, 12, 31))


def test_paths(tmp: Path) -> None:
    print("\n=== розкладання файлів по теках ===")
    db = Database(tmp / "test.db")
    root = tmp / "out"
    dl = FileDownloader(None, db, root)
    folder = tender_folder(root, "UA-2026-01-01-000001-a", "Ноутбуки/партія №1", "2026-01-15")
    check("тека за місяцем", folder.parent.name == "2026-01")
    check("слеш у назві прибрано", "/" not in folder.name)

    doc = {"scope": "bid", "owner_name": "ТОВ Х", "owner_edrpou": "111", "title": "акт.pdf"}
    first = dl.target_path(dict(doc), folder)
    second = dl.target_path(dict(doc), folder)
    check("однакові назви розведено", first != second, f"{first.name} / {second.name}")
    check("ЄДРПОУ у назві теки учасника", "111" in first.parent.name, first.parent.name)

    recorded = {"scope": "bid", "title": "акт.pdf",
                "local_path": str(folder / "Пропозиція" / "старе.pdf")}
    check("збережений шлях перевикористано",
          dl.target_path(recorded, folder).name == "старе.pdf")
    foreign = {"scope": "bid", "title": "акт.pdf", "local_path": r"D:\інша\тека\файл.pdf"}
    check("шлях поза текою завантажень ігнорується",
          dl.target_path(foreign, folder).name != "файл.pdf")
    check("розширення з MIME",
          dl._file_name({"title": "документ", "format": "application/pdf"}) == "документ.pdf")
    db.close()


def test_file_filter() -> None:
    print("\n=== фільтр типів файлів ===")
    check("підпис за MIME", is_signature({"title": "sign", "format": "application/pkcs7-signature"}))
    check("підпис за назвою", is_signature({"title": "sign.p7s", "format": ""}))
    check("підпис у назві документа", is_signature({"title": "Витяг_МВС.p7s"}))
    check("PDF не підпис", not is_signature({"title": "договір.pdf", "format": "application/pdf"}))
    check("архів не підпис", not is_signature({"title": "пропозиція.zip"}))
    check("розширення з назви", document_extension({"title": "Договір №5.PDF"}) == ".pdf")
    check("розширення з MIME", document_extension(
        {"title": "Договір", "format": "application/pdf"}) == ".pdf")
    check("без розширення", document_extension({"title": "Договір", "format": ""}) == "")
    check("нормалізація переліку",
          normalize_extensions(["PDF", ".docx", " xls ", "*.rtf", ""])
          == {".pdf", ".docx", ".xls", ".rtf"})
    check("порожній перелік", normalize_extensions(None) == set())

    # Замовники раз у раз називають файл датою. Правило «усе після останньої
    # крапки» робило з «…від 05.05.2026» розширення «.2026»: файл лягав на
    # диск без розширення, а фільтр типів викидав його як «не PDF».
    from app.core.downloader import extension_of, file_name_for
    dated = {"title": "Протокол відхилення від 05.05.2026", "format": "application/pdf"}
    check("дата в назві — не розширення", extension_of(dated["title"]) == "",
          extension_of(dated["title"]))
    check("тип беремо з MIME", document_extension(dated) == ".pdf", document_extension(dated))
    check("і файл лягає з розширенням", file_name_for(dated).endswith(".pdf"),
          file_name_for(dated))
    check("цифри в розширенні дозволені", extension_of("архів.7z") == ".7z")
    check("самі цифри — ні", extension_of("частина.5") == "")
    check("надто довгий хвіст — не розширення", extension_of("назва.абвгдеєжзи") == "")


def test_widgets() -> None:
    print("\n=== поля вводу ===")
    from PySide6.QtCore import QDate
    from PySide6.QtGui import QValidator
    from PySide6.QtWidgets import QApplication

    from app.ui.widgets.common import EDRPOU_RE, DateRange, EdrpouList, MoneyEdit

    app = QApplication.instance() or QApplication([])

    money = MoneyEdit("від")
    check("порожнє поле — без обмеження", money.value() is None)
    money.setText("1 234 567")
    check("сума з пробілами", money.value() == 1234567.0, money.value())
    # Кома — це роздільник копійок, а не ще один пробіл. Перевірка колись
    # закріплювала протилежне: «50000,00» читалося як 5 000 000, і фільтр
    # «від» виходив у сто разів більшим, ніж просив користувач.
    money.setText("50000,00")
    check("сума з комою", money.value() == 50000.0, money.value())
    money.setText("1 234,56")
    check("копійки не множать суму", money.value() == 1234.56, money.value())
    money.setText("1234.56")
    check("крапка теж роздільник копійок", money.value() == 1234.56, money.value())
    money.set_value(1234.56)
    check("копійки показуються, коли вони є", money.text() == "1 234,56", money.text())
    check("і читаються назад без втрат", money.value() == 1234.56, money.value())
    money.set_value(1234567)
    formatted = money.text()
    digits = "".join(ch for ch in formatted if ch.isdigit())
    check("форматування з групами",
          digits == "1234567" and "," not in formatted and len(formatted) == 9,
          repr(formatted))
    check("сформатоване читається назад", money.value() == 1234567.0)
    # Головне, заради чого поле переписане: після форматування текст мусить
    # лишатися прийнятним для валідатора, інакше ввід блокується.
    state, _, _ = money.validator().validate(formatted, len(formatted))
    check("сформатований текст лишається валідним",
          state == QValidator.State.Acceptable, state.name)
    for typed in ("5", "50000", "1 000", "1"):
        state, _, _ = money.validator().validate(typed, len(typed))
        check(f"ввід {typed!r} приймається", state == QValidator.State.Acceptable, state.name)
    money.set_value(None)
    check("скидання у порожнє", money.text() == "")

    check("ЄДРПОУ юрособи", EDRPOU_RE.findall("код 41263186 ТОВ") == ["41263186"])
    check("РНОКПП ФОП", EDRPOU_RE.findall("2381311679") == ["2381311679"])
    check("кілька кодів", EDRPOU_RE.findall("41263186, 44996712") == ["41263186", "44996712"])
    check("довгий номер не ріжеться", EDRPOU_RE.findall("123456789012345678") == [])
    check("короткий номер ігнорується", EDRPOU_RE.findall("12345") == [])

    codes = EdrpouList()
    codes.input.setText("41263186 та 44996712")
    codes._add()
    check("додано обидва коди", codes.values() == ["41263186", "44996712"], codes.values())
    codes.input.setText("ТОВ Ромашка")
    codes._add()
    # isVisible() у неприкріпленого віджета завжди False, тому дивимось на
    # намір показати підказку та на те, що ввід лишився на місці.
    check("непридатний ввід не з'їдається мовчки",
          codes.hint.isVisibleTo(codes) and codes.input.text() == "ТОВ Ромашка")
    codes.input.setText("41263186")
    codes._add()
    check("дублікат не додається двічі", codes.values() == ["41263186", "44996712"])
    check("підказка зникає після вдалого вводу", not codes.hint.isVisibleTo(codes))

    period = DateRange()
    period.set_values("2026-05-01", "2026-01-01")
    period.normalize()
    check("переставлені дати виправлено", period.values() == ("2026-01-01", "2026-05-01"),
          period.values())
    period.set_values("не дата", "")
    check("сміття замість дати не ламає поле",
          period.date_from.date().isValid() and period.date_to.date() == QDate.currentDate())
    app.processEvents()


def test_download_missing(tmp: Path) -> None:
    print("\n=== довантаження відсутніх файлів ===")
    db = Database(tmp / "retry.db")
    settings = Settings()
    settings.output_dir = str(tmp / "out")

    result = Pipeline(settings, SearchPreset(), db).download_missing()
    check("порожня база не падає", result.files_ok == 0 and not result.error, result.error)

    # Документи, зафіксовані під час збору без файлів, мають потрапляти в чергу.
    tender = {"id": "t3", "tenderID": "UA-2026-01-01-000003-a",
              "documents": [
                  {"id": "d1", "url": "https://x/1", "title": "умови.pdf",
                   "format": "application/pdf"},
                  {"id": "d2", "url": "https://x/2", "title": "sign.p7s",
                   "format": "application/pkcs7-signature"}]}
    parsed = parse_tender(tender)
    db.save_tender(parsed["row"], lots=[], items=[], bids=[], awards=[],
                   contracts=[], docs=parsed["docs"])
    queued = db.query("SELECT state, COUNT(*) n FROM documents GROUP BY state")
    check("документи лягли в чергу", [dict(r) for r in queued] == [{"state": "pending", "n": 2}],
          [dict(r) for r in queued])

    preset = SearchPreset()
    preset.skip_signatures = True
    pipeline = Pipeline(settings, preset, db)
    # Мережу не чіпаємо: перевіряємо лише те, що з черги береться в роботу.
    selected = pipeline._apply_file_filter(
        [dict(r) for r in db.pending_documents()])
    check("підпис відсіяно, документ лишився",
          [d["title"] for d in selected] == ["умови.pdf"], [d["title"] for d in selected])
    states = {r["state"]: r["n"] for r in
              db.query("SELECT state, COUNT(*) n FROM documents GROUP BY state")}
    check("підпис позначено відсіяним", states.get("filtered") == 1, states)

    # Відсіяне має повертатися в чергу, якщо фільтр вимкнули.
    preset.skip_signatures = False
    again = Pipeline(settings, preset, db)._apply_file_filter(
        [dict(r) for r in db.pending_documents()])
    check("після вимкнення фільтра підпис знову в черзі", len(again) == 2, len(again))
    db.close()


def test_report_stats() -> None:
    print("\n=== статистика звіту ===")
    from app.core.report import (
        Block, Profile, Report, compact, gini, hhi, median, pct, quantile, robust_z,
    )

    check("медіана парної кількості", median([1, 2, 3, 4]) == 2.5)
    check("медіана порожнього", median([]) == 0.0)
    check("квантиль як у Excel", quantile([1, 2, 3, 4], 0.25) == 1.75,
          quantile([1, 2, 3, 4], 0.25))
    check("HHI монополії", hhi([1.0]) == 10000.0, hhi([1.0]))
    check("HHI рівних часток", hhi([0.25] * 4) == 2500.0, hhi([0.25] * 4))
    check("Джині рівних", gini([5, 5, 5, 5]) == 0.0, gini([5, 5, 5, 5]))
    check("Джині нерівних", gini([1, 1, 1, 97]) > 0.6, gini([1, 1, 1, 97]))

    # Стійкий z-показник має ловити викид, якого звичайний z не помітив би:
    # одне велике значення саме роздуває і середнє, і стандартне відхилення.
    scores = robust_z([10, 11, 10, 12, 11, 10, 11, 900])
    check("викид знайдено", scores[-1] > 3.5, round(scores[-1], 1))
    check("норма не позначена", max(scores[:-1]) < 3.5, round(max(scores[:-1]), 1))
    check("без розкиду немає викидів", max(robust_z([5] * 10)) == 0.0)

    check("коротка сума", compact(12_345_678).endswith("млн"), compact(12_345_678))
    # Ділення на нуль десь у рушії не має перетворюватись на «inf млрд».
    check("нескінченність — риска", compact(float("inf")) == "—", compact(float("inf")))
    check("NaN — риска", compact(float("nan")) == "—", compact(float("nan")))
    check("відсоток", pct(0.1234) == "12,3%", pct(0.1234))

    # Гравець без блока — це рядок рейтингу, а не сторінка: портрет для нього
    # не будувався, тож і аркуша в книзі бути не повинно.
    block = Block("Профіль")
    block.tables.append(("ТМ", (["Марка"], [["Vinga"]])))
    report = Report()
    report.competitors = [Profile(edrpou="1", name="А", block=block),
                          Profile(edrpou="2", name="Б")]
    check("портрет є лише там, де є блок",
          [bool(p.block) for p in report.competitors] == [True, False])


def test_brands() -> None:
    print("\n=== торгові марки й документи ===")
    from app.core import brands

    check("латиниця", brands.detect("Ноутбук Vinga Iron S140") == ["Vinga"])
    check("лінійка моделі веде до бренду",
          brands.detect("Lenovo ThinkPad E14") == ["Lenovo"])
    check("кирилиця", brands.detect("монітор Філіпс 24") == ["Philips"])
    check("дефіс і пробіл рівноцінні",
          brands.detect("TP Link TL-SG1016D") == ["TP-Link"],
          brands.detect("TP Link TL-SG1016D"))
    # Найгірша помилка словника — знайти марку там, де її немає.
    check("немає хибних збігів", brands.detect("Шафа для одягу металева") == [])
    check("марка всередині слова не рахується", brands.detect("SHOPPING") == [])
    check("головна ТМ набору",
          brands.main_brand(["ПК Prologix", "ПК Prologix", "Монітор AOC"]) == "Prologix")

    check("підпис КЕП", brands.document_kind("sign.p7s") == "Підпис КЕП")
    check("сертифікат",
          brands.document_kind("Сертифікат відповідності.pdf").startswith("Сертифікат"))
    check("авторизаційний лист",
          brands.document_kind("Авторизаційний лист виробника.pdf").startswith("Лист"))
    check("невідомий файл", brands.document_kind("scan_001.jpg") == "")
    new = dict(brands.candidates(["Ноутбук Nekto X1", "Nekto X1 pro", "ще Nekto"],
                                 min_count=2))
    check("нова марка потрапляє в кандидати", new.get("Nekto") == 3, new)


#: Конкурент записаний так, як його пише реєстр: аналіз має скоротити форму
#: власності до «ТОВ», інакше вона з'їдає колонку в кожній таблиці звіту.
RIVAL = 'ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДАЛЬНІСТЮ "КОНКУРЕНТ"'


def _demo_workbook(tmp: Path) -> Path:
    """Маленька книга з усіма аркушами — щоб перевірити і читання, і аналіз."""
    from app.core.exporter import write_xlsx

    tenders = (["Номер закупівлі", "Дата оприлюднення", "Предмет закупівлі", "Опис",
                "Коди ДК021", "Статус", "Процедура", "Категорія предмета",
                "Очікувана вартість", "Валюта", "ПДВ включено", "Замовник",
                "ЄДРПОУ замовника", "Регіон", "Населений пункт", "Лотів", "Пропозицій",
                "Файлів", "Початок подання", "Кінець подання", "Остання зміна",
                "Посилання"],
               [[f"UA-2026-0{i}", f"2026-0{i}-10", "Ноутбуки", "", "30213100-6",
                 "Завершена", "Відкриті торги", "goods", 100000.0 + i, "UAH", "так",
                 "Замовник", "04527520", "Київська область", "Київ", 0, 2, 3,
                 "", "", "", ""] for i in range(1, 7)])
    bids = (["Номер закупівлі", "Дата закупівлі", "Учасник", "ЄДРПОУ учасника",
             "Регіон учасника", "Сума пропозиції", "Валюта", "Статус пропозиції",
             "Дата подання"],
            [row for i in range(1, 7) for row in (
                [f"UA-2026-0{i}", f"2026-0{i}-10", "НАШЕ ТОВ", "41263186",
                 "Київська область", 99000.0, "UAH", "active", f"2026-0{i}-11"],
                [f"UA-2026-0{i}", f"2026-0{i}-10", RIVAL, "12345678",
                 "Львівська область", 90000.0, "UAH", "active", f"2026-0{i}-11"])])
    contracts = (["Номер закупівлі", "Номер договору", "Постачальник",
                  "ЄДРПОУ постачальника", "Сума договору", "Валюта", "Статус договору",
                  "Дата підписання", "Замовник", "ЄДРПОУ замовника", "Регіон",
                  "Коди ДК021"],
                 [[f"UA-2026-0{i}", f"UA-2026-0{i}-a1", RIVAL, "12345678",
                   90000.0, "UAH", "active", f"2026-0{i}-20", "Замовник", "04527520",
                   "Київська область", "30213100-6"] for i in range(1, 6)]
                 # Останній договір — у євро: його має виключити очищення.
                 + [["UA-2026-06", "UA-2026-06-a1", RIVAL, "12345678", 5000.0,
                     "EUR", "active", "2026-06-20", "Замовник", "04527520",
                     "Київська область", "30213100-6"]])
    items = (["Номер закупівлі", "Дата", "Позиція", "Код ДК021", "Назва коду",
              "Кількість", "Одиниця", "Очікувана вартість лоту", "Ціна за одиницю",
              "Замовник", "ЄДРПОУ замовника", "Регіон"],
             [[f"UA-2026-0{i}", f"2026-0{i}-10", "Ноутбук Vinga Iron S140",
               "30213100-6", "Портативні комп'ютери", 5, "штука", 100000.0, 20000.0,
               "Замовник", "04527520", "Київська область"] for i in range(1, 7)])
    documents = (["Номер закупівлі", "Розділ", "Чий файл", "ЄДРПОУ власника",
                  "Назва файлу", "Тип документа", "Формат", "Дата оприлюднення",
                  "Розмір, байт", "Стан", "Шлях на диску", "Посилання"],
                 [[f"UA-2026-0{i}", "Пропозиція", RIVAL, "12345678",
                   "Авторизаційний лист виробника.pdf", "", "application/pdf",
                   f"2026-0{i}-11", 1000, "завантажено", "", ""] for i in range(1, 7)])
    path = tmp / "demo.xlsx"
    write_xlsx(path, {"Закупівлі": tenders, "Номенклатура": items, "Пропозиції": bids,
                      "Договори": contracts, "Документи": documents})
    return path


def test_xlsxload(tmp: Path) -> None:
    print("\n=== читання книги ===")
    from app.core import xlsxload

    check("ЄДРПОУ з провідним нулем", xlsxload.as_edrpou(4527520) == "04527520",
          xlsxload.as_edrpou(4527520))
    check("РНОКПП не добивається", xlsxload.as_edrpou("2882812765") == "2882812765")
    check("число з пробілами", xlsxload.as_num("1 234,50") == 1234.5,
          xlsxload.as_num("1 234,50"))
    check("дата з крапками", xlsxload.as_date("17.08.2026") == "2026-08-17")
    check("сміття замість дати", xlsxload.as_date("невідомо") == "")

    data = xlsxload.load(_demo_workbook(tmp))
    check("закупівлі прочитані", len(data.tenders) == 6, len(data.tenders))
    check("позиції прочитані", len(data.items) == 6, len(data.items))
    check("немає аркуша каталогу", "Товари каталогу" in data.missing, data.missing)
    # Найпідступніше місце зіставлення: «Замовник» не має забрати колонку
    # «ЄДРПОУ замовника» і навпаки.
    check("замовник і його код не переплутані",
          data.tenders[0]["buyer"] == "Замовник"
          and data.tenders[0]["buyer_edrpou"] == "04527520", data.tenders[0])


def test_insight(tmp: Path) -> None:
    print("\n=== аналітика ===")
    from app.core import insight, xlsxload
    from app.core.report import NEEDS_AI

    data = xlsxload.load(_demo_workbook(tmp))
    report = insight.analyse(data, own_edrpou=["41263186"])

    check("усі розділи на місці", len(report.sections) == 9, list(report.sections))
    check("прогноз стоїть одразу за ринком",
          list(report.sections).index("Прогнозування")
          == list(report.sections).index("Ринок") + 1, list(report.sections))
    check("підсумок першим", list(report.sections)[0] == "Підсумок", list(report.sections))

    check("повна форма власності скорочується",
          insight.short_org(RIVAL) == 'ТОВ "КОНКУРЕНТ"', insight.short_org(RIVAL))
    check("ФОП через тире", insight.short_org("ФІЗИЧНА ОСОБА – ПІДПРИЄМЕЦЬ ПЕТРЕНКО")
          == "ФОП ПЕТРЕНКО")
    check("довша форма має перевагу над коротшою",
          insight.short_org("ПРИВАТНЕ АКЦІОНЕРНЕ ТОВАРИСТВО «ІКС»") == "ПрАТ «ІКС»",
          insight.short_org("ПРИВАТНЕ АКЦІОНЕРНЕ ТОВАРИСТВО «ІКС»"))
    check("назву без форми не чіпаємо", insight.short_org("КОМЕЛ") == "КОМЕЛ")
    # У назвах філій форма стоїть у родовому відмінку — це та сама форма.
    branch = 'ФІЛІЯ "ЕНЕРГОРЕМТРАНС" ПУБЛІЧНОГО АКЦІОНЕРНОГО ТОВАРИСТВА "УЗ"'
    check("родовий відмінок теж скорочується",
          insight.short_org(branch) == 'ФІЛІЯ "ЕНЕРГОРЕМТРАНС" ПАТ "УЗ"',
          insight.short_org(branch))
    # М'який перенос посеред слова: очима назва звичайна, а збіг ламається.
    check("м'який перенос не заважає",
          insight.short_org("ТОВАРИСТВО З ОБМЕЖЕНОЮ ВІДПОВІДА­ЛЬНІСТЮ «ІКС»")
          == "ТОВ «ІКС»")

    rival = next(p for p in report.competitors if p.edrpou == "12345678")
    check("у звіті компанія — «ТОВ»", rival.name == 'ТОВ "КОНКУРЕНТ"', rival.name)
    check("сума конкурента без чужої валюти", rival.signed == 450000.0, rival.signed)
    check("ТМ конкурента розпізнана", rival.top_brand == "Vinga", rival.top_brand)
    check("товар конкурента", rival.main_product == "Ноутбук", rival.main_product)
    check("авторизації полічені", rival.authorizations == 6, rival.authorizations)
    # Шоста закупівля пішла в євро й вилетіла при очищенні, тож перемог
    # лишається п'ять із шести подань.
    check("результативність конкурента", round(rival.win_rate, 3) == 0.833, rival.win_rate)

    ours = next(p for p in report.ours if p.edrpou == "41263186")
    check("наші подавалися", ours.n_bids == 6, ours.n_bids)
    check("наші не виграли", ours.signed == 0.0, ours.signed)

    losses = dict(rival.block.tables)["Наші зустрічі"]
    verdicts = {row[7] for row in losses[1]}
    check("причина програшу — ціна", verdicts == {"програли за ціною"}, verdicts)

    clean_block = report.sections["Очищення"][0]
    reasons = {row[2] for row in dict(clean_block.tables)["Зауваження до даних"][1]}
    check("чужа валюта виключена", "Валюта не гривня" in reasons, reasons)

    supply = dict(report.sections["Постачання"][0].tables)["Канали постачання"]
    verdict = next(row[-1] for row in supply[1] if row[0] == "12345678")
    check("канал визначено як власну ТМ", "власна ТМ" in verdict, verdict)
    # Там, де довідник мовчить, аналіз має чесно казати про свою межу.
    brands_sheet = dict(report.sections["Товари і ТМ"][0].tables)["Торгові марки"]
    check("невідомий дистриб'ютор позначено чесно",
          all(row[-1] == NEEDS_AI for row in brands_sheet[1]), brands_sheet[1])

    compare = dict(report.sections["Порівняння"][0].tables)["Сильні та слабкі сторони"]
    check("є чесна позначка про межі аналізу",
          all(row[-1] == NEEDS_AI for row in compare[1]), compare[1][0][-1])

    # Книга звіту має вмістити все, що показує застосунок, — і жодного
    # аркуша понад те: розділ, гравець із портретом і зміст.
    from openpyxl import load_workbook

    from app.core.reportbook import chart_table, write_report

    path = tmp / "звіт-аналітики.xlsx"
    write_report(path, report)
    players = [p for p in [*report.ours, *report.competitors] if p.block]
    blocks = [block for group in report.sections.values() for block in group]
    blocks += [p.block for p in players]
    want_tables = sum(1 for b in blocks for _name, s in b.tables if s and s[1])
    want_charts = sum(1 for b in blocks for c in b.charts if chart_table(c)[1])
    wb = load_workbook(path)
    got_tables = sum(len(ws.tables) for ws in wb.worksheets)
    # Оглядовий аркуш повторює діаграми розділів, тож рахуємо без нього:
    # питання тут у тому, чи нічого не загубилося, а не скільки разів воно
    # намальоване.
    got_charts = sum(len(ws._charts) for ws in wb.worksheets if ws.title != "Графіки")
    check("усі таблиці звіту потрапили в книгу", got_tables == want_tables,
          f"{got_tables} з {want_tables}")
    check("усі діаграми звіту потрапили в книгу", got_charts == want_charts,
          f"{got_charts} з {want_charts}")
    titles = {c.title for group in report.sections.values() for b in group
              for c in b.charts if chart_table(c)[1]}
    check("оглядовий аркуш зібрав кожен графік розділів рівно раз",
          len(wb["Графіки"]._charts) == len(titles),
          f"{len(wb['Графіки']._charts)} з {len(titles)}")
    check("аркуш на зміст, огляд, розділ і гравця",
          len(wb.sheetnames) == 2 + len(report.sections) + len(players),
          wb.sheetnames)
    wb.close()


def _crowded_workbook(tmp: Path) -> Path:
    """Книга, де наше ТОВ свідомо стоїть далеко за межею топ-15.

    Двадцять конкурентів із сумами, що спадають, і одна маленька наша угода
    в хвості: саме та вибірка, на якій наше ТОВ зникало з графіків.
    """
    from app.core.exporter import write_xlsx

    players = [(f"{10000000 + i}", f"КОНКУРЕНТ {i:02d}", 900000.0 - i * 10000.0)
               for i in range(20)]
    players.append(("41263186", "НАШЕ ТОВ", 1000.0))

    tenders, bids, contracts, items = [], [], [], []
    for i, (edrpou, name, amount) in enumerate(players, start=1):
        number = f"UA-2026-{i:04d}"
        tenders.append([number, "2026-03-10", "Ноутбуки", "", "30213100-6",
                        "Завершена", "Відкриті торги", "goods", amount * 1.1, "UAH",
                        "так", "Замовник", "04527520", "Київська область", "Київ",
                        0, 1, 0, "", "", "", ""])
        bids.append([number, "2026-03-10", name, edrpou, "Київська область",
                     amount, "UAH", "active", "2026-03-11"])
        contracts.append([number, f"{number}-a1", name, edrpou, amount, "UAH",
                          "active", "2026-03-20", "Замовник", "04527520",
                          "Київська область", "30213100-6"])
        items.append([number, "2026-03-10", "Ноутбук Vinga Iron S140", "30213100-6",
                      "Портативні комп'ютери", 5, "штука", amount, amount / 5,
                      "Замовник", "04527520", "Київська область"])

    # Одне програне подання нашого ТОВ: ТМ звідти має потрапити в портфель,
    # але не в число проданих марок.
    bids.append(["UA-2026-0001", "2026-03-10", "НАШЕ ТОВ", "41263186",
                 "Київська область", 950000.0, "UAH", "active", "2026-03-11"])
    items.append(["UA-2026-0001", "2026-03-10", "Ноутбук Lenovo ThinkPad",
                  "30213100-6", "Портативні комп'ютери", 5, "штука", 900000.0,
                  180000.0, "Замовник", "04527520", "Київська область"])

    path = tmp / "crowded.xlsx"
    write_xlsx(path, {
        "Закупівлі": (["Номер закупівлі", "Дата оприлюднення", "Предмет закупівлі",
                       "Опис", "Коди ДК021", "Статус", "Процедура",
                       "Категорія предмета", "Очікувана вартість", "Валюта",
                       "ПДВ включено", "Замовник", "ЄДРПОУ замовника", "Регіон",
                       "Населений пункт", "Лотів", "Пропозицій", "Файлів",
                       "Початок подання", "Кінець подання", "Остання зміна",
                       "Посилання"], tenders),
        "Номенклатура": (["Номер закупівлі", "Дата", "Позиція", "Код ДК021",
                          "Назва коду", "Кількість", "Одиниця",
                          "Очікувана вартість лоту", "Ціна за одиницю", "Замовник",
                          "ЄДРПОУ замовника", "Регіон"], items),
        "Пропозиції": (["Номер закупівлі", "Дата закупівлі", "Учасник",
                        "ЄДРПОУ учасника", "Регіон учасника", "Сума пропозиції",
                        "Валюта", "Статус пропозиції", "Дата подання"], bids),
        "Договори": (["Номер закупівлі", "Номер договору", "Постачальник",
                      "ЄДРПОУ постачальника", "Сума договору", "Валюта",
                      "Статус договору", "Дата підписання", "Замовник",
                      "ЄДРПОУ замовника", "Регіон", "Коди ДК021"], contracts),
    })
    return path


def test_ours_never_hidden(tmp: Path) -> None:
    """Наше ТОВ видно на кожному графіку компаній, навіть поза топ-N."""
    print("\n=== наші ТОВ не зникають з графіків ===")
    from app.core import insight, xlsxload
    from app.core.insight import TOP_ROWS
    from app.core.players import TOP_RIVALS

    data = xlsxload.load(_crowded_workbook(tmp))
    report = insight.analyse(data, own_edrpou=["41263186"])
    ours = next(p for p in report.ours if p.edrpou == "41263186")
    check(f"наше ТОВ справді поза топ-{TOP_ROWS}", ours.rank > TOP_ROWS, ours.rank)

    def chart(section: str, needle: str):
        for block in report.sections[section]:
            for item in block.charts:
                if needle in item.title:
                    return item
        return None

    for section, needle, limit in (("Ринок", "Топ-", TOP_ROWS),
                                   ("Наші ТОВ", "Ми проти лідерів", TOP_RIVALS),
                                   ("Конкуренти", "Найбільші конкуренти", TOP_RIVALS)):
        item = chart(section, needle)
        check(f"графік «{needle}…» є в розділі «{section}»", item is not None)
        if not item:
            continue
        series = item.series[0]
        spot = [i for i, label in enumerate(series.labels) if "НАШЕ" in label.upper()]
        check(f"[{section}] наше ТОВ на графіку", len(spot) == 1, series.labels)
        if not spot:
            continue
        # Головне правило: наше ТОВ дописане в хвіст, а не втиснуте в топ.
        check(f"[{section}] воно стоїть за межею топ-{limit}", spot[0] >= limit, spot[0])
        check(f"[{section}] воно підсвічене", spot[0] in series.accent, series.accent)
        check(f"[{section}] номер місця видно в підписі",
              f"№{ours.rank}" in series.labels[spot[0]], series.labels[spot[0]])
        check(f"[{section}] решту топ-{limit} не обрізано",
              len(series.labels) == limit + 1, len(series.labels))

    # Друга правка: явна назва графіка про ТМ і період прямо в заголовку.
    brands = chart("Порівняння", "товарних марок")
    check("графік ТМ перейменовано", brands is not None,
          [c.title for c in report.sections["Порівняння"][0].charts])
    if brands:
        check("у назві сказано, що рахується",
              brands.title.startswith("Кількість різних товарних марок, проданих"),
              brands.title)
        check("у назві є період", "з 10.03.2026 до 20.03.2026" in brands.title,
              brands.title)
        check("слова «ширина портфеля» більше немає",
              "ортфел" not in brands.title, brands.title)
    # Програне подання дає ТМ у портфелі, але не в проданих марках.
    check("портфель нашого ТОВ — дві ТМ", len(ours.brands) == 2, ours.brands)
    check("продана — лише одна", len(ours.sold_brands) == 1, ours.sold_brands)


def test_charts() -> None:
    print("\n=== графіки й таблиці звіту ===")
    from PySide6.QtGui import QPixmap
    from PySide6.QtWidgets import QApplication

    from app.core.report import ChartData, Series
    from app.ui.widgets.charts import KINDS, _nice_step, build
    from app.ui.widgets.table import DataTable

    QApplication.instance() or QApplication([])

    check("крок сітки округлюється", _nice_step(97) == 25, _nice_step(97))
    # Дробовий крок на штуках дав би підписи «0 0 1 1».
    check("цілий крок для штук", _nice_step(1, whole=True) == 1.0,
          _nice_step(1, whole=True))

    series = [Series("Ряд", ["а", "б", "в"], [1.0, 5.0, 3.0]),
              Series("Ще", ["а", "б", "в"], [2.0, 1.0, 4.0])]
    for kind in KINDS:
        data = ChartData(f"Тест {kind}", kind, list(series))
        if kind == "scatter":
            data = ChartData("Точки", kind,
                             [Series("т", points=[(1.0, 2.0), (2.0, 3.0)], accent={1})])
        widget = build(data, "dark")
        widget.resize(600, 300)
        pixmap = QPixmap(widget.size())
        widget.render(pixmap)
        check(f"графік {kind} малюється", not pixmap.isNull())

    empty = build(ChartData("Порожньо", "bar", []), "light")
    empty.resize(300, 200)
    empty.render(QPixmap(empty.size()))
    check("порожній графік не падає", True)

    # Розрив у лінії — на ньому тримається графік прогнозу: факт займає ліву
    # частину осі, прогноз із межами — праву, і жодна точка не має малюватися
    # там, де значення немає. Нуль замість ``None`` намалював би обвал до осі.
    gapped = ChartData("Факт і прогноз", "line", [
        Series("Факт", ["01", "02", "03", "04"], [10.0, 12.0, 11.0, None]),
        Series("Прогноз", ["01", "02", "03", "04"], [None, None, 11.0, 13.0])])
    widget = build(gapped, "dark")
    widget.resize(600, 300)
    pixmap = QPixmap(widget.size())
    widget.render(pixmap)
    check("лінія з розривом малюється", not pixmap.isNull())
    only_gaps = build(ChartData("Самі дірки", "line",
                                [Series("Ряд", ["01", "02"], [None, None])]), "dark")
    only_gaps.resize(300, 200)
    only_gaps.render(QPixmap(only_gaps.size()))
    check("ряд із самих дірок не падає", True)

    table = DataTable(["Компанія", "Сума угод, грн", "Частка"],
                      [["А", 1234.5, 0.25], ["Б", None, 0.75], ["В", 90000.0, 0.0]])
    check("гроші з копійками, коли вони є",
          table.model.data(table.model.index(0, 1)) == "1 234,50",
          table.model.data(table.model.index(0, 1)))
    check("частка як відсоток", table.model.data(table.model.index(0, 2)) == "25,0%",
          table.model.data(table.model.index(0, 2)))
    check("порожня клітинка", table.model.data(table.model.index(1, 1)) == "")
    check("кругла сума без копійок", table.model.data(table.model.index(2, 1)) == "90 000",
          table.model.data(table.model.index(2, 1)))


#: Заголовки аркушів для синтетичних книг у тестах.
DEMO_HEADERS = {
    "tenders": ["Номер закупівлі", "Дата оприлюднення", "Предмет закупівлі", "Опис",
                "Коди ДК021", "Статус", "Процедура", "Категорія предмета",
                "Очікувана вартість", "Валюта", "ПДВ включено", "Замовник",
                "ЄДРПОУ замовника", "Регіон", "Населений пункт", "Лотів", "Пропозицій",
                "Файлів", "Початок подання", "Кінець подання", "Остання зміна",
                "Посилання"],
    "items": ["Номер закупівлі", "Дата", "Позиція", "Код ДК021", "Назва коду",
              "Кількість", "Одиниця", "Очікувана вартість лоту", "Ціна за одиницю",
              "Замовник", "ЄДРПОУ замовника", "Регіон"],
    "bids": ["Номер закупівлі", "Дата закупівлі", "Учасник", "ЄДРПОУ учасника",
             "Регіон учасника", "Сума пропозиції", "Валюта", "Статус пропозиції",
             "Дата подання"],
    "contracts": ["Номер закупівлі", "Номер договору", "Постачальник",
                  "ЄДРПОУ постачальника", "Сума договору", "Валюта", "Статус договору",
                  "Дата підписання", "Замовник", "ЄДРПОУ замовника", "Регіон",
                  "Коди ДК021"],
    "documents": ["Номер закупівлі", "Розділ", "Чий файл", "ЄДРПОУ власника",
                  "Назва файлу", "Тип документа", "Формат", "Дата оприлюднення",
                  "Розмір, байт", "Стан", "Шлях на диску", "Посилання"],
}


def _demo_headers(name: str) -> list[str]:
    return list(DEMO_HEADERS[name])


def test_fast_reader(tmp: Path) -> None:
    print("\n=== швидкий читач XLSX ===")
    import threading
    from app.core import insight, xlsxload
    from app.core.xlsxfast import FastWorkbook, _is_date_format

    check("формат дати впізнано", _is_date_format("dd.mm.yyyy"))
    check("грошовий формат — не дата", not _is_date_format('#,##0.00" грн"'),
          _is_date_format('#,##0.00" грн"'))
    check("відсоток — не дата", not _is_date_format("0.00%"))

    path = _demo_workbook(tmp)
    with FastWorkbook(path) as book:
        check("аркуші знайдено", "Закупівлі" in book.names, book.names)
        fast = {name: [list(row) for row in book.rows(name)] for name in book.names}

    # Головна перевірка швидкого читача — що він дає рівно те саме, що й
    # openpyxl: інакше пришвидшення нічого не варте.
    from openpyxl import load_workbook
    slow_book = load_workbook(path, read_only=True, data_only=True)
    same = True
    for sheet in slow_book.worksheets:
        slow = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        quick = fast.get(sheet.title, [])
        if len(slow) != len(quick):
            same = False
            print(f"    {sheet.title}: рядків {len(quick)} проти {len(slow)}")
            break
        for a, b in zip(slow, quick):
            trimmed = list(a)[:len(b)]
            if [x if x != "" else None for x in trimmed] != [x if x != "" else None for x in b]:
                same = False
                print(f"    {sheet.title}: {trimmed} != {b}")
                break
    slow_book.close()
    check("швидкий читач збігається з openpyxl", same)

    quick_data = xlsxload.load(path)
    slow_data = xlsxload._load_with(path, xlsxload._openpyxl_streams, None)
    check("Dataset однаковий обома шляхами",
          quick_data.tenders == slow_data.tenders
          and quick_data.contracts == slow_data.contracts,
          f"{len(quick_data.tenders)} / {len(slow_data.tenders)}")

    # Книга не з тих: читач має тихо віддати роботу openpyxl, а не впасти.
    broken = tmp / "broken.xlsx"
    broken.write_bytes("це не книга Excel".encode("utf-8"))
    failed = False
    try:
        xlsxload.load(broken)
    except Exception:
        failed = True
    check("зіпсований файл дає помилку, а не мовчазний нуль", failed)

    # Зупинка аналізу має спрацьовувати, а не чекати кінця.
    stop = threading.Event()
    stop.set()
    stopped = False
    try:
        insight.analyse(quick_data, own_edrpou=["41263186"], cancel_event=stop)
    except insight.Cancelled:
        stopped = True
    check("аналіз зупиняється на вимогу", stopped)


def test_workbook_listing(tmp: Path) -> None:
    print("\n=== перелік книг у теці ===")
    from app.core.xlsxload import find_workbooks

    folder = tmp / "downloads"
    (folder / "2026-08" / "UA-1").mkdir(parents=True, exist_ok=True)
    (folder / "prozorro-дані.xlsx").write_bytes(b"x")
    (folder / "~$тимчасовий.xlsx").write_bytes(b"x")
    # Документ закупівлі — теж .xlsx, але в підтеці: у списку йому не місце.
    (folder / "2026-08" / "UA-1" / "Додаток 1.xlsx").write_bytes(b"x")
    names = [item.name for item in find_workbooks(folder)]
    check("вивантаження знайдено", names == ["prozorro-дані.xlsx"], names)
    check("порожня тека не падає", find_workbooks(tmp / "немає") == [])


def test_scale(tmp: Path) -> None:
    print("\n=== поведінка на великій вибірці ===")
    import random
    import time
    from app.core import insight, xlsxload
    from app.core.exporter import write_xlsx

    random.seed(7)
    n = 1200
    tenders, items, bids, contracts, documents = [], [], [], [], []
    suppliers = [(f"4000000{i}", f"ТОВ \"ГРАВЕЦЬ {i}\"") for i in range(20)]
    suppliers.append(("41263186", "ТОВ \"ПАРТНЕР АЙ ТІ\""))
    for i in range(n):
        tid = f"UA-2026-{i:05d}-a"
        date = f"2026-{random.randint(1, 12):02d}-15"
        value = round(random.lognormvariate(11, 1), 2)
        tenders.append([tid, date, "Ноутбуки", "", "30213100-6", "Завершена",
                        "Відкриті торги", "goods", value, "UAH", "так", "Замовник",
                        "04527520", "Київська область", "Київ", 0, 2, 3, "", "", "", ""])
        items.append([tid, date, "Ноутбук Vinga Iron", "30213100-6", "Ноутбуки", 3,
                      "штука", value, round(value / 3, 2), "Замовник", "04527520",
                      "Київська область"])
        winner, name = random.choice(suppliers)
        amount = round(value * random.uniform(0.8, 1.0), 2)
        bids.append([tid, date, name, winner, "Київська область", amount, "UAH",
                     "active", date])
        contracts.append([tid, f"{tid}-a1", name, winner, amount, "UAH", "active", date,
                          "Замовник", "04527520", "Київська область", "30213100-6"])
        documents.append([tid, "Пропозиція", name, winner, "Сертифікат ISO.pdf", "",
                          "application/pdf", date, 1000, "завантажено", "", ""])
    path = tmp / "scale.xlsx"
    write_xlsx(path, {
        "Закупівлі": (_demo_headers("tenders"), tenders),
        "Номенклатура": (_demo_headers("items"), items),
        "Пропозиції": (_demo_headers("bids"), bids),
        "Договори": (_demo_headers("contracts"), contracts),
        "Документи": (_demo_headers("documents"), documents),
    })
    started = time.perf_counter()
    data = xlsxload.load(path)
    report = insight.analyse(data, own_edrpou=["41263186"])
    spent = time.perf_counter() - started
    check(f"{n} закупівель розібрано за {spent:.1f} с", spent < 20, f"{spent:.1f} с")
    check("усі закупівлі прочитані", len(data.tenders) == n, len(data.tenders))

    # Результативність не має суперечити сама собі: перемог із подань не
    # може бути більше, ніж самих подань.
    bad = [p.label for p in report.competitors
           if p.n_won_bids > p.n_bids or (p.win_rate or 0) > 1]
    check("результативність узгоджена", not bad, bad)
    totals = sum(p.signed for p in report.competitors) + sum(p.signed for p in report.ours)
    market = sum(row[6] for row in
                 dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"][1])
    check("сума профілів не перевищує ринок", totals <= market + 1,
          f"{totals:.0f} проти {market:.0f}")


def test_tracked_competitors(tmp: Path) -> None:
    print("\n=== відстежувані конкуренти ===")
    from app.config import COMPETITOR_COMPANIES, KNOWN_COMPANIES, Settings
    from app.core import insight, xlsxload

    fresh = Settings()
    check("конкуренти є в типових налаштуваннях",
          fresh.competitors == list(COMPETITOR_COMPANIES), fresh.competitors)
    # Типовий фільтр — клас 30210000-4, а не весь розділ 30: 25 кодів проти
    # 401, тобто пошук без паперу, меблів і калькуляторів.
    from app.config import DEFAULT_CPV_PREFIXES
    from app.core.classifiers import expand_prefixes, label_for
    check("типовий клас ДК021 — 3021", fresh.preset.cpv_prefixes == ["3021"],
          fresh.preset.cpv_prefixes)
    check("це саме 30210000-4",
          label_for("3021").startswith("30210000-4"), label_for("3021"))
    codes = expand_prefixes(DEFAULT_CPV_PREFIXES)
    check("клас розгортається у 25 кодів", len(codes) == 25, len(codes))
    check("сам клас теж у переліку", "30210000-4" in codes, codes[:3])
    # Межу класу закріплюємо явно: 30210000-4 — це обчислювальна техніка,
    # тобто комп'ютери, ноутбуки, планшети, робочі станції та сканери.
    check("комп'ютери всередині",
          {"30213100-6", "30213300-8", "30214000-2", "30216110-0"} <= set(codes),
          sorted({"30213100-6", "30213300-8", "30214000-2", "30216110-0"} - set(codes)))
    # А монітори, принтери й картриджі — уже інші класи (30230000, 30120000).
    # Якщо їх теж треба збирати, у типовий фільтр додається префікс «3023».
    outside = {"30231300-0", "30232110-8", "30125100-2"}
    check("монітори, принтери й картриджі — поза класом",
          not (outside & set(codes)), sorted(outside & set(codes)))
    check("папір і меблі — і поготів",
          not ({"30197630-1", "39130000-2"} & set(codes)))
    check("коди конкурентів — вісім цифр",
          all(len(code) == 8 and code.isdigit() for code in COMPETITOR_COMPANIES),
          list(COMPETITOR_COMPANIES))
    check("наші й конкуренти не перетинаються",
          not set(fresh.own_edrpou) & set(fresh.competitors))
    check("усі відомі компанії мають назву",
          all(KNOWN_COMPANIES.values()), KNOWN_COMPANIES)

    # Головне, заради чого список існує: такого конкурента розбирають поіменно,
    # навіть якщо в цій вибірці за обсягом він не потрапив би у верхівку.
    data = xlsxload.load(_demo_workbook(tmp))
    tracked = "24083083"
    report = insight.analyse(data, own_edrpou=["41263186"], tracked=[tracked])
    profile = next((p for p in report.competitors if p.edrpou == tracked), None)
    check("відстежуваний конкурент у звіті", profile is not None)
    if profile:
        check("його показано назвою, а не кодом", profile.name == "КОМЕЛ", profile.label)


def _excel_workbook(path: Path) -> Path:
    """Книга в діалекті самого Excel, а не ``openpyxl``.

    Наші вивантаження зберігають текст просто в клітинці (``inlineStr``) і не
    мають ані спільної таблиці рядків, ані форматів дати. Файл, збережений
    людиною в Excel, виглядає інакше: рядки лежать у ``sharedStrings.xml``,
    дати — це числа зі стилем, порожні колонки просто відсутні. Ці гілки
    читача інакше ніде не перевіряються.
    """
    import zipfile

    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    strings = ["Номер закупівлі", "Дата оприлюднення", "Предмет закупівлі",
               "Очікувана вартість", "Замовник", "ЄДРПОУ замовника",
               "UA-2026-000001-a", "Ноутбуки", "Лікарня №1"]
    shared = "".join(f"<si><t>{text}</t></si>" for text in strings)

    # Рядок 3 пропущено навмисно, у рядку 4 бракує колонки C.
    rows = """
      <row r="1">
        <c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>1</v></c>
        <c r="C1" t="s"><v>2</v></c><c r="D1" t="s"><v>3</v></c>
        <c r="E1" t="s"><v>4</v></c><c r="F1" t="s"><v>5</v></c>
      </row>
      <row r="2">
        <c r="A2" t="s"><v>6</v></c><c r="B2" s="1"><v>46000</v></c>
        <c r="C2" t="s"><v>7</v></c><c r="D2"><v>125000.5</v></c>
        <c r="E2" t="s"><v>8</v></c><c r="F2"><v>4527520</v></c>
      </row>
      <row r="4">
        <c r="A4" t="inlineStr"><is><t>UA-2026-000002-a</t></is></c>
        <c r="B4" s="2"><v>46001</v></c>
        <c r="D4"><v>7000</v></c>
        <c r="E4" t="str"><v>Школа</v></c><c r="F4" t="b"><v>1</v></c>
      </row>"""

    parts = {
        "[Content_Types].xml":
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-'
            'package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.'
            'openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.'
            'openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.styles+xml"/></Types>',
        "_rels/.rels":
            f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            f'relationships"><Relationship Id="rId1" Type="{rel}/officeDocument" '
            f'Target="xl/workbook.xml"/></Relationships>',
        "xl/workbook.xml":
            f'<workbook xmlns="{ns}" xmlns:r="{rel}"><sheets>'
            f'<sheet name="Закупівлі" sheetId="1" r:id="rId1"/></sheets></workbook>',
        "xl/_rels/workbook.xml.rels":
            f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
            f'relationships">'
            f'<Relationship Id="rId1" Type="{rel}/worksheet" Target="worksheets/sheet1.xml"/>'
            f'<Relationship Id="rId2" Type="{rel}/sharedStrings" Target="sharedStrings.xml"/>'
            f'<Relationship Id="rId3" Type="{rel}/styles" Target="styles.xml"/>'
            f'</Relationships>',
        "xl/sharedStrings.xml":
            f'<sst xmlns="{ns}" count="{len(strings)}" uniqueCount="{len(strings)}">'
            f'{shared}</sst>',
        "xl/styles.xml":
            f'<styleSheet xmlns="{ns}">'
            f'<numFmts count="1"><numFmt numFmtId="164" formatCode="dd.mm.yyyy"/></numFmts>'
            f'<cellXfs count="3"><xf numFmtId="0"/><xf numFmtId="14"/><xf numFmtId="164"/>'
            f'</cellXfs></styleSheet>',
        "xl/worksheets/sheet1.xml":
            f'<worksheet xmlns="{ns}"><sheetData>{rows}</sheetData></worksheet>',
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as book:
        for name, text in parts.items():
            book.writestr(name, text)
    return path


def test_excel_dialect(tmp: Path) -> None:
    print("\n=== книга в діалекті Excel ===")
    from openpyxl import load_workbook

    from app.core import xlsxload
    from app.core.xlsxfast import FastWorkbook

    path = _excel_workbook(tmp / "excel.xlsx")
    with FastWorkbook(path) as book:
        quick = [list(row) for row in book.rows("Закупівлі")]
    slow_book = load_workbook(path, read_only=True, data_only=True)
    slow = [list(row) for row in slow_book["Закупівлі"].iter_rows(values_only=True)]
    slow_book.close()

    check("спільна таблиця рядків прочитана",
          quick[0][:2] == ["Номер закупівлі", "Дата оприлюднення"], quick[0][:2])
    check("вбудований формат дати став датою",
          str(quick[1][1])[:10] == "2025-12-09", quick[1][1])
    check("власний формат дати теж", str(quick[2][1])[:10] == "2025-12-10", quick[2][1])
    check("число лишилось числом", quick[1][3] == 125000.5, quick[1][3])
    check("текст із формули", quick[2][4] == "Школа", quick[2][4])
    check("булеве значення", quick[2][5] is True, quick[2][5])
    # Пропущена колонка C має лишити діру рівно на своєму місці, інакше
    # всі наступні значення поїдуть на одну колонку вліво.
    check("пропущена колонка лишається порожньою", quick[2][2] is None, quick[2])
    # Порожні рядки порівнюємо окремо: пропущений у файлі рядок openpyxl
    # віддає порожнім кортежем, а наш читач просто не згадує. Обидва потім
    # однаково відкидаються, зате наш не сплутає такий привид із заголовком.
    check("той самий результат, що й в openpyxl",
          quick == [row for row in slow if row], f"{quick}\n      != {slow}")
    check("порожній рядок не вигадується", [] not in quick, quick)

    data = xlsxload.load(path)
    check("книга розібралась у Dataset", len(data.tenders) == 2, len(data.tenders))
    check("ЄДРПОУ з провідним нулем відновлено",
          data.tenders[0]["buyer_edrpou"] == "04527520", data.tenders[0]["buyer_edrpou"])
    check("дата у форматі ISO", data.tenders[0]["date"] == "2025-12-09",
          data.tenders[0]["date"])


def _empty_dataset(path: Path):
    from app.core.xlsxload import Dataset
    return Dataset(path=path)


def test_degenerate(tmp: Path) -> None:
    print("\n=== вироджені вибірки ===")
    from app.core import insight
    from app.core.xlsxload import Dataset

    # Порожня книга: звіт має скластися й нічого не поділити на нуль.
    report = insight.analyse(_empty_dataset(tmp / "порожньо.xlsx"),
                             own_edrpou=["41263186"])
    check("порожня вибірка не валить аналіз", len(report.sections) == 8,
          list(report.sections))
    # Книга з порожнього звіту має скластися: аркуші будуть, таблиць у них
    # не буде — і це не помилка, а чесна відповідь «нічого не знайдено».
    from app.core.reportbook import write_report
    book = write_report(tmp / "порожній-звіт.xlsx", report)
    check("книга з порожнього звіту записується", book.exists(),
          book.stat().st_size)

    def tender(number: int, **kw):
        row = {"tender_id": f"UA-{number}", "date": "2026-03-01", "title": "Ноутбуки",
               "description": "", "cpv_list": "30213100-6", "status": "Завершена",
               "method": "Відкриті торги", "category": "goods", "value": 100000.0,
               "currency": "UAH", "vat": "так", "buyer": "Замовник",
               "buyer_edrpou": "04527520", "region": "Київська область",
               "locality": "Київ", "n_lots": 0, "n_bids": 1, "n_docs": 0,
               "tender_start": "", "tender_end": "", "modified": "", "url": ""}
        row.update(kw)
        return row

    def contract(number: int, edrpou: str, amount: float, currency: str = "UAH"):
        return {"tender_id": f"UA-{number}", "contract_id": f"UA-{number}-a1",
                "name": f"ТОВ {edrpou}", "edrpou": edrpou, "amount": amount,
                "currency": currency, "status": "active", "signed": "2026-03-10",
                "buyer": "Замовник", "buyer_edrpou": "04527520",
                "region": "Київська область", "cpv_list": "30213100-6"}

    # Закупівлі є, грошей немає: жодного договору й жодного переможця.
    only_tenders = Dataset(path=tmp / "без-угод.xlsx",
                           tenders=[tender(i) for i in range(5)])
    report = insight.analyse(only_tenders, own_edrpou=["41263186"])
    market = report.sections["Ринок"][0]
    check("ринок без угод рахується", dict(market.tiles).get("Сума угод") == "0 грн",
          dict(market.tiles).get("Сума угод"))
    check("замовники все одно є",
          len(dict(market.tables)["Замовники"][1]) == 1,
          dict(market.tables)["Замовники"][1])

    # Уся вибірка в чужій валюті — після очищення не лишається нічого.
    foreign = Dataset(path=tmp / "євро.xlsx", tenders=[tender(i) for i in range(5)],
                     contracts=[contract(i, "12345678", 1000.0, "EUR") for i in range(5)])
    report = insight.analyse(foreign, own_edrpou=["41263186"])
    issues = dict(report.sections["Очищення"][0].tables)["Зауваження до даних"][1]
    currency = [row for row in issues if row[2] == "Валюта не гривня"]
    check("усі угоди в євро виключено", len(currency) == 5, len(currency))
    check("виключення позначене в таблиці",
          all(row[4] == "виключено" for row in currency), currency[:1])
    check("ринок порожній, а не від'ємний",
          dict(report.sections["Ринок"][0].tiles).get("Сума угод") == "0 грн")

    # Один постачальник на весь ринок — межа концентрації.
    single = Dataset(path=tmp / "монополія.xlsx",
                     tenders=[tender(i) for i in range(6)],
                     contracts=[contract(i, "12345678", 50000.0) for i in range(6)])
    report = insight.analyse(single, own_edrpou=["41263186"])
    tiles = dict(report.sections["Ринок"][0].tiles)
    check("HHI монополії — максимум", tiles["HHI"].startswith("10 000"), tiles["HHI"])
    check("CR3 монополії — 100%", tiles["CR3 / CR5 / CR10"] == "100% / 100% / 100%",
          tiles["CR3 / CR5 / CR10"])
    check("Джині монополії — нуль", tiles["Нерівність (Джині)"] == "0,000",
          tiles["Нерівність (Джині)"])

    # Одна-єдина угода: статистика викидів не має нічого «знаходити».
    lonely = Dataset(path=tmp / "одна.xlsx", tenders=[tender(0)],
                     contracts=[contract(0, "12345678", 100000.0)])
    report = insight.analyse(lonely, own_edrpou=["41263186"])
    dropped = dict(report.sections["Очищення"][0].tiles)["Виключено угод"]
    check("на одній угоді викидів немає", dropped.startswith("0"), dropped)


def test_degenerate_ui(tmp: Path) -> None:
    print("\n=== інтерфейс на вироджених даних ===")
    from PySide6.QtGui import QPixmap
    from PySide6.QtWidgets import QApplication

    from app.config import Settings
    from app.core import insight
    from app.core.xlsxload import Dataset
    from app.ui.pages.analytics_page import AnalyticsPage

    QApplication.instance() or QApplication([])
    page = AnalyticsPage(Settings())
    page.resize(1200, 800)

    report = insight.analyse(Dataset(path=tmp / "порожньо.xlsx"), own_edrpou=["41263186"])
    page.report = report
    page.show_report(report)
    check("порожній звіт показується", page.tabs.count() == 8, page.tabs.count())

    # Проходимо всі вкладки: кожна будується під час першого показу, і саме
    # там ловляться помилки в кресленні порожніх графіків.
    for index in range(page.tabs.count()):
        page.tabs.setCurrentIndex(index)
        pixmap = QPixmap(page.size())
        page.render(pixmap)
    check("усі вкладки порожнього звіту малюються", True)

    # Другий звіт поспіль: попередній має піти, а не накопичитись.
    page.show_report(report)
    check("повторний показ не множить вкладки", page.tabs.count() == 8, page.tabs.count())


def test_reader_fallback(tmp: Path) -> None:
    print("\n=== запасний читач ===")
    from app.core import xlsxload

    path = _demo_workbook(tmp)
    expected = xlsxload.load(path)

    class Broken:
        """Швидкий читач, який завжди відмовляє."""

        def __init__(self, *args, **kwargs):
            raise RuntimeError("книга не така")

    original = xlsxload.FastWorkbook
    xlsxload.FastWorkbook = Broken
    try:
        # Найважливіше в запасному шляху — що він непомітний: ті самі дані,
        # без винятку назовні.
        fallback = xlsxload.load(path)
    finally:
        xlsxload.FastWorkbook = original
    check("openpyxl підхоплює після відмови швидкого читача",
          fallback.tenders == expected.tenders
          and fallback.documents == expected.documents,
          f"{len(fallback.tenders)} / {len(expected.tenders)}")
    check("аркуші порахували однаково",
          fallback.counts() == expected.counts(), fallback.counts())


def test_deal_sources(tmp: Path) -> None:
    print("\n=== договір чи рішення про переможця ===")
    from app.core import insight
    from app.core.xlsxload import Dataset

    def tender(number: int):
        return {"tender_id": f"UA-{number}", "date": "2026-04-01", "title": "Ноутбуки",
                "description": "", "cpv_list": "30213100-6", "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 200000.0,
                "currency": "UAH", "vat": "так", "buyer": "Замовник",
                "buyer_edrpou": "04527520", "region": "Київська область",
                "locality": "Київ", "n_lots": 0, "n_bids": 1, "n_docs": 0,
                "tender_start": "", "tender_end": "", "modified": "", "url": ""}

    def contract(number, edrpou, amount, status="active", suffix="a1"):
        return {"tender_id": f"UA-{number}", "contract_id": f"UA-{number}-{suffix}",
                "name": "ТОВ Гравець", "edrpou": edrpou, "amount": amount,
                "currency": "UAH", "status": status, "signed": "2026-04-10",
                "buyer": "Замовник", "buyer_edrpou": "04527520",
                "region": "Київська область", "cpv_list": "30213100-6"}

    def award(number, edrpou, amount, status="active"):
        return {"tender_id": f"UA-{number}", "date": "2026-04-01", "name": "ТОВ Гравець",
                "edrpou": edrpou, "amount": amount, "currency": "UAH",
                "status": status, "decided": "2026-04-05"}

    data = Dataset(
        path=tmp / "джерела.xlsx",
        tenders=[tender(i) for i in range(4)],
        contracts=[
            contract(0, "11111111", 100000.0),                      # звичайний
            contract(1, "11111111", 60000.0, suffix="a1"),          # два лоти
            contract(1, "11111111", 40000.0, suffix="a2"),
            contract(2, "11111111", 90000.0, status="cancelled"),   # скасований
        ],
        awards=[
            award(0, "11111111", 100000.0),      # дубль договору — не рахувати
            award(2, "11111111", 95000.0),       # заміна скасованому договору
            award(3, "22222222", 70000.0),       # закупівля без договору
        ],
    )
    report = insight.analyse(data, own_edrpou=["41263186"])
    ranking = dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"]
    sums = {row[1]: row[6] for row in ranking[1]}

    # 100 000 за перший + 100 000 за два лоти другого + 95 000 за рішення
    # там, де договір скасовано. Дубль договору й рішення не подвоюється.
    check("договір і рішення не подвоюються", sums.get("11111111") == 295000.0,
          sums.get("11111111"))
    check("закупівля без договору теж рахується", sums.get("22222222") == 70000.0,
          sums.get("22222222"))
    total = dict(report.sections["Ринок"][0].tiles)["Сума угод"]
    check("сума ринку сходиться", total == "365,0 тис. грн", total)
    # Угод більше, ніж закупівель: у другій закупівлі два лоти, тобто два
    # окремі договори. Саме тому «кількість угод» і «кількість закупівель» —
    # різні показники, і плутати їх не можна.
    deals = dict(report.sections["Ринок"][0].tables)["Процедури"][1]
    check("угод п'ять на чотири закупівлі", sum(row[1] for row in deals) == 5,
          [row[1] for row in deals])
    ranking = dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"][1]
    lots = next(row for row in ranking if row[1] == "11111111")
    check("три закупівлі, чотири угоди в одного гравця",
          (lots[4], lots[5]) == (4, 3), (lots[4], lots[5]))


def test_cancel_midway(tmp: Path) -> None:
    print("\n=== зупинка посеред аналізу ===")
    import threading
    from app.core import insight, xlsxload

    data = xlsxload.load(_demo_workbook(tmp))
    stop = threading.Event()
    seen: list[str] = []

    def on_progress(stage, done, total):
        seen.append(stage)
        if done >= 2:              # зупиняємо вже після очищення
            stop.set()

    stopped = False
    try:
        insight.analyse(data, own_edrpou=["41263186"],
                        on_progress=on_progress, cancel_event=stop)
    except insight.Cancelled:
        stopped = True
    check("аналіз зупинився посеред роботи", stopped)
    check("до зупинки встиг зробити перші етапи", len(seen) >= 2, seen)
    check("а не всі вісім", len(seen) < 8, seen)


def test_export_names(tmp: Path) -> None:
    print("\n=== назви аркушів у вивантаженні ===")
    from openpyxl import load_workbook

    from app.core.report import Block, Profile, Report
    from app.core.reportbook import write_report

    report = Report()
    # Назва розділу із символами, яких Excel не пускає в назву аркуша.
    block = Block("Блок")
    block.tables.append(("Таблиця", (["Колонка"], [["значення"]])))
    report.add("Товари/ТМ: розбір [1]", block)
    # Порожні таблиці в книгу не потрапляють — інакше половина аркушів була б
    # порожньою.
    empty = Block("Порожній")
    empty.tables.append(("Без рядків", (["Колонка"], [])))
    report.add("Розділ", empty)

    # Два конкуренти з однаковою довгою назвою — класика реєстру: та сама
    # мережа під різними ЄДРПОУ.
    long_name = "Дуже довга назва компанії, яку Excel не подужає"
    people = []
    for edrpou in ("1", "2"):
        profile = Profile(edrpou=edrpou, name=long_name, block=Block("Портрет"))
        profile.block.tables.append(("Угоди", (["Сума, грн"], [[1]])))
        people.append(profile)
    report.competitors = people

    path = write_report(tmp / "назви.xlsx", report)
    names = load_workbook(path).sheetnames
    check("назви вкладаються в межу Excel", all(len(n) <= 31 for n in names), names)
    check("назви різні", len(set(names)) == len(names), names)
    check("заборонені символи прибрано",
          not any(set(n) & set("[]:*?/\\") for n in names), names)
    check("аркуш на зміст, огляд, розділ і кожного гравця", len(names) == 6, names)


def _demo_report():
    """Звіт з усіма типами діаграм — щоб перевірити книгу цілком."""
    from app.core.report import Block, ChartData, Profile, Report, Series

    block = Block("Ринок у цифрах", "Що видно з очищених даних.")
    block.tiles = [("Сума угод", "12,3 млн грн")]
    block.notes = ["Три гравці тримають 41% ринку."]
    block.charts = [
        ChartData("Топ постачальників", "hbar",
                  [Series("Сума угод", ["ТОВ А", "ТОВ Б"], [900.0, 750.0],
                          accent={1})], unit="грн"),
        ChartData("Закупівель по місяцях", "bar",
                  [Series("Закупівель", ["01", "02"], [10, 20])],
                  unit="шт", money_axis=False),
        ChartData("Процедури", "pie",
                  [Series("Сума", ["відкриті", "без торгів"], [5.0, 3.0])], unit="грн"),
        ChartData("Динаміка", "line",
                  [Series("План", ["01", "02"], [1.0, 2.0]),
                   Series("Факт", ["01", "02"], [1.0, 1.5])], unit="грн"),
        ChartData("Накопичено", "area",
                  [Series("Сума", ["01", "02"], [1.0, 3.0])], unit="грн"),
        ChartData("Дисконт", "hbar", [Series("Дисконт", ["ТОВ А"], [12.5])],
                  unit="%", money_axis=False),
        ChartData("Ціна проти кількості", "scatter",
                  [Series("Позиції",
                          points=[(float(i), float(i % 7)) for i in range(3000)],
                          accent={1, 2})], money_axis=False),
        ChartData("Порожній", "bar", [Series("Нічого", [], [])]),
    ]
    block.tables = [
        ("Рейтинг", (["Компанія", "Сума угод, грн", "Частка"],
                     [["ТОВ А", 900.0, 0.45], ["=ТОВ Б", 750.5, 0.37]])),
        ("Порожня", (["Колонка"], [])),
    ]
    report = Report(source="дані.xlsx", generated="2026-08-31 15:00",
                    period=("2026-01-01", "2026-08-31"))
    report.notes = ["Курс валют не перераховувався."]
    # «Підсумок» переносить до себе графіки з інших розділів — саме так це
    # робить `Analyzer.summary`, і саме через це оглядовий аркуш мусить
    # відсіювати повтори.
    digest = Block("Головне")
    digest.charts = [block.charts[0],
                     ChartData("Ми проти ринку", "hbar",
                               [Series("Сума", ["ТОВ А"], [900.0])], unit="грн")]
    report.add("Підсумок", digest)
    report.add("Ринок", block)

    profile = Profile(edrpou="1234", name="ТОВ А", is_ours=True, rank=1,
                      block=Block("Портрет"), strengths=["широкий портфель ТМ"])
    profile.block.charts.append(
        ChartData("Ключові замовники", "hbar",
                  [Series("Сума", ["Лікарня"], [10.0])], unit="грн"))
    profile.block.tables.append(("Угоди", (["Закупівля", "Сума, грн"], [["UA-1", 10.0]])))
    report.ours = [profile]
    return report


def test_report_overview(tmp: Path) -> None:
    print("\n=== оглядовий аркуш: усі графіки звіту в одному місці ===")
    from openpyxl import load_workbook

    from app.core.reportbook import write_report

    path = write_report(tmp / "огляд.xlsx", _demo_report())
    wb = load_workbook(path)
    check("аркуш «Графіки» стоїть одразу за змістом",
          wb.sheetnames[:2] == ["Зміст", "Графіки"], wb.sheetnames[:3])
    overview = wb["Графіки"]

    def title_of(chart) -> str:
        """Назва діаграми текстом: після читання книги вона — дерево об'єктів."""
        try:
            return "".join(run.t or "" for para in chart.title.tx.rich.p
                           for run in (para.r or ()))
        except AttributeError:
            return ""

    # «Топ постачальників» стоїть і в «Підсумку», і в «Ринку» — на огляді має
    # бути один раз.
    titles = [title_of(chart) for chart in overview._charts]
    check("зібрано всі графіки розділів", len(titles) == 8, titles)
    check("повтор між розділами відсіяно", len(set(titles)) == 8, sorted(titles))
    check("спільний графік узятий із першого розділу",
          titles.count("Топ постачальників") == 1, titles)
    check("порожній графік не потрапив і сюди", "Порожній" not in titles, titles)

    # Головне: аркуш не переписує до себе даних — інакше хмара на тисячу
    # рядків відсунула б усе наступне за обрій.
    numbers = [overview.cell(row=r, column=c).value
               for r in range(1, overview.max_row + 1)
               for c in range(1, 25)
               if isinstance(overview.cell(row=r, column=c).value, (int, float))]
    check("даних на аркуші немає — лише малюнки", numbers == [], numbers[:5])

    def sources(chart) -> set[str]:
        out = set()
        for item in chart.series:
            for holder in (item.cat, item.val, item.xVal, item.yVal):
                ref = getattr(holder, "numRef", None) or getattr(holder, "strRef", None)
                if ref and ref.f and "!" in ref.f:
                    out.add(ref.f.split("!")[0].strip("'"))
        return out

    used = {name for chart in overview._charts for name in sources(chart)}
    check("кожна діаграма читає дані свого розділу",
          used and "Графіки" not in used, sorted(used))
    check("розділи-джерела всі на місці", used == {"Підсумок", "Ринок"}, sorted(used))

    spots = [(chart.anchor._from.col, chart.anchor._from.row)
             for chart in overview._charts]
    check("діаграми стоять у дві колонки",
          {col for col, _row in spots} == {0, 12}, sorted({c for c, _r in spots}))
    check("жодна не лягла поверх іншої", len(set(spots)) == len(spots), spots)

    contents = [wb["Зміст"].cell(row=r, column=2).value
                for r in range(1, wb["Зміст"].max_row + 1)]
    check("огляд названо у змісті", "Усі графіки звіту" in contents, contents)
    wb.close()


def test_report_book(tmp: Path) -> None:
    print("\n=== книга звіту: діаграми праворуч від своїх даних ===")
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    from app.core.reportbook import SCATTER_CAP, write_report

    report = _demo_report()
    path = write_report(tmp / "звіт.xlsx", report)
    wb = load_workbook(path)
    market = wb["Ринок"]

    kinds = [type(chart).__name__ for chart in market._charts]
    check("кожен графік став діаграмою Excel", len(kinds) == 7, kinds)
    check("типи діаграм збережено",
          sorted(kinds) == ["AreaChart", "BarChart", "BarChart", "BarChart",
                            "LineChart", "PieChart", "ScatterChart"], sorted(kinds))
    check("порожній графік пропущено", "Порожній" not in
          [market.cell(row=r, column=1).value for r in range(1, market.max_row + 1)])
    check("смугова діаграма горизонтальна",
          [c.type for c in market._charts if type(c).__name__ == "BarChart"].count("bar") == 2,
          [getattr(c, "type", "") for c in market._charts])

    def used_columns(chart) -> int:
        """Найправіша колонка, з якої діаграма бере числа."""
        right = 0
        for item in chart.series:
            for holder in (item.cat, item.val, item.xVal, item.yVal):
                ref = getattr(holder, "numRef", None) or getattr(holder, "strRef", None)
                if not ref or not ref.f:
                    continue
                for cell in ref.f.split("!")[-1].replace("$", "").split(":"):
                    letters = "".join(ch for ch in cell if ch.isalpha())
                    right = max(right, column_index_from_string(letters))
        return right

    placed = [(chart.anchor._from.col + 1, used_columns(chart))
              for chart in market._charts]
    check("діаграма стоїть праворуч від своїх даних",
          all(anchor > data for anchor, data in placed), placed)

    # Таблиця даних під діаграмою — та сама, з якої вона намальована.
    labels = {market.cell(row=r, column=1).value for r in range(1, market.max_row + 1)}
    check("дані діаграми лежать поруч у книзі", {"ТОВ А", "ТОВ Б"} <= labels)
    percent = next(market.cell(row=r, column=2) for r in range(1, market.max_row + 1)
                   if market.cell(row=r, column=2).value == 12.5)
    check("відсотки графіка не множаться вдруге",
          percent.number_format == '0.0"%"', percent.number_format)

    excel_table = next(iter(market.tables.values()))
    table = excel_table.ref
    top = int("".join(ch for ch in table.split(":")[0] if ch.isdigit())) + 1
    # Стрілки відбору openpyxl додає лише тоді, коли сам вигадує колонки;
    # ми описуємо колонки самі, тож і фільтр маємо ставити самі.
    check("у таблиці є стрілки відбору",
          excel_table.autoFilter is not None and excel_table.autoFilter.ref == table,
          excel_table.autoFilter)
    check("колонки таблиці названі як заголовки",
          [c.name for c in excel_table.tableColumns] == ["Компанія", "Сума угод, грн",
                                                         "Частка"],
          [c.name for c in excel_table.tableColumns])
    check("частка в таблиці — відсотком",
          market.cell(row=top, column=3).number_format == "0.0%",
          market.cell(row=top, column=3).number_format)
    check("сума з копійками там, де вони є",
          market.cell(row=top, column=2).number_format == "#,##0.00",
          market.cell(row=top, column=2).number_format)
    check("текст із «=» не став формулою",
          market.cell(row=top + 1, column=1).value == "=ТОВ Б",
          market.cell(row=top + 1, column=1).value)

    scatter = next(chart for chart in market._charts
                   if type(chart).__name__ == "ScatterChart")
    span = scatter.series[0].xVal.numRef.f.split("!")[-1].replace("$", "").split(":")
    written = int("".join(ch for ch in span[1] if ch.isdigit())) - \
        int("".join(ch for ch in span[0] if ch.isdigit())) + 1
    check("хмару проріджено до межі", written <= SCATTER_CAP, written)

    check("портрет гравця став аркушем", "Ми1 ТОВ А" in wb.sheetnames, wb.sheetnames)
    check("порожня таблиця не потрапила в книгу",
          sum(len(ws.tables) for ws in wb.worksheets) == 2,
          {ws.title: len(ws.tables) for ws in wb.worksheets})
    ours = wb["Ми1 ТОВ А"]
    column = [ours.cell(row=r, column=1).value for r in range(1, ours.max_row + 1)]
    check("сильні сторони перенесені", "+  широкий портфель ТМ" in column, column)
    check("зміст веде на аркуші",
          any(str(wb["Зміст"].cell(row=r, column=3).value or "").startswith("=HYPERLINK")
              for r in range(1, wb["Зміст"].max_row + 1)))
    wb.close()


def test_profile_book(tmp: Path) -> None:
    print("\n=== книга по одній компанії ===")
    from openpyxl import load_workbook

    from app.core.report import Profile
    from app.core.reportbook import write_profile_report

    report = _demo_report()
    # Назва довша за межу Excel і з символами, яких він не пускає в аркуш.
    profile = report.ours[0]
    profile.name = "ТОВ «ТОРГОВИЙ ДІМ/ПАРТНЕР: АЙ ТІ» та інші дуже довгі слова"
    profile.weaknesses = ["вузька географія"]
    path = write_profile_report(tmp / "компанія.xlsx", report, profile)
    wb = load_workbook(path)

    check("у книзі один аркуш — сама компанія", len(wb.sheetnames) == 1, wb.sheetnames)
    check("назва аркуша прийнятна для Excel",
          len(wb.sheetnames[0]) <= 31 and not set(wb.sheetnames[0]) & set("[]:*?/\\"),
          wb.sheetnames)
    ws = wb.worksheets[0]
    column = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    want_charts = sum(1 for c in profile.block.charts if c.series)
    check("діаграми компанії на місці", len(ws._charts) == want_charts,
          f"{len(ws._charts)} з {want_charts}")
    check("таблиці компанії на місці", len(ws.tables) == len(profile.block.tables),
          f"{len(ws.tables)} з {len(profile.block.tables)}")
    check("сильні та слабкі сторони перенесені",
          "+  широкий портфель ТМ" in column and "−  вузька географія" in column, column)
    check("видно, звідки й за який період дані",
          any("дані.xlsx" in str(v or "") and "2026-01-01" in str(v or "")
              for v in column), column[:4])
    check("аркуша «Зміст» тут немає", "Зміст" not in wb.sheetnames, wb.sheetnames)
    wb.close()

    # Гравець без портрета трапляється: у рейтингу він є, сторінки в нього немає.
    empty = write_profile_report(tmp / "без-блоку.xlsx", report,
                                 Profile(edrpou="9999", name="Без блоку"))
    check("гравець без портрета не валить запис", empty.exists())


def test_profile_export_button(tmp: Path) -> None:
    print("\n=== кнопка «Зберегти звіт по компанії» ===")
    from openpyxl import load_workbook
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox, QPushButton

    from app.config import Settings
    from app.ui.pages import analytics_page as page_module
    from app.ui.pages.analytics_page import AnalyticsPage

    QApplication.instance() or QApplication([])
    report = _demo_report()
    page = AnalyticsPage(Settings())
    page.report = report
    page.show_report(report)

    profile = report.ours[0]
    widget = page._profile_widget(profile)
    buttons = [b for b in widget.findChildren(QPushButton) if "компані" in b.text()]
    check("на сторінці компанії є кнопка звіту", len(buttons) == 1,
          [b.text() for b in widget.findChildren(QPushButton)])

    saved = tmp / "звіт-по-компанії.xlsx"
    shown: list[str] = []

    # Діалоги в тесті нікому показувати: підміняємо їх на місці, у просторі
    # імен сторінки, і повертаємо назад одразу після натискання.
    class Dialog:
        @staticmethod
        def getSaveFileName(*_args, **_kw):
            return str(saved), ""

    class Box:
        @staticmethod
        def information(_parent, _title, text=""):
            shown.append(text)

        @staticmethod
        def critical(_parent, _title, text=""):
            shown.append("ПОМИЛКА: " + text)

    page_module.QFileDialog, page_module.QMessageBox = Dialog, Box
    try:
        buttons[0].click()
    finally:
        page_module.QFileDialog, page_module.QMessageBox = QFileDialog, QMessageBox

    check("натискання зберегло книгу", saved.exists(), shown[:1])
    check("під час запису не було помилок",
          not any(s.startswith("ПОМИЛКА") for s in shown), shown)
    if saved.exists():
        book = load_workbook(saved)
        check("у книзі аркуш саме цієї компанії", len(book.sheetnames) == 1,
              book.sheetnames)
        book.close()


def test_number_formats() -> None:
    print("\n=== межі форматування чисел ===")
    from app.core.report import compact, count, median, money, pct, robust_z

    check("від'ємна сума", money(-1234.5, 2) == "-1 234,50", money(-1234.5, 2))
    check("нуль", money(0) == "0", money(0))
    check("мільярд", compact(2_500_000_000) == "2,5 млрд", compact(2_500_000_000))
    check("не число", money("хтозна") == "—", money("хтозна"))
    check("нескінченність", money(float("inf")) == "—", money(float("inf")))
    check("NaN", money(float("nan")) == "—", money(float("nan")))
    check("від'ємний відсоток", pct(-0.075) == "-7,5%", pct(-0.075))
    check("кількість із розрядами", count(1234567) == "1 234 567", count(1234567))
    check("кількість не з числа", count(None) == "—", count(None))

    # Статистика не повинна ламатися на сміттєвих значеннях.
    check("медіана ігнорує нечислове", median([1, "два", 3, None]) == 2.0,
          median([1, "два", 3, None]))
    check("z-показник на сміттєвому ряду",
          robust_z([1, "два", 3, None, 5, 7, 9, 11])[1] == 0.0)
    check("z-показник на однакових значеннях", max(robust_z([7] * 8)) == 0.0)


def _rivalry_dataset(tmp: Path):
    """Вибірка, у якій є всі способи програти закупівлю.

    Три причини поразки означають три різні проблеми, і плутати їх не можна:
    нас не було на подачі — це моніторинг; нас перебили ціною — це ціна; наша
    ціна була нижчою, а закупівля пішла не до нас — це підготовка пропозиції.
    """
    from app.core.xlsxload import Dataset

    OURS, RIVAL, OTHER = "41263186", "12345678", "99999999"

    def tender(number: int, cpv: str = "30213100-6"):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01", "title": "Ноутбуки",
                "description": "", "cpv_list": cpv, "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 150000.0,
                "currency": "UAH", "vat": "так", "buyer": f"Замовник {number}",
                "buyer_edrpou": f"0452752{number}", "region": "Київська область",
                "locality": "Київ", "n_lots": 0, "n_bids": 2, "n_docs": 0,
                "tender_start": "", "tender_end": "", "modified": "", "url": ""}

    def bid(number, edrpou, amount):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01",
                "name": f"ТОВ {edrpou}", "edrpou": edrpou, "region": "Київська область",
                "amount": amount, "currency": "UAH", "status": "active",
                "submitted": "2026-05-02"}

    def contract(number, edrpou, amount):
        return {"tender_id": f"UA-{number}", "contract_id": f"UA-{number}-a1",
                "name": f"ТОВ {edrpou}", "edrpou": edrpou, "amount": amount,
                "currency": "UAH", "status": "active", "signed": "2026-05-10",
                "buyer": f"Замовник {number}", "buyer_edrpou": f"0452752{number}",
                "region": "Київська область", "cpv_list": "30213100-6"}

    return Dataset(
        path=tmp / "суперництво.xlsx",
        tenders=[tender(i) for i in range(1, 6)],
        bids=[
            bid(1, OURS, 110000.0), bid(1, RIVAL, 100000.0),   # перебили ціною
            bid(2, OURS, 90000.0), bid(2, RIVAL, 100000.0),    # ціна була наша
            bid(3, RIVAL, 100000.0),                           # нас не було
            bid(4, OURS, 80000.0), bid(4, RIVAL, 95000.0),     # виграли ми
            bid(5, OTHER, 120000.0),                           # повз нас узагалі
        ],
        contracts=[
            contract(1, RIVAL, 100000.0),
            contract(2, RIVAL, 100000.0),
            contract(3, RIVAL, 100000.0),
            contract(4, OURS, 80000.0),
            contract(5, OTHER, 120000.0),
        ],
    ), OURS, RIVAL


def test_loss_reasons(tmp: Path) -> None:
    print("\n=== чому програли ===")
    from app.core import insight

    data, ours, rival = _rivalry_dataset(tmp)
    report = insight.analyse(data, own_edrpou=[ours])
    profile = next(p for p in report.competitors if p.edrpou == rival)

    losses = dict(profile.block.tables)["Наші зустрічі"]
    verdicts = {row[0]: row[7] for row in losses[1]}
    check("перебили ціною", verdicts.get("UA-1") == "програли за ціною", verdicts)
    check("ціна була наша, а закупівля — ні",
          verdicts.get("UA-2") == "ціна була нижча — програли не за ціною", verdicts)
    check("нас не було на подачі", verdicts.get("UA-3") == "ми не подавалися", verdicts)
    check("виграну нами закупівлю сюди не тягнемо", "UA-4" not in verdicts, verdicts)

    gaps = {row[0]: row[6] for row in losses[1]}
    check("розрив ціни там, де програли", gaps.get("UA-1") == 0.1, gaps.get("UA-1"))
    check("від'ємний розрив там, де наша була нижчою",
          gaps.get("UA-2") == -0.1, gaps.get("UA-2"))

    summary = dict(report.sections["Конкуренти"][0].tables)["Чому програли"][1]
    row = next(r for r in summary if r[0] == rival)
    check("у зведенні: перемог у нього три", row[2] == 3, row)
    check("у зведенні: не подавалися — одна", row[3] == 1, row)
    check("у зведенні: програли за ціною — одна", row[4] == 1, row)
    check("у зведенні: програли не за ціною — одна", row[5] == 1, row)


def test_head_to_head(tmp: Path) -> None:
    print("\n=== очні зустрічі та охоплення ===")
    from app.core import insight

    data, ours, rival = _rivalry_dataset(tmp)
    report = insight.analyse(data, own_edrpou=[ours])
    block = report.sections["Наші ТОВ"][0]

    meetings = dict(block.tables)["Очні зустрічі"][1]
    row = next(r for r in meetings if r[0] == rival)
    # Спільними вважаємо ті закупівлі, де подавалися обидва: UA-1, UA-2, UA-4.
    check("спільних закупівель три", row[2] == 3, row)
    check("одну виграли ми", row[3] == 1, row)
    check("дві виграв конкурент", row[4] == 2, row)
    check("сума його перемог у зустрічах", row[6] == 200000.0, row)

    coverage = dict(block.tables)["Пропущені закупівлі"][1]
    missed = {r[0] for r in coverage}
    # UA-3 і UA-5 — у нашій галузі, але без нашої участі.
    check("пропущені закупівлі знайдено", missed == {"UA-3", "UA-5"}, missed)
    note = " ".join(block.notes)
    check("охоплення пораховано", "60,0%" in note, note)


def test_column_matching() -> None:
    print("\n=== зіставлення колонок ===")
    from app.core.xlsxload import TABLES, _match_columns

    spec = TABLES["contracts"]["columns"]
    headers = ["Номер закупівлі", "Номер договору", "Постачальник",
               "ЄДРПОУ постачальника", "Сума договору", "Валюта", "Статус договору",
               "Дата підписання", "Замовник", "ЄДРПОУ замовника", "Регіон", "Коди ДК021"]
    found = _match_columns(headers, spec)
    check("усі колонки знайдено", len(found) == len(spec), sorted(set(spec) - set(found)))
    check("постачальник і його код не переплутані",
          (found["name"], found["edrpou"]) == (2, 3), (found["name"], found["edrpou"]))
    check("замовник і його код теж",
          (found["buyer"], found["buyer_edrpou"]) == (8, 9),
          (found["buyer"], found["buyer_edrpou"]))

    # Колонки переставлені, додано зайві, регістр і пробіли інші.
    shuffled = ["зайва", "  СУМА ДОГОВОРУ  ", "Номер закупівлі", "ще одна",
                "ЄДРПОУ постачальника", "Постачальник"]
    found = _match_columns(shuffled, spec)
    check("порядок колонок не важливий", found.get("tender_id") == 2, found.get("tender_id"))
    check("регістр і зайві пробіли не заважають", found.get("amount") == 1,
          found.get("amount"))
    check("зайві колонки не привласнюються",
          found.get("name") == 5 and found.get("edrpou") == 4,
          (found.get("name"), found.get("edrpou")))
    check("чого немає — того немає", "currency" not in found, found.get("currency"))

    # Одна колонка не може дістатися двом полям одразу.
    check("жодна колонка не зайнята двічі",
          len(set(found.values())) == len(found), found)


def test_settings_robustness(tmp: Path) -> None:
    print("\n=== стійкість налаштувань ===")
    import json

    from app.config import DEFAULT_COMPETITORS, DEFAULT_OWN_EDRPOU, Settings

    missing = Settings.load(tmp / "немає-такого.json")
    check("без файлу беруться типові", missing.own_edrpou == DEFAULT_OWN_EDRPOU,
          missing.own_edrpou)
    check("і конкуренти теж", missing.competitors == DEFAULT_COMPETITORS,
          missing.competitors)

    broken = tmp / "зіпсовані.json"
    broken.write_text("{це не json", encoding="utf-8")
    check("зіпсований файл не валить запуск",
          Settings.load(broken).competitors == DEFAULT_COMPETITORS)

    # Файл налаштувань тепер лежить у репозиторії, тож на іншій машині в ньому
    # буде чужий шлях до теки завантажень. Програма має це пережити.
    alien = tmp / "чужі.json"
    alien.write_text(json.dumps({"output_dir": "Z:\\\\немає\\\\теки",
                                 "own_edrpou": ["41263186"]}, ensure_ascii=False),
                     encoding="utf-8")
    settings = Settings.load(alien)
    check("чужий шлях читається як є", settings.output_dir.startswith("Z:"),
          settings.output_dir)
    from app.core.xlsxload import find_workbooks
    check("неіснуюча тека дає порожній перелік, а не помилку",
          find_workbooks(settings.output_dir) == [])

    # Часткові налаштування: бракує майже всіх ключів.
    partial = tmp / "частково.json"
    partial.write_text('{"theme": "light"}', encoding="utf-8")
    settings = Settings.load(partial)
    check("бракуючі ключі добираються типовими",
          settings.theme == "light" and settings.competitors == DEFAULT_COMPETITORS,
          (settings.theme, settings.competitors))

    # Збереження й читання назад мають давати те саме.
    roundtrip = tmp / "туди-назад.json"
    original = Settings()
    original.competitors = ["24083083", "40733799"]
    original.save(roundtrip)
    restored = Settings.load(roundtrip)
    check("збережене читається назад", restored.competitors == original.competitors,
          restored.competitors)
    check("пресет теж переживає збереження",
          restored.preset.cpv_prefixes == original.preset.cpv_prefixes,
          restored.preset.cpv_prefixes)


def test_own_tm_verdict(tmp: Path) -> None:
    print("\n=== власна ТМ у каналах постачання ===")
    from app.core import brands, insight
    from app.core.xlsxload import Dataset

    # Довідник знає власників трьох марок — перевіряємо, що аналіз справді
    # робить із цього висновок, а не лише зберігає запис.
    owners = brands.own_tm()
    check("Artline має власника", owners["Artline"]["owner"] == "АРТЛАЙН ІНТЕГРАЦІЯ",
          owners["Artline"])
    check("код власника теж", owners["Artline"]["edrpou"] == "40733799",
          owners["Artline"])

    def tender(number, title):
        return {"tender_id": f"UA-{number}", "date": "2026-06-01", "title": title,
                "description": "", "cpv_list": "30213100-6", "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 100000.0,
                "currency": "UAH", "vat": "так", "buyer": "Замовник",
                "buyer_edrpou": "04527520", "region": "Київська область",
                "locality": "Київ", "n_lots": 0, "n_bids": 1, "n_docs": 0,
                "tender_start": "", "tender_end": "", "modified": "", "url": ""}

    def item(number, description):
        return {"tender_id": f"UA-{number}", "date": "2026-06-01",
                "description": description, "cpv": "30213100-6", "cpv_name": "Ноутбуки",
                "quantity": 5.0, "unit": "штука", "lot_value": 100000.0,
                "unit_price": 20000.0, "buyer": "Замовник", "buyer_edrpou": "04527520",
                "region": "Київська область"}

    def contract(number, edrpou):
        return {"tender_id": f"UA-{number}", "contract_id": f"UA-{number}-a1",
                "name": "ТОВ АРТЛАЙН ІНТЕГРАЦІЯ", "edrpou": edrpou, "amount": 100000.0,
                "currency": "UAH", "status": "active", "signed": "2026-06-10",
                "buyer": "Замовник", "buyer_edrpou": "04527520",
                "region": "Київська область", "cpv_list": "30213100-6"}

    data = Dataset(
        path=tmp / "власна-тм.xlsx",
        tenders=[tender(i, "Ноутбуки") for i in range(4)],
        items=[item(i, "Ноутбук Artline Business B57") for i in range(4)],
        contracts=[contract(i, "40733799") for i in range(4)],
    )
    report = insight.analyse(data, own_edrpou=["41263186"], tracked=["40733799"])
    supply = dict(report.sections["Постачання"][0].tables)["Канали постачання"][1]
    row = next(r for r in supply if r[0] == "40733799")
    check("ТМ розпізнана в позиціях", row[3] == "Artline", row[3])
    check("довідник підставив власника", row[5].startswith("Artline"), row[5])
    check("висновок — власна ТМ", "власна ТМ" in row[10], row[10])

    profile = next(p for p in report.competitors if p.edrpou == "40733799")
    check("у портреті теж видно власну ТМ",
          any("власну ТМ" in trait for trait in profile.traits), profile.traits)


def _market_dataset(tmp: Path, tenders_spec):
    """Ринок із заданих трійок «очікувана вартість, сума договору, учасники»."""
    from app.core.xlsxload import Dataset

    tenders, contracts, bids = [], [], []
    for i, (expected, signed, players) in enumerate(tenders_spec):
        tid = f"UA-{i:04d}"
        tenders.append({
            "tender_id": tid, "date": f"2026-0{i % 6 + 1}-15", "title": "Ноутбуки",
            "description": "", "cpv_list": "30213100-6", "status": "Завершена",
            "method": "Відкриті торги" if players > 1 else "Звіт про договір",
            "category": "goods", "value": float(expected), "currency": "UAH",
            "vat": "так", "buyer": f"Замовник {i % 7}", "buyer_edrpou": f"0452752{i % 7}",
            "region": "Київська область", "locality": "Київ", "n_lots": 0,
            "n_bids": players, "n_docs": 0, "tender_start": "", "tender_end": "",
            "modified": "", "url": ""})
        contracts.append({
            "tender_id": tid, "contract_id": f"{tid}-a1", "name": f"ТОВ {i % 5}",
            "edrpou": f"1234567{i % 5}", "amount": float(signed), "currency": "UAH",
            "status": "active", "signed": f"2026-0{i % 6 + 1}-25",
            "buyer": f"Замовник {i % 7}", "buyer_edrpou": f"0452752{i % 7}",
            "region": "Київська область", "cpv_list": "30213100-6"})
        for k in range(players):
            bids.append({
                "tender_id": tid, "date": f"2026-0{i % 6 + 1}-15",
                "name": f"ТОВ {k}", "edrpou": f"9999999{k}",
                "region": "Київська область", "amount": float(signed) * (1 + 0.05 * k),
                "currency": "UAH", "status": "active",
                "submitted": f"2026-0{i % 6 + 1}-16"})
    return Dataset(path=tmp / "ринок.xlsx", tenders=tenders, contracts=contracts, bids=bids)


def test_big_deals_survive(tmp: Path) -> None:
    print("\n=== велика закупівля — не викид ===")
    from app.core import insight

    # Ринок як у житті: сотні дрібних прямих договорів і кілька великих
    # закупівель. На реальних даних саме тут статистика викидала третину
    # грошей, бо поріг 3,5σ від медіани в 10 тис. грн оголошував помилкою
    # кожен договір понад 18 млн.
    spec = [(20_000, 20_000, 1) for _ in range(400)]
    spec += [(196_000_000, 196_000_000, 1),   # збігається з планом до копійки
             (98_000_000, 91_000_000, 3),     # трохи дешевше плану
             (66_000_000, 69_000_000, 2)]     # трохи дорожче плану
    data = _market_dataset(tmp, spec)
    report = insight.analyse(data, own_edrpou=["41263186"])

    total = sum(row[6] for row in
                dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"][1])
    expected_total = 400 * 20_000 + 196_000_000 + 91_000_000 + 69_000_000
    check("жодної гривні не загублено", round(total) == expected_total,
          f"{total:,.0f} проти {expected_total:,.0f}".replace(",", " "))

    tiles = dict(report.sections["Очищення"][0].tiles)
    check("нічого не виключено", tiles["Виключено угод"].startswith("0"),
          tiles["Виключено угод"])
    notes = " ".join(report.sections["Очищення"][0].notes)
    check("сказано, що суми підтверджені", "підтверджена очікуваною вартістю" in notes)

    # А ось коли підтвердити нічим — статистика має спрацювати.
    loose = _market_dataset(tmp, [(0, 20_000, 1) for _ in range(400)]
                            + [(0, 500_000_000, 1)])
    report = insight.analyse(loose, own_edrpou=["41263186"])
    issues = dict(report.sections["Очищення"][0].tables)["Зауваження до даних"][1]
    caught = [r for r in issues if r[2] == "Нетипова сума для галузі"]
    check("без очікуваної вартості викид ловиться", len(caught) == 1, len(caught))

    # І коли сума суперечить плану — теж.
    wrong = _market_dataset(tmp, [(20_000, 20_000, 1) for _ in range(10)]
                            + [(1_000_000, 50_000_000, 1)])
    report = insight.analyse(wrong, own_edrpou=["41263186"])
    issues = dict(report.sections["Очищення"][0].tables)["Зауваження до даних"][1]
    check("суперечність із планом ловиться",
          any("значно більший" in r[2] for r in issues),
          [r[2] for r in issues])


def test_multilot_not_penalised(tmp: Path) -> None:
    print("\n=== багатолотова закупівля ===")
    from app.core import insight
    from app.core.xlsxload import Dataset

    # Одна закупівля на 10 млн, розбита на 20 лотів по 500 тис. Кожен лот
    # окремо — 5% від очікуваної вартості, тобто рівно на межі «мізерного».
    tender = {"tender_id": "UA-1", "date": "2026-05-01", "title": "Ноутбуки",
              "description": "", "cpv_list": "30213100-6", "status": "Завершена",
              "method": "Відкриті торги", "category": "goods", "value": 10_000_000.0,
              "currency": "UAH", "vat": "так", "buyer": "Замовник",
              "buyer_edrpou": "04527520", "region": "Київська область",
              "locality": "Київ", "n_lots": 20, "n_bids": 2, "n_docs": 0,
              "tender_start": "", "tender_end": "", "modified": "", "url": ""}
    contracts = [{"tender_id": "UA-1", "contract_id": f"UA-1-a{k}", "name": "ТОВ Лот",
                  "edrpou": "12345678", "amount": 500_000.0, "currency": "UAH",
                  "status": "active", "signed": "2026-05-20", "buyer": "Замовник",
                  "buyer_edrpou": "04527520", "region": "Київська область",
                  "cpv_list": "30213100-6"} for k in range(20)]
    data = Dataset(path=tmp / "лоти.xlsx", tenders=[tender], contracts=contracts)
    report = insight.analyse(data, own_edrpou=["41263186"])
    ranking = dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"][1]
    check("усі 20 лотів у підсумку", ranking and ranking[0][6] == 10_000_000.0,
          ranking[0][6] if ranking else None)
    check("двадцять угод, одна закупівля",
          ranking and (ranking[0][4], ranking[0][5]) == (20, 1),
          (ranking[0][4], ranking[0][5]) if ranking else None)


def test_savings_by_competition(tmp: Path) -> None:
    print("\n=== економія та конкуренція ===")
    from app.core import insight

    # 100 прямих договорів рівно за планом і 20 конкурентних із економією 20%.
    spec = [(100_000, 100_000, 1) for _ in range(100)]
    spec += [(100_000, 80_000, 3) for _ in range(20)]
    report = insight.analyse(_market_dataset(tmp, spec), own_edrpou=["41263186"])
    tiles = dict(report.sections["Ринок"][0].tiles)
    # Медіана по всій вибірці була б 0% — і ховала б усю економію ринку.
    check("економія рахується на конкурентних торгах",
          tiles["Економія на торгах"] == "20,0%", tiles["Економія на торгах"])
    check("зекономлені гроші показано",
          tiles["Зекономлено на ринку"].startswith("400,0 тис"),
          tiles["Зекономлено на ринку"])

    table = dict(dict(report.sections["Ринок"][0].tables))["Економія та конкуренція"][1]
    rows = {r[0].split(" (")[0]: r for r in table}
    check("прямі договори — нульова економія", rows["Без конкуренції"][6] == 0.0,
          rows["Без конкуренції"])
    check("конкурентні — двадцять відсотків", rows["Троє і більше"][6] == 0.2,
          rows["Троє і більше"])
    check("зважена економія теж", rows["Троє і більше"][5] == 0.2, rows["Троє і більше"])


def test_monthly_one_axis(tmp: Path) -> None:
    print("\n=== один часовий ряд для плану й факту ===")
    from app.core import insight
    from app.core.xlsxload import Dataset

    # Закупівля оприлюднена в березні, договір підписаний у липні.
    tender = {"tender_id": "UA-1", "date": "2026-03-10", "title": "Ноутбуки",
              "description": "", "cpv_list": "30213100-6", "status": "Завершена",
              "method": "Відкриті торги", "category": "goods", "value": 1_000_000.0,
              "currency": "UAH", "vat": "так", "buyer": "Замовник",
              "buyer_edrpou": "04527520", "region": "Київська область",
              "locality": "Київ", "n_lots": 0, "n_bids": 2, "n_docs": 0,
              "tender_start": "", "tender_end": "", "modified": "", "url": ""}
    contract = {"tender_id": "UA-1", "contract_id": "UA-1-a1", "name": "ТОВ",
                "edrpou": "12345678", "amount": 900_000.0, "currency": "UAH",
                "status": "active", "signed": "2026-07-05", "buyer": "Замовник",
                "buyer_edrpou": "04527520", "region": "Київська область",
                "cpv_list": "30213100-6"}
    data = Dataset(path=tmp / "місяці.xlsx", tenders=[tender], contracts=[contract])
    report = insight.analyse(data, own_edrpou=["41263186"])
    rows = {r[0].rstrip(" *"): r for r in
            dict(report.sections["Ринок"][0].tables)["Динаміка по місяцях"][1]}
    # План і факт мають стояти в тому самому місяці — інакше березень виглядав
    # би як «план без грошей», а липень як «гроші без плану».
    check("план і факт в одному місяці",
          rows["2026-03"][2] == 1_000_000.0 and rows["2026-03"][3] == 900_000.0,
          (rows["2026-03"][2], rows["2026-03"][3]))
    check("дата підпису — окремою колонкою",
          rows["2026-07"][4] == 900_000.0 and rows["2026-03"][4] == 0.0,
          (rows["2026-03"][4], rows["2026-07"][4]))
    # Неповноту місяця таблиця показує окремою колонкою, а графік — зірочкою
    # в підписі: у таблиці для цього є місце, у підпису осі — ні.
    table = dict(report.sections["Ринок"][0].tables)["Динаміка по місяцях"][1]
    check("у таблиці є колонка про повноту місяця",
          all(row[1] in ("так", "") for row in table), [row[:2] for row in table])
    check("обидва крайні місяці неповні",
          all(row[1] == "" for row in table), [row[:2] for row in table])
    chart = next(c for c in report.sections["Ринок"][0].charts
                 if c.title.startswith("Динаміка"))
    check("на графіку неповні місяці із зірочкою",
          all(label.endswith(" *") for label in chart.series[0].labels),
          chart.series[0].labels)
    check("у підказці пояснено, що означає зірочка",
          "неповні місяці" in chart.hint, chart.hint)


def test_foreign_currency_reported(tmp: Path) -> None:
    print("\n=== угоди в чужій валюті ===")
    from app.core import insight
    from app.core.xlsxload import Dataset

    tenders, contracts = [], []
    for i, currency in enumerate(["UAH", "UAH", "USD", "EUR"]):
        tid = f"UA-{i}"
        tenders.append({"tender_id": tid, "date": "2026-05-01", "title": "Ноутбуки",
                        "description": "", "cpv_list": "30213100-6",
                        "status": "Завершена", "method": "Відкриті торги",
                        "category": "goods", "value": 100_000.0, "currency": currency,
                        "vat": "так", "buyer": "Замовник", "buyer_edrpou": "04527520",
                        "region": "Київська область", "locality": "Київ", "n_lots": 0,
                        "n_bids": 1, "n_docs": 0, "tender_start": "", "tender_end": "",
                        "modified": "", "url": ""})
        contracts.append({"tender_id": tid, "contract_id": f"{tid}-a1", "name": "ТОВ",
                          "edrpou": "12345678", "amount": 100_000.0,
                          "currency": currency, "status": "active",
                          "signed": "2026-05-10", "buyer": "Замовник",
                          "buyer_edrpou": "04527520", "region": "Київська область",
                          "cpv_list": "30213100-6"})
    data = Dataset(path=tmp / "валюта.xlsx", tenders=tenders, contracts=contracts)
    report = insight.analyse(data, own_edrpou=["41263186"])
    block = report.sections["Очищення"][0]
    note = " ".join(block.notes)
    check("про валютні угоди сказано", "іноземній валюті" in note, note[-120:])
    check("суми за валютами названо", "100 000 USD" in note and "100 000 EUR" in note, note)
    sheet = dict(block.tables).get("Угоди в іноземній валюті")
    check("їх перелічено окремою таблицею", sheet and len(sheet[1]) == 2,
          len(sheet[1]) if sheet else None)
    # І вони справді не потрапили в гроші ринку.
    ranking = dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"][1]
    check("у підсумок ринку не входять", ranking[0][6] == 200_000.0, ranking[0][6])


def test_fresh_start(tmp: Path) -> None:
    print("\n=== чистий аркуш перед збором ===")
    from app.config import Settings
    from app.core.db import Database

    check("очищення ввімкнене типово", Settings().fresh_start is True)

    db = Database(tmp / "чистка.db")
    # Індекс здобувається дорого — він має пережити очищення.
    db.index_put([("UA-2026-000001-a", "uuid-1", "2026-05-01", "2026-05-02",
                   "complete", "aboveThreshold", "Замовник", "04527520", "Київська")])
    db.coverage_mark("2026-05-01", 17000, True)
    db.save_tender(
        {"uuid": "uuid-1", "tender_id": "UA-2026-000001-a", "title": "Ноутбуки",
         "description": "", "status": "complete", "method_type": "aboveThreshold",
         "main_category": "goods", "date_created": "2026-05-01", "date_modified": "",
         "tender_start": "", "tender_end": "", "value_amount": 100000.0,
         "value_currency": "UAH", "vat_included": 1, "pe_name": "Замовник",
         "pe_edrpou": "04527520", "pe_region": "Київська", "pe_locality": "Київ",
         "n_lots": 0, "n_bids": 1, "n_docs": 1, "cpv_list": "30213100-6",
         "fetched_at": "2026-05-02"},
        lots=[], items=[("i1", "uuid-1", None, "Ноутбук", "30213100-6", "Ноутбуки", 5, "шт")],
        bids=[], awards=[],
        contracts=[("c1", "uuid-1", "UA-1-a1", "active", "2026-05-10", 90000.0,
                    "UAH", "ТОВ", "12345678")],
        docs=[("k1", "d1", "uuid-1", "UA-2026-000001-a", "tender", "documents", "", "",
               "", "умови.pdf", "application/pdf", "2026-05-01", "https://x/1", "",
               0, "", "pending", "")])
    db.save_product({"id": "p1", "title": "Ноутбук", "brand": "Vinga", "description": "",
                     "category": "", "cpv": "30213100-6", "cpv_name": "", "barcode": "",
                     "barcode_scheme": "", "status": "active", "marketplace": "",
                     "vendor": "", "price_low": 1.0, "price_high": 2.0,
                     "price_currency": "UAH", "price_vat": 1, "price_date": "",
                     "n_images": 0, "images": "", "n_specs": 1, "description_len": 0,
                     "date_created": "", "date_modified": "", "expiration_date": "",
                     "url": ""},
                    [("p1", "Бренд", "Vinga", None, "")])

    before = {t: db.scalar(f"SELECT COUNT(*) FROM {t}") for t in db.COLLECTED}
    filled = {t for t, n in before.items() if n}
    check("картки записалися",
          filled == {"tenders", "items", "contracts", "documents", "products",
                     "product_specs"}, before)

    removed = db.reset_collected()
    after = {t: db.scalar(f"SELECT COUNT(*) FROM {t}") for t in db.COLLECTED}
    check("зібране прибрано", not any(after.values()), after)
    # Звітуємо лише про те, що справді було: рядок «lots — 0» нічого не додає.
    check("повідомлено, що саме прибрано", set(removed) == filled, removed)
    # А ось індекс і покриття стрічки змін мають лишитися: їх здобувають
    # годинами обходу, і до предмета збору вони стосунку не мають.
    check("індекс уцілів", db.index_size() == 1, db.index_size())
    check("покриття стрічки змін уціліло", db.coverage_days() == {"2026-05-01"},
          db.coverage_days())
    check("повторне очищення не падає", db.reset_collected() == {})
    db.close()


def test_urls_only_when_downloading() -> None:
    print("\n=== посилання лише тоді, коли качаємо ===")
    from app.core.extract import parse_tender

    tender = {
        "id": "uuid-1", "tenderID": "UA-2026-000001-a", "title": "Ноутбуки",
        "status": "complete", "procurementMethodType": "aboveThreshold",
        "dateCreated": "2026-05-01", "value": {"amount": 100000, "currency": "UAH"},
        "procuringEntity": {"name": "Замовник",
                            "identifier": {"id": "04527520"},
                            "address": {"region": "Київська область"}},
        "items": [], "lots": [], "bids": [], "awards": [], "contracts": [],
        "documents": [
            {"id": "d1", "title": "Сертифікат відповідності.pdf",
             "url": "https://public-docs.prozorro.gov.ua/get/abc?Signature=" + "x" * 120,
             "format": "application/pdf", "datePublished": "2026-05-01"},
        ],
    }
    with_urls = parse_tender(tender, keep_urls=True)["docs"]
    without = parse_tender(tender, keep_urls=False)["docs"]
    check("з файлами посилання зберігається", with_urls[0][12].startswith("https://"),
          with_urls[0][12][:40])
    check("без файлів посилання порожнє", without[0][12] == "", without[0][12][:40])
    # Аналітика впізнає сертифікати й авторизаційні листи саме за назвою —
    # вона має лишатися завжди.
    check("назва файлу лишається", without[0][9] == "Сертифікат відповідності.pdf",
          without[0][9])
    check("тип і розділ теж", (without[0][10], without[0][4]) == ("application/pdf",
                                                                  "tender"),
          (without[0][10], without[0][4]))
    check("ключ рядка не залежить від режиму", with_urls[0][0] == without[0][0],
          (with_urls[0][0], without[0][0]))

    from app.core import brands
    check("сертифікат усе одно розпізнається",
          brands.document_kind(without[0][9]).startswith("Сертифікат"),
          brands.document_kind(without[0][9]))


def test_analytics_is_file_only(tmp: Path) -> None:
    print("\n=== аналітика читає лише файл ===")
    import sqlite3

    # 1. Статично: у жодному модулі аналітики немає ані бази, ані sqlite.
    root = Path(__file__).resolve().parent.parent / "app"
    modules = [root / "core" / name for name in
               ("insight.py", "players.py", "benchmark.py", "report.py",
                "xlsxload.py", "xlsxfast.py", "brands.py")]
    modules.append(root / "ui" / "pages" / "analytics_page.py")
    guilty = []
    for module in modules:
        text = module.read_text(encoding="utf-8")
        for mark in ("sqlite", "core.db", "from .db", "Database"):
            if mark in text:
                guilty.append(f"{module.name}: {mark}")
    check("у модулях аналітики немає бази", not guilty, guilty)

    # 2. Сторінка аналітики не отримує бази навіть у конструкторі.
    from app.ui.pages.analytics_page import AnalyticsPage
    import inspect
    params = list(inspect.signature(AnalyticsPage.__init__).parameters)
    check("AnalyticsPage не приймає базу", params == ["self", "settings", "parent"],
          params)

    # 3. На живому запуску: будь-яке звертання до sqlite — помилка.
    from app.core import insight, xlsxload
    from app.core.exporter import write_xlsx

    path = tmp / "лише-файл.xlsx"
    write_xlsx(path, {
        "Закупівлі": (_demo_headers("tenders"),
                      [["UA-1", "2026-05-01", "Ноутбуки", "", "30213100-6", "Завершена",
                        "Відкриті торги", "goods", 100000.0, "UAH", "так", "Замовник",
                        "04527520", "Київська область", "Київ", 0, 1, 0, "", "", "", ""]]),
        "Договори": (_demo_headers("contracts"),
                     [["UA-1", "UA-1-a1", "ТОВ Гравець", "12345678", 95000.0, "UAH",
                       "active", "2026-05-10", "Замовник", "04527520",
                       "Київська область", "30213100-6"]]),
    })

    original = sqlite3.connect

    def refuse(*args, **kwargs):
        raise AssertionError("аналітика звернулася до бази")

    sqlite3.connect = refuse
    try:
        data = xlsxload.load(path)
        report = insight.analyse(data, own_edrpou=["41263186"])
    finally:
        sqlite3.connect = original
    check("аналіз пройшов без жодного дотику до sqlite", len(report.sections) == 9,
          list(report.sections))
    ranking = dict(report.sections["Ринок"][0].tables)["Рейтинг постачальників"][1]
    check("і порахував саме те, що у файлі", ranking[0][6] == 95000.0, ranking)

    # 4. Дві різні книги дають два різні звіти — файл, а не спільне сховище.
    other = tmp / "інша-книга.xlsx"
    write_xlsx(other, {
        "Закупівлі": (_demo_headers("tenders"),
                      [["UA-9", "2026-06-01", "Монітори", "", "30231300-0", "Завершена",
                        "Відкриті торги", "goods", 7000.0, "UAH", "так", "Інший замовник",
                        "04527529", "Львівська область", "Львів", 0, 1, 0, "", "", "", ""]]),
        "Договори": (_demo_headers("contracts"),
                     [["UA-9", "UA-9-a1", "ТОВ Інший", "87654321", 6500.0, "UAH",
                       "active", "2026-06-10", "Інший замовник", "04527529",
                       "Львівська область", "30231300-0"]]),
    })
    second = insight.analyse(xlsxload.load(other), own_edrpou=["41263186"])
    ranking2 = dict(second.sections["Ринок"][0].tables)["Рейтинг постачальників"][1]
    check("друга книга — інший ринок", ranking2[0][6] == 6500.0, ranking2)
    check("жодного сліду першої книги",
          {row[1] for row in ranking2} == {"87654321"},
          {row[1] for row in ranking2})


def test_forecast_models() -> None:
    print("\n=== моделі прогнозу ===")
    from app.core import forecast as fc

    # Квантиль Стьюдента звіряємо з друкованою таблицею: на ньому тримається
    # ширина всіх інтервалів, і помилка тут була б непомітною й системною.
    check("квантиль Стьюдента t(0,90; 4)", abs(fc.t_quantile(0.90, 4) - 1.5332) < 5e-4,
          round(fc.t_quantile(0.90, 4), 4))
    check("квантиль Стьюдента t(0,975; 10)", abs(fc.t_quantile(0.975, 10) - 2.2281) < 5e-4,
          round(fc.t_quantile(0.975, 10), 4))
    check("квантиль ширший за нормальний", fc.t_quantile(0.90, 4) > 1.2816)

    months = [fc.month_shift("2026-01", i) for i in range(12)]
    check("наступний місяць після грудня", fc.month_next("2026-12") == "2027-01")

    # Точна пряма: модель із трендом зобов'язана продовжити її без похибки.
    line = [10.0 + 2 * i for i in range(10)]
    out = fc.forecast("пряма", months[:10], line, horizon=3)
    check("пряма продовжується точно",
          abs(out.points[0].value - 30.0) < 1e-6 and abs(out.points[2].value - 34.0) < 1e-6,
          [p.value for p in out.points])
    check("на прямій похибка нульова", out.mase is not None and out.mase < 1e-6, out.mase)
    check("тренд у прогнозі додатний", out.trend and out.trend > 0, out.trend)

    # Усі моделі мають міритися на однаковій кількості звірок — інакше та,
    # що почала пізніше, виграє на коротшому й легшому хвості ряду.
    checks = {s.checks for s in out.scores if s.mase is not None}
    check("усі моделі перевірено однаково", len(checks) == 1, checks)

    # Стала: жодна модель не має вигадати тренду.
    flat = fc.forecast("стала", months[:8], [5.0] * 8, horizon=3)
    check("стала лишається сталою",
          all(abs(p.value - 5.0) < 1e-9 for p in flat.points), [p.value for p in flat.points])
    check("на сталому ряду сказано, що він сталий",
          any("сталий" in note for note in flat.notes), flat.notes)

    # Оптимальний старт: перше спостереження — викид, і рівень не має за нього
    # чіплятися. Саме тут прогноз ринку помилявся на 8%.
    ses = fc.Ses().fit([1.0, 9.0, 10.0, 11.0, 9.0, 10.0])
    check("рівень не прив'язаний до першої точки", ses.predict(1)[0] > 6.0,
          round(ses.predict(1)[0], 3))

    # Замало місяців — чесна відмова, а не число з повітря.
    short = fc.forecast("короткий", months[:2], [1.0, 2.0], horizon=3)
    check("два місяці — прогнозу немає", not short.ok and short.reason, short.reason)
    gap = fc.forecast("з діркою", ["2026-01", "2026-02", "2026-05"], [1.0, 2.0, 3.0])
    check("розрив у ряду — прогнозу немає", not gap.ok and "розрив" in gap.reason, gap.reason)

    # Спадний ряд не має давати від'ємних грошей.
    falling = fc.forecast("спад", months[:8], [80.0, 70.0, 60.0, 50.0, 40.0, 30.0, 20.0, 10.0],
                          horizon=6)
    check("прогноз не йде нижче нуля",
          all(p.value >= 0 and p.low >= 0 for p in falling.points),
          [(p.value, p.low) for p in falling.points])

    # Інтервал має розширюватися з горизонтом — інакше він не інтервал.
    noisy = [10.0, 13.0, 9.0, 14.0, 11.0, 15.0, 10.0, 16.0, 12.0, 17.0]
    wide = fc.forecast("шум", months[:10], noisy, horizon=4)
    spans = [p.high - p.low for p in wide.points]
    check("межі розходяться з горизонтом", spans[-1] >= spans[0] - 1e-9,
          [round(s, 2) for s in spans])
    check("прогноз усередині своїх меж",
          all(p.low <= p.value <= p.high for p in wide.points))


def test_forecast_choice_risk() -> None:
    print("\n=== межі враховують і невизначеність вибору моделі ===")
    from app.core import forecast as fc

    months = [fc.month_shift("2026-01", i) for i in range(12)]
    flat, drift = fc.Flat(), fc.Drift()

    # Смуга з однієї моделі — сперечатися нема з ким.
    y = [10.0, 11.0, 9.0, 10.5, 10.0, 9.5, 10.2, 10.1]
    alone = fc._rival_spread(y, [flat], flat, 3, flat.fit(y).predict(3))
    check("одна модель — розкиду немає", alone == [0.0, 0.0, 0.0], alone)

    # Стала проти дрейфу на ряду з нахилом: що далі горизонт, то дужче вони
    # розходяться — і саме на стільки має рости інтервал.
    slope = [10.0 + 2 * i for i in range(8)]
    pair = fc._rival_spread(slope, [flat, drift], flat, 3, flat.fit(slope).predict(3))
    check("моделі розходяться — розкид додатний", pair[0] > 0, [round(v, 2) for v in pair])
    check("розкид росте з горизонтом", pair[2] > pair[1] > pair[0],
          [round(v, 2) for v in pair])

    # Наскрізь: та сама вибірка з урахуванням вибору й без нього.
    noisy = [10.0, 13.0, 9.0, 14.0, 11.0, 15.0, 10.0, 16.0, 12.0, 17.0, 13.0, 18.0]
    honest = fc.forecast("шум", months, noisy, horizon=3)
    original = fc._rival_spread
    try:
        fc._rival_spread = lambda *a, **k: [0.0] * a[3]
        naive_bounds = fc.forecast("шум", months, noisy, horizon=3)
    finally:
        fc._rival_spread = original
    wide = honest.points[0].high - honest.points[0].low
    narrow = naive_bounds.points[0].high - naive_bounds.points[0].low
    check("суперників пораховано", honest.rivals >= 1, honest.rivals)
    check("врахування вибору не звужує інтервал", wide >= narrow - 1e-9,
          (round(wide, 2), round(narrow, 2)))
    if honest.rivals > 1:
        check("а коли моделі сперечаються — розширює", wide > narrow,
              (round(wide, 2), round(narrow, 2)))
    check("точка прогнозу від цього не зсувається",
          abs(honest.points[0].value - naive_bounds.points[0].value) < 1e-9,
          (honest.points[0].value, naive_bounds.points[0].value))


def test_forecast_seasonality() -> None:
    print("\n=== сезонність вмикається сама ===")
    import time
    from app.core import forecast as fc

    # Три роки з грудневим сплеском: рівно те, чого не видно на пів року.
    spike = [0.0] * 11 + [60.0]
    y = [100.0 + 3 * t + spike[t % 12] for t in range(36)]
    months = [fc.month_shift("2024-01", i) for i in range(36)]
    started = time.time()
    out = fc.forecast("сезонний", months, y, horizon=3)
    spent = time.time() - started
    check("сезонні моделі з'явилися в наборі",
          any("езонн" in s.model for s in out.scores), [s.model for s in out.scores])
    check("обрано сезонну модель", "езонн" in out.model, out.model)
    # Наступний за груднем місяць — січень без сплеску: несезонна модель
    # потягнула б грудневі +60 далі.
    check("січень після грудня без сплеску", out.points[0].value < 100 + 3 * 36 + 30,
          round(out.points[0].value, 1))
    check("сезонний прогноз близький до правди",
          abs(out.points[0].value - (100 + 3 * 36)) < 25, round(out.points[0].value, 1))
    check("на 36 точках звірок уже вистачає", out.checks >= 6, out.checks)
    check("перебір сітки не гальмує", spent < 20.0, f"{spent:.1f} с")

    # На пів року сезонні моделі не мають вмикатися взагалі.
    half = fc.forecast("пів року", months[:6], y[:6], horizon=3)
    check("на 6 місяцях сезонності немає",
          not any("езонн" in s.model for s in half.scores), [s.model for s in half.scores])


def _lagged_workbook(tmp: Path) -> Path:
    """Вісім місяців із відомою затримкою договору — рівно 10 днів.

    Закупівлі виходять 1, 15 і 25 числа, договір з'являється через десять
    днів. Останню серпневу закупівлю (25.08) договір ще не наздогнав — саме
    та ситуація, через яку останній місяць будь-якої вибірки виглядає
    провальним. Тут вона відтворена так, що правильну відповідь можна
    порахувати на папері: видно дві закупівлі з трьох, тобто дві третини.
    """
    from app.core.exporter import write_xlsx

    days = ("01", "15", "25")
    rows, deals = [], []
    for month in range(1, 9):
        for day in days:
            code = f"UA-2026-{month:02d}-{day}"
            rows.append([code, f"2026-{month:02d}-{day}", "Ноутбуки", "", "30213100-6",
                         "Завершена", "Звіт про договір", "goods", 100000.0, "UAH",
                         "так", "Замовник", "04527520", "Київська область", "Київ",
                         0, 1, 0, "", "", "", ""])
            signed = date(2026, month, int(day)) + timedelta(days=10)
            if signed <= date(2026, 8, 25):
                deals.append([code, code + "-a1", RIVAL, "12345678", 90000.0, "UAH",
                              "active", signed.isoformat(), "Замовник", "04527520",
                              "Київська область", "30213100-6"])
    tenders = (["Номер закупівлі", "Дата оприлюднення", "Предмет закупівлі", "Опис",
                "Коди ДК021", "Статус", "Процедура", "Категорія предмета",
                "Очікувана вартість", "Валюта", "ПДВ включено", "Замовник",
                "ЄДРПОУ замовника", "Регіон", "Населений пункт", "Лотів",
                "Пропозицій", "Файлів", "Початок подання", "Кінець подання",
                "Остання зміна", "Посилання"], rows)
    contracts = (["Номер закупівлі", "Номер договору", "Постачальник",
                  "ЄДРПОУ постачальника", "Сума договору", "Валюта",
                  "Статус договору", "Дата підписання", "Замовник",
                  "ЄДРПОУ замовника", "Регіон", "Коди ДК021"], deals)
    path = tmp / "lagged.xlsx"
    write_xlsx(path, {"Закупівлі": tenders, "Договори": contracts})
    return path


def test_forecast_completeness(tmp: Path) -> None:
    print("\n=== поправка на неповноту останніх місяців ===")
    from app.core import insight, xlsxload

    data = xlsxload.load(_lagged_workbook(tmp))
    report = insight.analyse(data, own_edrpou=["41263186"], horizon=2)
    blocks = report.sections["Прогнозування"]
    check("розділ прогнозу з двох блоків і більше", len(blocks) >= 2, len(blocks))

    curve = dict(blocks[1].tables)["Крива розвитку договорів"][1]
    known = {row[0]: row[1] for row in curve}
    check("до десятого дня не видно нічого", known.get(7) == 0.0, known)
    check("на десятий день видно все", known.get(10) == 1.0, known)

    full = dict(blocks[1].tables)["Повнота місяців"][1]
    august = [row for row in full if row[0] == "2026-08" and row[1] == "Сума угод"]
    check("серпень позначено неповним", len(august) == 1, full)
    if august:
        row = august[0]
        check("розвиток серпня — дві третини", abs(row[4] - 2 / 3) < 0.005, row[4])
        check("вибірка уривається посеред серпня", 0.5 < row[3] < 1.0, row[3])
        check("спостережено дві угоди з трьох", row[2] == 180000.0, row[2])
        check("поправка піднімає, а не опускає", row[6] > row[2], (row[2], row[6]))
        # У таблиці частки округлені до чотирьох знаків, тож звіряємо
        # відносно: точне ділення дало б розбіжність у сотню гривень.
        check("поправка — це ділення на повноту",
              abs(row[6] / (row[2] / (row[3] * row[4])) - 1) < 0.005,
              (row[6], round(row[2] / (row[3] * row[4]), 2)))
    older = [row for row in full if row[0] < "2026-08" and row[4] < 0.999]
    check("зрілі місяці не коригуються за розвитком", not older, older)

    points = dict(blocks[0].tables)["Прогноз по місяцях"][1]
    ahead = {row[1] for row in points if row[0] == "Сума угод"}
    check("прогноз рівно на заданий горизонт", ahead == {"2026-09", "2026-10"}, ahead)

    # Зворотна перевірка: програма вдає, що збір урвався кінцем липня, і
    # звіряє скориговане число з тим, що в липні опинилося насправді. На цих
    # даних відповідь відома точно — договір відстає рівно на десять днів,
    # тож видно дві закупівлі з трьох, а поправка має влучити в яблучко.
    check_sheet = dict(blocks[1].tables).get("Перевірка поправки")
    check("поправку перевірено ретроспективно", bool(check_sheet and check_sheet[1]),
          check_sheet[1] if check_sheet else None)
    if check_sheet and check_sheet[1]:
        row = check_sheet[1][0]
        check("без поправки місяць недорахований на третину",
              abs(row[4] + 1 / 3) < 0.01, row[4])
        check("з поправкою похибки немає", abs(row[5]) < 0.01, row[5])
        check("скориговане збігається з тим, що виявилося насправді",
              abs(row[2] - row[3]) < 1.0, (row[2], row[3]))
    check("перевірка не підглядає в майбутнє",
          all(r[1] < r[3] for r in (check_sheet[1] if check_sheet else [])),
          check_sheet[1] if check_sheet else None)

    chart = next(c for c in blocks[0].charts if c.title.startswith("Сума угод"))
    rows_by_name = {s.name: s for s in chart.series}
    labels = rows_by_name["Факт"].labels
    check("вісь без повторів і за зростанням",
          labels == sorted(set(labels)) and len(labels) == 10, labels)
    fact = rows_by_name["Факт"].values
    line = rows_by_name["Прогноз"].values
    check("факт закінчується там, де починається прогноз",
          fact[-1] is None and fact[-3] is not None and line[-1] is not None
          and line[-4] is None, (fact[-4:], line[-4:]))
    check("лінії стикаються в останньому факті", line[-3] is not None, line[-4:])

    # Прогноз відходить не від видимої частини останнього місяця, а від
    # оцінки повного — інакше модель продовжувала б обірваний ряд. Щоб це не
    # читалося як розрив, скоригований ряд іде суцільно через усю історію.
    fixed = rows_by_name["Оцінка повного місяця"]
    check("поправка малюється під фактом, а не поверх нього",
          chart.series.index(fixed) < chart.series.index(rows_by_name["Факт"]),
          [s.name for s in chart.series])
    history = len(labels) - 2
    check("скоригований ряд суцільний по всій історії",
          all(v is not None for v in fixed.values[:history])
          and all(v is None for v in fixed.values[history:]),
          fixed.values)
    check("прогноз стартує саме зі скоригованої точки",
          line[history - 1] == fixed.values[history - 1] > fact[history - 1],
          (line[history - 1], fixed.values[history - 1], fact[history - 1]))


def test_rejection_reasons() -> None:
    print("\n=== причини відмов: розбір формулювання ===")
    from app.core import rejections as rj

    # Формулювання з живих карток (02.09.2026): у вибірці 72 відмов саме ці
    # чотири підстави й трапилися.
    check("відмова від договору",
          rj.classify("Переможець письмово відмовився від укладення договору на умовах, "
                      "визначених замовником у запиті пропозицій постачальників")
          == "Відмова від договору")
    check("не підписав договір",
          rj.classify("Переможець не підписав договір, на умовах, визначених замовником "
                      "у запиті пропозицій постачальників, у строк, визначений пунктом 66")
          == "Не підписав договір у строк")
    check("технічна специфікація",
          rj.classify("Тендерна пропозиція учасника не відповідає умовам технічної "
                      "специфікації та іншим вимогам щодо предмета закупівлі тендерної "
                      "документації") == "Невідповідність технічним вимогам")
    check("умови оголошення",
          rj.classify("Пропозиція учасника не відповідає умовам, визначеним в оголошенні "
                      "про проведення спрощеної закупівлі, та вимогам до предмета закупівлі")
          == "Невідповідність вимогам документації")
    check("підстави статті 17",
          rj.classify("Не надав у спосіб, зазначений в тендерній документації, документи, "
                      "що підтверджують відсутність підстав, визначених у підпунктах 3, 5, "
                      "6 і 12 пункту 47 Особливостей") == "Не надав документи")
    # Обидві підстави посилаються на той самий пункт 47, але кажуть протилежне:
    # одна — «він не довів, що не такий», друга — «він таки такий». Голе
    # посилання на пункт колись відносило другу до першої.
    check("підпадає під підстави — не те саме, що не надав довідку",
          rj.classify("Учасник підпадає під підстави, встановлені пунктом 47 Особливостей")
          == "Підпадає під підстави для відмови")
    check("голе посилання на пункт нічого не вигадує",
          rj.classify("Відхилено на підставі пункту 47 Особливостей") == rj.OTHER)

    # Підстави з пункту 44 Особливостей, яких у цій вибірці не було, але які
    # трапляються на інших ринках.
    check("недостовірна інформація",
          rj.classify("Учасник зазначив у тендерній пропозиції недостовірну інформацію")
          == "Недостовірна інформація")
    check("не усунув невідповідності",
          rj.classify("Учасник не виправив виявлені замовником невідповідності")
          == "Не усунув невідповідності")
    check("аномально низька ціна",
          rj.classify("Учасник не надав обґрунтування аномально низької ціни")
          == "Аномально низька ціна")
    check("кваліфікаційні критерії",
          rj.classify("Учасник не відповідає кваліфікаційним критеріям, установленим "
                      "статтею 16 Закону") == "Не відповідає кваліфікаційним критеріям")
    check("строк дії пропозиції",
          rj.classify("Тендерна пропозиція є такою, строк дії якої закінчився")
          == "Строк дії пропозиції минув")
    # «Не надав забезпечення» починається тим самим «не надав», що й «не надав
    # документи», а слово «документація» стоїть поруч — без застереження в
    # шаблоні перемагала б чужа категорія.
    check("забезпечення не плутається з документами",
          rj.classify("Переможець не надав забезпечення виконання договору у формі, "
                      "визначеній тендерною документацією") == "Не надав забезпечення")

    # Замовники зчіплюють підстави в один рядок — вирішує позиція, а не
    # порядок категорій у довіднику.
    check("перемагає перша названа підстава",
          rj.classify("не відповідає умовам технічної специфікації; не відповідає "
                      "вимогам, установленим у тендерній документації")
          == "Невідповідність технічним вимогам")
    check("підстава важливіша за переказ обставин",
          rj.classify("Тендерна пропозиція не відповідає умовам технічної специфікації",
                      "Учасник згодом відмовився від укладення договору")
          == "Невідповідність технічним вимогам")
    check("без підстави читаємо пояснення",
          rj.classify("", "Переможець письмово відмовився від укладення договору")
          == "Відмова від договору")
    check("немає тексту — немає причини", rj.classify("", "") == rj.UNSTATED)
    check("незнайоме формулювання не вигадується",
          rj.classify("Рішення уповноваженої особи №17") == rj.OTHER)

    check("ознаки документів",
          rj.marks("не надав сертифікат", "відсутній лист виробника")
          == ["сертифікат", "авторизаційний лист"])
    check("без тексту ознак немає", rj.marks("", "") == [])
    check("підстава відміни з переліку",
          rj.cancel_label("noDemand") == "Відсутня потреба в закупівлі")
    check("невідомий код підстави лишається як є",
          rj.cancel_label("newReason") == "newReason")
    # Скасоване рішення — це та сама подія вдруге (51 із 55 у вибірці мали
    # `unsuccessful`-двійника), тож відмовою вважаємо лише відхилення.
    check("відмова — це лише unsuccessful",
          (rj.is_rejected("unsuccessful"), rj.is_rejected("cancelled"),
           rj.is_rejected("active")) == (True, False, False))


def test_rejection_extract(tmp: Path) -> None:
    print("\n=== причини відмов: розбір картки й запис у базу ===")
    tender = {
        "id": "u1", "tenderID": "UA-2026-05-01-000001-a",
        "awards": [
            {"id": "aw1", "status": "unsuccessful", "date": "2026-05-10",
             "value": {"amount": 100.0, "currency": "UAH"},
             "title": "Переможець письмово відмовився від укладення договору",
             "description": "Лист-відмова б/н", "qualified": False,
             "suppliers": [{"name": "ТОВ Один",
                            "identifier": {"scheme": "UA-EDR", "id": "12345678"}}]},
            {"id": "aw2", "status": "active", "date": "2026-05-12",
             "value": {"amount": 120.0, "currency": "UAH"},
             "title": "Визнаний переможцем", "qualified": True, "eligible": True,
             "suppliers": [{"name": "ТОВ Два",
                            "identifier": {"scheme": "UA-EDR", "id": "87654321"}}]},
        ],
        "cancellations": [
            {"id": "c1", "status": "active", "reasonType": "noDemand",
             "reason": "Потреба відпала", "date": "2026-05-20"},
        ],
    }
    parsed = parse_tender(tender)
    first, second = parsed["awards"]
    check("підставу знято з картки", first[9].startswith("Переможець письмово"), first[9])
    check("пояснення знято з картки", first[10] == "Лист-відмова б/н", first[10])
    check("qualified — 0, а не порожньо", first[11] == 0, first[11])
    # `eligible` приходить не завжди (виміряно: 18 зі 106 нечинних рішень),
    # і нуль замість «невідомо» читався б як «не відповідає критеріям».
    check("відсутнє eligible лишається невідомим", first[12] is None, first[12])
    check("у чинному рішенні title — не причина", second[9] == "Визнаний переможцем")
    # Конвеєр бере ЄДРПОУ переможця з дев'ятого поля; нові колонки дописані в
    # кінець саме тому.
    check("ЄДРПОУ переможця лишився восьмим", first[8] == "12345678", first[8])
    check("відміну закупівлі розібрано",
          parsed["cancellations"] == [("c1", "u1", "active", "noDemand",
                                       "Потреба відпала", None, "2026-05-20")],
          parsed["cancellations"])

    db = Database(tmp / "відмови.db")
    db.save_tender(parsed["row"], lots=[], items=[], bids=[], awards=parsed["awards"],
                   contracts=[], docs=[], cancellations=parsed["cancellations"])
    rows = {r["id"]: dict(r) for r in db.query("SELECT * FROM awards")}
    check("причина збереглася", rows["aw1"]["reason"].startswith("Переможець"))
    check("відміна збереглася",
          db.scalar("SELECT reason_type FROM cancellations") == "noDemand")
    # Типове очищення прибирає зібране, але не індекс: відміни — зібране.
    db.reset_collected()
    check("відміни прибираються разом зі збором",
          db.scalar("SELECT COUNT(*) FROM cancellations") == 0)
    db.close()

    # База, зібрана попередньою версією: колонок причин у ній ще немає, і
    # `CREATE TABLE IF NOT EXISTS` їх не додасть.
    import sqlite3
    old = tmp / "стара.db"
    conn = sqlite3.connect(old)
    conn.execute("CREATE TABLE awards (id TEXT PRIMARY KEY, tender_uuid TEXT NOT NULL,"
                 " lot_id TEXT, status TEXT, date TEXT, value_amount REAL, currency TEXT,"
                 " supplier_name TEXT, supplier_edrpou TEXT)")
    conn.commit()
    conn.close()
    upgraded = Database(old)
    columns = {r[1] for r in upgraded.query("PRAGMA table_info(awards)")}
    check("стару базу дописано без втрат",
          {"reason", "explanation", "qualified", "eligible"} <= columns, sorted(columns))
    upgraded.close()


def test_rejection_export(tmp: Path) -> None:
    print("\n=== причини відмов: книга даних ===")
    from app.core import rawexport
    from app.core.xlsxload import TABLES, _match_columns

    db = Database(tmp / "вивантаження-відмов.db")
    row = {"uuid": "u1", "tender_id": "UA-1", "title": "Ноутбуки", "description": "",
           "status": "complete", "method_type": "aboveThreshold", "main_category": "goods",
           "date_created": "2026-05-01T10:00:00+03:00", "date_modified": "", "tender_start": "",
           "tender_end": "", "value_amount": 100.0, "value_currency": "UAH", "vat_included": 1,
           "pe_name": "Замовник", "pe_edrpou": "04527520", "pe_region": "Київська область",
           "pe_locality": "Київ", "n_lots": 0, "n_bids": 1, "n_docs": 0,
           "cpv_list": "30213100-6", "fetched_at": ""}
    db.save_tender(
        row, lots=[], items=[], bids=[], contracts=[], docs=[],
        awards=[("aw1", "u1", None, "unsuccessful", "2026-05-10", 100.0, "UAH",
                 "ТОВ Один", "12345678",
                 "Переможець письмово відмовився від укладення договору", "", 0, None),
                ("aw2", "u1", None, "active", "2026-05-12", 120.0, "UAH",
                 "ТОВ Два", "87654321", "Визнаний переможцем", "", 1, 1)],
        cancellations=[("c1", "u1", "active", "noDemand", "Потреба відпала", None,
                        "2026-05-20")])

    headers, rows = rawexport.awards(db)
    by_id = {r[3]: r for r in rows}
    check("категорія порахована при вивантаженні",
          by_id["12345678"][9] == "Відмова від договору", by_id["12345678"][9])
    # «Визнаний переможцем» — не причина; лишити його в колонці причин
    # означало б, що кожен другий переможець «відхилений».
    check("у чинного рішення причини немає",
          (by_id["87654321"][8], by_id["87654321"][9]) == ("", ""), by_id["87654321"])

    cancel_headers, cancel_rows = rawexport.cancellations(db)
    check("підстава відміни підписана людською мовою",
          cancel_rows[0][7] == "Відсутня потреба в закупівлі", cancel_rows[0])

    # Книгу читає аналітика, і зіставляє вона за назвами колонок — тож обидва
    # боки мають зійтися без окремої домовленості.
    found = _match_columns(list(headers), TABLES["awards"]["columns"])
    check("аналітика знаходить причину в книзі",
          {"reason", "explanation"} <= set(found), sorted(found))
    found_cancel = _match_columns(list(cancel_headers), TABLES["cancellations"]["columns"])
    check("аналітика знаходить відміни в книзі",
          set(TABLES["cancellations"]["columns"]) == set(found_cancel), sorted(found_cancel))

    # Те саме зведення, що й за кнопкою «Зведення»: воно рахується з бази, а не
    # з книги, тож причини мають зійтися по обох шляхах.
    from app.core import analytics
    sheets = analytics.build_sheets(db)
    reasons = sheets["Причини відмов"][1]
    check("причини у зведенні", reasons[0][:2] == ["Відмова від договору", 1], reasons)
    players = sheets["Скасовані перемоги"][1]
    check("зірвана перемога в зведенні",
          players[0][:3] == ["12345678", "ТОВ Один", 1], players)
    db.close()


def _rejection_dataset(tmp: Path):
    """Вибірка, у якій перемоги зриваються всіма відомими способами."""
    from app.core.xlsxload import Dataset

    OURS, RIVAL = "41263186", "12345678"

    def tender(number: int):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01", "title": "Ноутбуки",
                "description": "", "cpv_list": "30213100-6", "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 150000.0,
                "currency": "UAH", "vat": "так", "buyer": f"Замовник {number}",
                "buyer_edrpou": f"0452752{number}", "region": "Київська область",
                "locality": "Київ", "n_lots": 0, "n_bids": 2, "n_docs": 0,
                "tender_start": "", "tender_end": "", "modified": "", "url": ""}

    def bid(number, edrpou, amount):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01", "name": f"ТОВ {edrpou}",
                "edrpou": edrpou, "region": "Київська область", "amount": amount,
                "currency": "UAH", "status": "active", "submitted": "2026-05-02"}

    def award(number, edrpou, status, amount, reason="", explanation=""):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01", "name": f"ТОВ {edrpou}",
                "edrpou": edrpou, "amount": amount, "currency": "UAH", "status": status,
                "decided": "2026-05-10", "reason": reason, "explanation": explanation}

    def contract(number, edrpou, amount):
        return {"tender_id": f"UA-{number}", "contract_id": f"UA-{number}-a1",
                "name": f"ТОВ {edrpou}", "edrpou": edrpou, "amount": amount,
                "currency": "UAH", "status": "active", "signed": "2026-05-10",
                "buyer": f"Замовник {number}", "buyer_edrpou": f"0452752{number}",
                "region": "Київська область", "cpv_list": "30213100-6"}

    return Dataset(
        path=tmp / "відмови.xlsx",
        tenders=[tender(i) for i in range(1, 5)],
        bids=[bid(1, RIVAL, 100000.0),
              bid(2, OURS, 90000.0), bid(2, RIVAL, 100000.0),
              bid(3, RIVAL, 100000.0), bid(4, RIVAL, 100000.0)],
        awards=[
            award(1, RIVAL, "active", 100000.0),
            # Нас відхилили за технікою, і закупівля пішла до конкурента.
            award(2, OURS, "unsuccessful", 90000.0,
                  "Тендерна пропозиція учасника не відповідає умовам технічної специфікації"),
            award(2, RIVAL, "active", 100000.0),
            # Одна подія двома рядками: спершу скасоване рішення, потім відхилення.
            award(3, RIVAL, "cancelled", 100000.0, "Визнаний переможцем"),
            award(3, RIVAL, "unsuccessful", 100000.0,
                  "Переможець письмово відмовився від укладення договору"),
            # Замовник обмежився зміною статусу.
            award(4, RIVAL, "unsuccessful", 100000.0),
        ],
        contracts=[contract(1, RIVAL, 100000.0), contract(2, RIVAL, 100000.0)],
        cancellations=[{"tender_id": "UA-4", "date": "2026-05-01", "buyer": "Замовник 4",
                        "buyer_edrpou": "04527524", "value": 150000.0,
                        "cancelled": "2026-05-20", "status": "active",
                        "reason": "Відсутня потреба в закупівлі",
                        "explanation": "Потреба відпала"}],
    ), OURS, RIVAL


def test_rejection_section(tmp: Path) -> None:
    print("\n=== причини відмов: розділ звіту ===")
    from app.core import insight

    data, ours, rival = _rejection_dataset(tmp)
    report = insight.analyse(data, own_edrpou=[ours])
    check("розділ «Відмови» стоїть після прогнозу",
          list(report.sections).index("Відмови") == list(report.sections).index("Ринок") + 2,
          list(report.sections))
    block = report.sections["Відмови"][0]
    tiles = dict(block.tiles)
    # Скасоване рішення в UA-3 — двійник відхилення, і рахувати його окремо
    # означало б чотири відмови замість трьох.
    check("скасоване рішення не рахується вдруге",
          tiles["Скасовано перемог"] == "3", tiles)
    # Знаменник — ухвалені рішення: два чинних і три відхилення.
    check("частка від ухвалених рішень", tiles["Частка рішень"] == "60,0%", tiles)
    check("сума скасованих перемог", tiles["Сума скасованих перемог"].startswith("290"), tiles)
    check("закупівель зачепило три", tiles["Закупівель зачепило"] == "3", tiles)
    check("наші втрати видно окремо",
          tiles["Наших перемог зірвано"].startswith("1 на 90"), tiles)

    reasons = {row[0]: row[1] for row in dict(block.tables)["Причини"][1]}
    check("причини розкладено", reasons == {"Відмова від договору": 1,
                                            "Невідповідність технічним вимогам": 1,
                                            "Причину не зазначено": 1}, reasons)
    firms = {row[0]: row for row in dict(block.tables)["Компанії"][1]}
    check("у конкурента дві зірвані перемоги", firms[rival][2] == 2, firms[rival])
    check("частка зривів рахується від його ж перемог",
          firms[rival][4] == 0.5, firms[rival])
    check("наша компанія позначена", firms[ours][7] == "так", firms[ours])
    cases = dict(block.tables)["Випадки"][1]
    check("у переліку випадків три рядки", len(cases) == 3, len(cases))

    notes = " ".join(block.notes)
    check("порожню підставу названо прямо", "У 1 випадку з 3" in notes, notes)
    # Число у висновок підставляється з даних, тож форма іменника має
    # узгоджуватися сама: «зачепило 1 разів» вилізло б на першій же вибірці.
    check("число узгоджене з іменником", "зачепило 1 раз на" in notes, notes)

    cancelled = report.sections["Відмови"][1]
    check("відмінені закупівлі окремим блоком",
          cancelled.title == "Відмінені закупівлі", cancelled.title)
    check("підстава відміни в таблиці",
          dict(cancelled.tables)["Підстави"][1][0][0] == "Відсутня потреба в закупівлі",
          dict(cancelled.tables)["Підстави"][1])

    # Вкладка будується під час першого показу — саме там ловляться помилки в
    # кресленні нових діаграм.
    from PySide6.QtGui import QPixmap
    from PySide6.QtWidgets import QApplication

    from app.config import Settings
    from app.ui.pages.analytics_page import AnalyticsPage

    QApplication.instance() or QApplication([])
    page = AnalyticsPage(Settings())
    page.resize(1200, 800)
    page.report = report
    page.show_report(report)
    page.tabs.setCurrentIndex([page.tabs.tabText(i) for i in range(page.tabs.count())]
                              .index("Відмови"))
    page.render(QPixmap(page.size()))
    check("вкладка «Відмови» малюється", True)


def test_rejection_in_profiles(tmp: Path) -> None:
    print("\n=== причини відмов: портрети й очні зустрічі ===")
    from app.core import insight

    data, ours, rival = _rejection_dataset(tmp)
    report = insight.analyse(data, own_edrpou=[ours])

    profile = next(p for p in report.competitors if p.edrpou == rival)
    check("зірвані перемоги в портреті", profile.n_rejected == 2, profile.n_rejected)
    check("головна причина названа",
          profile.reject_reason in ("Відмова від договору", "Причину не зазначено"),
          profile.reject_reason)
    tiles = dict(profile.block.tiles)
    check("плитка зривів", tiles["Зірвано перемог"].startswith("2 на 200"), tiles)
    check("таблиця зірваних перемог є",
          len(dict(profile.block.tables)["Зірвані перемоги"][1]) == 2)

    # Головне: там, де ціна була наша, а закупівля пішла не до нас, більше не
    # доводиться здогадуватись — підставу назвав сам замовник.
    verdicts = {row[0]: row[7] for row in dict(profile.block.tables)["Наші зустрічі"][1]}
    check("причину програшу взято з картки",
          verdicts.get("UA-2") == "нас відхилили: невідповідність технічним вимогам",
          verdicts)

    mine = next(p for p in report.ours if p.edrpou == ours)
    check("наша зірвана перемога порахована", mine.n_rejected == 1, mine.n_rejected)
    summary = " ".join(report.sections["Підсумок"][0].notes)
    check("у підсумку сказано, скільки перемог втрачено після відхилення",
          "втратили після відхилення" in summary, summary)


def test_rejection_double_count(tmp: Path) -> None:
    print("\n=== причини відмов: сума рішень — не сума грошей ринку ===")
    from app.core import insight
    from app.core.rejections import DETAIL_ROWS
    from app.core.xlsxload import Dataset

    def tender(number: int):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01", "title": "Ноутбуки",
                "description": "", "cpv_list": "30213100-6", "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 150000.0,
                "currency": "UAH", "vat": "так", "buyer": "Замовник",
                "buyer_edrpou": "04527520", "region": "Київська область",
                "locality": "Київ", "n_lots": 0, "n_bids": 2, "n_docs": 0,
                "tender_start": "", "tender_end": "", "modified": "", "url": ""}

    def award(number, edrpou, status, amount, reason=""):
        return {"tender_id": f"UA-{number}", "date": "2026-05-01", "name": f"ТОВ {edrpou}",
                "edrpou": edrpou, "amount": amount, "currency": "UAH", "status": status,
                "decided": "2026-05-10", "reason": reason, "explanation": ""}

    # Одна закупівля, у якій замовник поспіль відхилив двох переможців, і
    # договір урешті підписав третій. Грошей ринок не втратив — але скасованих
    # рішень тут два, і сума їх обох до ринку не додається.
    refusal = "Переможець письмово відмовився від укладення договору"
    data = Dataset(
        path=tmp / "подвійний-рахунок.xlsx",
        tenders=[tender(1)],
        awards=[award(1, "11111111", "unsuccessful", 100000.0, refusal),
                award(1, "22222222", "unsuccessful", 110000.0, refusal),
                award(1, "33333333", "active", 120000.0)],
        contracts=[{"tender_id": "UA-1", "contract_id": "UA-1-a1", "name": "ТОВ 33333333",
                    "edrpou": "33333333", "amount": 120000.0, "currency": "UAH",
                    "status": "active", "signed": "2026-05-12", "buyer": "Замовник",
                    "buyer_edrpou": "04527520", "region": "Київська область",
                    "cpv_list": "30213100-6"}],
    )
    block = insight.analyse(data).sections["Відмови"][0]
    tiles = dict(block.tiles)
    check("скасованих перемог дві", tiles["Скасовано перемог"] == "2", tiles)
    check("а закупівля одна", tiles["Закупівель зачепило"] == "1", tiles)
    check("сума — це сума рішень",
          tiles["Сума скасованих перемог"].startswith("210"), tiles)
    notes = " ".join(block.notes)
    check("про повторний рахунок сказано прямо",
          "враховані кілька разів" in notes and "не грошей, що зникли з ринку" in notes,
          notes)
    check("число узгоджене й тут", "У 1 закупівлі поспіль" in notes
          and "у 1 закупівлі на" in notes, notes)

    # Довгий перелік випадків обрізається — і про це теж треба сказати, бо
    # інакше «300 з 300» у таблиці читається як «оце всі відмови».
    many = Dataset(
        path=tmp / "багато-відмов.xlsx",
        tenders=[tender(i) for i in range(1, DETAIL_ROWS + 2)],
        awards=[award(i, f"{i:08d}", "unsuccessful", 1000.0 + i, refusal)
                for i in range(1, DETAIL_ROWS + 2)],
    )
    long_block = insight.analyse(many).sections["Відмови"][0]
    cases = dict(long_block.tables)["Випадки"][1]
    check("перелік обрізано до межі", len(cases) == DETAIL_ROWS, len(cases))
    check("про обрізання сказано",
          any("Повний перелік" in note for note in long_block.notes), long_block.notes)


def test_excel_serial_time() -> None:
    print("\n=== порядковий номер Excel → дата ===")
    from app.core.xlsxfast import _from_serial

    # 0,572916666… доби — це рівно 13:45. У подвійній точності
    # timedelta(days=...) давав 13:44:59,999999, і час у книзі читався б на
    # секунду раніше, ніж його показує Excel.
    check("час без похибки", _from_serial(46143.572916666664).isoformat()
          == "2026-05-01T13:45:00", _from_serial(46143.572916666664).isoformat())
    check("ціла доба — опівніч", _from_serial(46143).isoformat() == "2026-05-01T00:00:00")
    check("північ наступного дня не з'їдається",
          _from_serial(46143.99999999).date().isoformat() == "2026-05-01")
    check("сміття не падає", _from_serial(1e12) == 1e12)


def test_progress_scale(tmp: Path) -> None:
    print("\n=== спільна шкала поступу: читання + аналіз ===")
    from app.core import insight, xlsxload

    # Сторінка аналітики склеює дві смуги в одну, зсуваючи аналіз на кількість
    # аркушів книги. Числа мусять збігатися з тим, що модулі справді роблять:
    # доданий аркуш або новий крок інакше зсувають відсоток, і смуга стрибає.
    steps: list[tuple[int, int]] = []
    insight.analyse(xlsxload.Dataset(path=tmp / "порожньо.xlsx"),
                    on_progress=lambda stage, done, total: steps.append((done, total)))
    check("аналіз оголошує стільки кроків, скільки в константі",
          {total for _done, total in steps} == {insight.ANALYSIS_STEPS},
          {total for _done, total in steps})
    check("останній крок доходить до кінця шкали",
          max(done for done, _total in steps) == insight.ANALYSIS_STEPS,
          max(done for done, _total in steps))
    check("кроки не повторюються й не задкують",
          [d for d, _t in steps] == sorted(set(d for d, _t in steps)),
          [d for d, _t in steps])

    read: list[tuple[int, int]] = []
    xlsxload.load(_empty_book(tmp), lambda stage, done, total: read.append((done, total)))
    check("читання оголошує стільки кроків, скільки логічних таблиць",
          {total for _done, total in read} == {len(xlsxload.TABLES)},
          {total for _done, total in read})


def _empty_book(tmp: Path) -> Path:
    from openpyxl import Workbook

    path = tmp / "шкала.xlsx"
    book = Workbook()
    book.active.title = "Закупівлі"
    book.active.append(["Номер закупівлі", "Дата оприлюднення"])
    book.save(path)
    return path


def test_impossible_dates(tmp: Path) -> None:
    print("\n=== неможлива дата не валить аналіз ===")
    from app.core import insight
    from app.core.xlsxload import Dataset, as_date

    # Раніше досить було, щоб дефіси стояли на своїх місцях: «2026-13-45»
    # проходило як дата, а прогноз робив із неї справжній ``date`` — і весь
    # аналіз падав із ValueError: month must be in 1..12.
    check("неможливий місяць відкидається", as_date("2026-13-45") == "", as_date("2026-13-45"))
    check("неможливий день відкидається", as_date("2026-02-30") == "")
    check("29 лютого високосного року — дата", as_date("29.02.2024") == "2024-02-29")
    check("29 лютого звичайного — ні", as_date("29.02.2026") == "")
    check("сміття з цифр не стає датою", as_date("45.99.2026") == "")
    check("звичайна дата не постраждала", as_date("2026-05-01T10:00:00+03:00") == "2026-05-01")
    check("місяць лише з можливої дати", insight.month_of("2026-05-01") == "2026-05")
    check("з неможливої — порожньо", insight.month_of("2026-13-01") == "")

    def tender(number: int, day: str):
        return {"tender_id": f"UA-{number}", "date": day, "title": "Ноутбуки",
                "description": "", "cpv_list": "30213100-6", "status": "Завершена",
                "method": "Відкриті торги", "category": "goods", "value": 150000.0,
                "currency": "UAH", "vat": "так", "buyer": "Замовник",
                "buyer_edrpou": "04527520", "region": "Київська область",
                "locality": "Київ", "n_lots": 0, "n_bids": 1, "n_docs": 0,
                "tender_start": "", "tender_end": "", "modified": "", "url": ""}

    def contract(number: int, day: str):
        return {"tender_id": f"UA-{number}", "contract_id": f"UA-{number}-a1",
                "name": "ТОВ", "edrpou": "12345678", "amount": 100000.0,
                "currency": "UAH", "status": "active", "signed": day,
                "buyer": "Замовник", "buyer_edrpou": "04527520",
                "region": "Київська область", "cpv_list": "30213100-6"}

    days = ["2026-03-01", "2026-04-01", "2026-05-01", "2026-06-01", "2026-07-01",
            "2026-13-45"]
    data = Dataset(path=tmp / "зіпсована-дата.xlsx",
                   tenders=[tender(i, d) for i, d in enumerate(days, 1)],
                   contracts=[contract(i, d) for i, d in enumerate(days, 1)])
    report = insight.analyse(data)
    check("аналіз доходить до кінця", len(report.sections) >= 8, list(report.sections))
    check("прогноз усе одно є", "Прогнозування" in report.sections, list(report.sections))


def test_visibility_matches_deals(tmp: Path) -> None:
    print("\n=== крива видимості = правило угод ===")
    from datetime import date as _date

    from app.core.outlook import OutlookMixin

    published = _date(2026, 5, 1)
    award_one = (_date(2026, 5, 10), 100.0)
    award_two = (_date(2026, 5, 25), 50.0)
    deal = (_date(2026, 5, 20), 90.0)

    def visible(steps, age):
        return round(sum(delta for lag, delta in steps if lag <= age), 2)

    # Доки договору немає, гроші закупівлі — це рішення про переможця.
    only_awards = OutlookMixin._visibility(published, (), [award_one])
    # 10 травня — це дев'ятий день від оприлюднення 1 травня.
    check("до договору видно рішення", visible(only_awards, 8) == 0.0
          and visible(only_awards, 9) == 100.0, only_awards)

    # Щойно виходить договір, рішення перестають рахуватися — і ті, що
    # з'явилися пізніше, теж: ``_deals`` дивиться на факт договору, а не на
    # дати. Багатолотова закупівля (лот 1 з договором, лот 2 ще з рішенням)
    # інакше показувала б у кривій більше грошей, ніж бачить сам аналіз.
    mixed = OutlookMixin._visibility(published, [deal], [award_one, award_two])
    check("до договору — рішення", visible(mixed, 15) == 100.0, mixed)
    check("після договору — лише договір", visible(mixed, 19) == 90.0, mixed)
    check("пізніше рішення не додається", visible(mixed, 40) == 90.0, mixed)


def test_chart_edges() -> None:
    print("\n=== діаграми на вироджених даних ===")
    from PySide6.QtGui import QPixmap
    from PySide6.QtWidgets import QApplication

    from app.core.report import ChartData, Series
    from app.ui.widgets.charts import build

    QApplication.instance() or QApplication([])
    # `None` у ряду — це «даних немає» (див. Series), і лінія його вміла
    # завжди, а стовпчики, смуги й кільце падали ще на пошуку максимуму.
    # Нескінченність приходить не з рушія, а з книги: float("1e400") — це вже
    # inf, і Qt кидав на ньому OverflowError просто з paintEvent.
    cases = {
        "порожній ряд": [Series("Ряд", [], [])],
        "без рядів": [],
        "усі нулі": [Series("Ряд", ["A", "B"], [0.0, 0.0])],
        "від'ємні": [Series("Ряд", ["A", "B"], [-10.0, -20.0])],
        "розрив усередині": [Series("Ряд", ["A", "B", "C"], [1.0, None, 3.0])],
        "самі розриви": [Series("Ряд", ["A", "B"], [None, None])],
        "нескінченність": [Series("Ряд", ["A", "B"], [float("inf"), 1.0])],
        "NaN": [Series("Ряд", ["A", "B"], [float("nan"), 1.0])],
        "значень менше за мітки": [Series("Ряд", ["A", "B", "C"], [1.0])],
        "виділення поза межами": [Series("Ряд", ["A"], [1.0], accent={9})],
    }
    failed: list[str] = []
    for name, series in cases.items():
        for kind in ("bar", "hbar", "pie", "line", "area", "hist", "scatter"):
            try:
                view = build(ChartData(name, kind, series, unit="грн"), "dark")
                view.resize(400, 240)
                view.render(QPixmap(view.size()))
            except Exception as exc:
                failed.append(f"{kind}/{name}: {type(exc).__name__}")
    check("жодна діаграма не падає", not failed, failed[:5])

    # Розрив не має ставати нулем: лінія прогнозу саме на цьому й тримається.
    line = build(ChartData("розрив", "line",
                           [Series("Ряд", ["A", "B", "C"], [10.0, None, 30.0])]), "dark")
    line.resize(400, 240)
    line.render(QPixmap(line.size()))
    check("точок стільки, скільки відомих значень", len(line._hits) == 2, len(line._hits))


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    with tempfile.TemporaryDirectory(prefix="prozorro-test-") as raw:
        tmp = Path(raw)
        test_classifiers()
        test_safe_names()
        test_extract()
        test_index_shards()
        test_file_filter()
        test_batch_files()
        test_market()
        test_market_api()
        test_gone_cards(tmp)
        test_product_groups()
        test_two_product_tables(tmp)
        test_search_quota()
        test_search_plan(tmp)
        test_unnamed_procedure(tmp)
        test_buyer_split(tmp)
        test_search_progress(tmp)
        test_host_policies()
        test_output_dir(tmp)
        test_output_dir_inside_project(tmp)
        test_one_date()
        test_resolve_choice(tmp)
        test_summary_resolver(tmp)
        test_index_chunks()
        test_paths(tmp)
        test_download_missing(tmp)
        test_widgets()
        test_report_stats()
        test_brands()
        test_xlsxload(tmp)
        test_insight(tmp)
        test_ours_never_hidden(tmp)
        test_fast_reader(tmp)
        test_workbook_listing(tmp)
        test_tracked_competitors(tmp)
        test_excel_dialect(tmp)
        test_degenerate(tmp)
        test_degenerate_ui(tmp)
        test_reader_fallback(tmp)
        test_deal_sources(tmp)
        test_cancel_midway(tmp)
        test_export_names(tmp)
        test_report_overview(tmp)
        test_report_book(tmp)
        test_profile_book(tmp)
        test_profile_export_button(tmp)
        test_number_formats()
        test_loss_reasons(tmp)
        test_head_to_head(tmp)
        test_rejection_reasons()
        test_rejection_extract(tmp)
        test_rejection_export(tmp)
        test_rejection_section(tmp)
        test_rejection_in_profiles(tmp)
        test_rejection_double_count(tmp)
        test_excel_serial_time()
        test_progress_scale(tmp)
        test_impossible_dates(tmp)
        test_visibility_matches_deals(tmp)
        test_chart_edges()
        test_column_matching()
        test_settings_robustness(tmp)
        test_own_tm_verdict(tmp)
        test_big_deals_survive(tmp)
        test_multilot_not_penalised(tmp)
        test_savings_by_competition(tmp)
        test_monthly_one_axis(tmp)
        test_foreign_currency_reported(tmp)
        test_fresh_start(tmp)
        test_urls_only_when_downloading()
        test_analytics_is_file_only(tmp)
        test_scale(tmp)
        test_charts()
        test_forecast_models()
        test_forecast_choice_risk()
        test_forecast_seasonality()
        test_forecast_completeness(tmp)
    if FAILED:
        print(f"\nНЕ ПРОЙШЛО {len(FAILED)}: {', '.join(FAILED)}")
        return 1
    print("\nУСЕ ГАРАЗД")
    return 0


if __name__ == "__main__":
    sys.exit(main())
